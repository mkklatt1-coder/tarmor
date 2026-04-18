from django.shortcuts import render, redirect
from .forms import MeterUploadForm
from django.forms import inlineformset_factory
from django.http import HttpResponse
from .models import MeterReading, cascade_meter_update, Meter
from equipment.models import Equipment
from equipment.models import Equipment, Meter
from .forms import MeterUploadForm
from django.db import IntegrityError
import openpyxl, csv, io
from django.contrib import messages

def meters(request):
    return render(request, 'meters/meters.html', {'content_type': 'text/html'})

def new_reading(request):
    query = request.GET.get('new_reading')
    unit = None
    formset = None

    if query:
        unit = Equipment.objects.filter(Equipment_Number=query).first()

        if unit:
            MeterFormSet = inlineformset_factory(
                Equipment, 
                MeterReading, 
                form=MeterUploadForm, 
                extra=unit.meters.count(),
                can_delete=False
            )
            if request.method == "POST":
                formset = MeterFormSet(request.POST, instance=unit)
                if formset.is_valid():
                    try:
                        readings = formset.save(commit=False)
                        unit_meters = unit.meters.all()
                        master_date = formset.forms[0].cleaned_data.get('Date')
                        
                        for i, (reading, form) in enumerate(zip(readings, formset.forms)):
                            diff = form.cleaned_data.get('reading_diff')
                            current_val = form.cleaned_data.get('Meter_Reading')
                            is_replaced = form.cleaned_data.get('Meter_Replaced')
                            
                            reading.Date = master_date
                            reading.Equipment = unit
                            reading.Meter_Type = unit_meters[i]
                            
                            last_log = MeterReading.objects.filter(Meter_Type=reading.Meter_Type).order_by('-Date', '-id').first()
                            if last_log:
                                last_total = last_log.Total_Meter_Value
                                last_physical_reading = last_log.Meter_Reading 
                            else:
                                last_total = 0
                                last_physical_reading = 0
                            if is_replaced == 'Yes':
                                reading.Reading_Difference = 0
                                reading.Total_Meter_Value = last_total
                                reading.Meter_Reading = current_val or 0
                            elif diff:
                                reading.Reading_Difference = diff
                                reading.Total_Meter_Value = last_total + diff
                                reading.Meter_Reading = last_physical_reading + diff
                            else:
                                actual_diff = (current_val or 0) - last_physical_reading
                                reading.Reading_Difference = actual_diff
                                reading.Total_Meter_Value = last_total + actual_diff
                                reading.Meter_Reading = current_val
                            
                            reading.save()
                        return redirect('meters:meters')
                    
                    except IntegrityError:
                        return render(request, 'meters/new_reading.html', {
                            'formset': formset, 
                            'unit': unit,
                            'error_message': f"Error: A reading for {unit} on {master_date} already exists."
                        })
            else:
                formset = MeterFormSet(instance=unit, queryset=MeterReading.objects.none())
                for form, mtr in zip(formset.forms, unit.meters.all()):
                    form.meter_label = mtr.meter_type
        else:
            "Please select equipment number to begin"
            pass
    return render(request, 'meters/new_reading.html', {'formset': formset, 'unit': unit})

def edit_reading(request):
    query_unit = request.GET.get('unit_search')
    query_date = request.GET.get('date_search')
    unit = Equipment.objects.filter(Equipment_Number=query_unit).first() if query_unit else None
    formset = None
        
    if unit:
        existing_readings = MeterReading.objects.filter(Equipment=unit, Date=query_date)
        
        MeterFormSet = inlineformset_factory(
            Equipment, 
            MeterReading, 
            form=MeterUploadForm, 
            extra=0, 
            can_delete=True 
        )

        if request.method == "POST":
            formset = MeterFormSet(request.POST, instance=unit, queryset=existing_readings)
            if formset.is_valid():
                affected = []
                
                for form in formset.deleted_forms:
                    if form.instance.pk:
                        affected.append((form.instance.Meter_Type, form.instance.Date))
                        form.instance.delete()
                    
                for form in formset.forms:
                    if not form.cleaned_data.get('DELETE'):
                        reading = form.save(commit=False)
                        
                        reading.Equipment = unit
                        
                        last_log = MeterReading.objects.filter(Meter_Type=reading.Meter_Type, Date__lt=reading.Date).order_by('-Date', '-id').first()
                        last_total = last_log.Total_Meter_Value if last_log else 0
                        last_phys = last_log.Meter_Reading if last_log else 0
                        diff = form.cleaned_data.get('reading_diff')
                        curr = form.cleaned_data.get('Meter_Reading')
                
                        if reading.Meter_Replaced == 'Yes':
                            reading.Reading_Difference = 0
                            reading.Total_Meter_Value = last_total
                        elif diff:
                            reading.Reading_Difference = diff
                            reading.Total_Meter_Value = last_total + diff
                            reading.Meter_Reading = last_phys + diff
                        else:
                            diff_calc = (curr or 0) - last_phys
                            reading.Reading_Difference = diff_calc
                            reading.Total_Meter_Value = last_total + diff_calc
                            reading.Meter_Reading = curr

                        reading.save()
                        affected.append((reading.Meter_Type, reading.Date))
                        
                for mtr_type, mtr_date in set(affected):
                    cascade_meter_update(mtr_type, mtr_date)
            
                return redirect('meters:meters')
            else:
                print("VALIDATION FAILED:", formset.errors)
        else:
            formset = MeterFormSet(instance=unit, queryset=existing_readings)
            for form in formset.forms:
                if form.instance.pk:
                    form.meter_label = form.instance.Meter_Type.meter_type

    return render(request, 'meters/edit_reading.html', {'formset': formset, 'unit': unit})

def search_readings(request):
    # Start with all records
    readings = MeterReading.objects.all().select_related('Equipment', 'Meter_Type')

    # Get Filter Parameters
    eq_num = request.GET.get('Equipment_Number')
    asset_type = request.GET.get('Asset_Type')
    eq_type = request.GET.get('Equipment_Type')
    status = request.GET.get('Equipment_Status')
    make = request.GET.get('Make')
    model = request.GET.get('Model')
    sort_by = request.GET.get('sort', '-Date') # Default to newest first

    # Apply Filters (icontains makes it case-insensitive)
    if eq_num:
        readings = readings.filter(Equipment__Equipment_Number__icontains=eq_num)
    if asset_type:
        readings = readings.filter(Equipment__Asset_Type__name__icontains=asset_type)
    if eq_type:
        readings = readings.filter(Equipment__Equipment_Type__Equipment_Type__icontains=eq_type)
    if status:
        readings = readings.filter(Equipment__Equipment_Status__icontains=status)
    if make:
        readings = readings.filter(Equipment__Make__icontains=make)
    if model:
        readings = readings.filter(Equipment__Model__icontains=model)

    # Sorting Logic
    readings = readings.order_by(sort_by)

    return render(request, 'meters/search_readings.html', {
        'equipment_list': readings,
        'sort_by': sort_by,
    })
    
def export_readings_excel(request):
    readings = MeterReading.objects.all().select_related('Equipment', 'Meter_Type')
    
    eq_num = request.GET.get('Equipment_Number')
    asset_type = request.GET.get('Asset_Type')
    eq_type = request.GET.get('Equipment_Type')
    status = request.GET.get('Equipment_Status')
    make = request.GET.get('Make')
    model = request.GET.get('Model')
    
    if eq_num:
        readings = readings.filter(Equipment__Equipment_Number__icontains=eq_num)
        if asset_type:
            readings = readings.filter(Equipment__Asset_Type__name__icontains=asset_type)
        if eq_type:
            readings = readings.filter(Equipment__Equipment_Type__Equipment_Type__icontains=eq_type)
        if status:
            readings = readings.filter(Equipment__Equipment_Status__icontains=status)
        if make:
            readings = readings.filter(Equipment__Make__icontains=make)
        if model:
            readings = readings.filter(Equipment__Model__icontains=model)

    # 2. Create the Excel Workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Meter_Readings_Export.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Meter Readings"

    # 3. Add Headers
    headers = ['Equipment Number', 'Date', 'Meter Type', 'Reading', 'Difference', 'Total Value', 'Replaced']
    ws.append(headers)

    # 4. Add Data Rows
    for item in readings:
        ws.append([
            item.Equipment.Equipment_Number,
            item.Date.strftime('%Y-%m-%d') if item.Date else '',
            item.Meter_Type.meter_type,
            item.Meter_Reading,
            item.Reading_Difference,
            item.Total_Meter_Value,
            item.Meter_Replaced
        ])

    wb.save(response)
    return response

def mass_upload_readings(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        uploaded_file = request.FILES['excel_file']
        filename = uploaded_file.name.lower()
        rows = []
        
        try:
            if filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(min_row=2, values_only=True))

            elif filename.endswith('.csv'):
                content = uploaded_file.read().decode('utf-8-sig') 
                csv_reader = csv.reader(io.StringIO(content))
                next(csv_reader)
                rows = list(csv_reader)

            else:
                messages.error(request, "Please upload a .csv or .xlsx file.")
                return redirect('meters:mass_upload')

            for row in rows:
                if not any(row): continue
                
                eq_num, date_val, mtr_type_name, curr_val = row[:4]
                
                if isinstance(curr_val, str):
                    curr_val = float(curr_val) if curr_val.strip() else 0

                unit = Equipment.objects.filter(Equipment_Number=eq_num).first()
                if not unit:
                    continue

                mtr_type = Meter.objects.filter(equipment=unit, meter_type=mtr_type_name).first()
                if not mtr_type:
                    continue

                last_log = MeterReading.objects.filter(
                    Meter_Type=mtr_type, 
                    Date__lt=date_val
                ).order_by('-Date', '-id').first()
                
                last_total = last_log.Total_Meter_Value if last_log else 0
                last_phys = last_log.Meter_Reading if last_log else 0

                reading = MeterReading(
                    Date=date_val,
                    Equipment=unit,
                    Meter_Type=mtr_type,
                    Meter_Reading=curr_val,
                    Meter_Replaced='No'
                )

                actual_diff = (curr_val or 0) - last_phys
                reading.Reading_Difference = actual_diff
                reading.Total_Meter_Value = last_total + actual_diff

                reading.save()
                
                cascade_meter_update(mtr_type, date_val)

            messages.success(request, "Mass upload completed successfully!")
            return redirect('meters:meters')

        except Exception as e:
            messages.error(request, f"Error processing file: {e}")

    return render(request, 'meters/mass_upload.html')

def download_meter_template(request):
   
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="meter_upload_template.csv"'

    writer = csv.writer(response)
    writer.writerow(['Eq_Number', 'Date', 'Meter_Type', 'Current_Reading'])

    meters = Meter.objects.all().select_related('equipment')
    for m in meters:
        writer.writerow([
            m.equipment.Equipment_Number, 
            "", 
            m.meter_type, 
            ""  
        ])

    return response