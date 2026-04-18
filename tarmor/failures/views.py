from django.shortcuts import render, redirect, get_object_or_404
from .forms import SystemFormSet, SystemEditFormSet, SystemForm, modelformset_factory, ComponentEditForm, FailureTypeForm, FailureTypeEditForm, ActionForm, ActionEditForm
from .models import System, Component, FailureType, Action
from django.contrib import messages
import csv, io
from django.db import transaction
from django.db.models import Q
from urllib.parse import urlencode
import openpyxl
from django.http import HttpResponse, JsonResponse
from django import forms

def failure_codes(request):
    return render(request, 'failures/failure_codes.html')

def system(request):
    return render(request, 'failures/system.html')

def components(request):
    return render(request, 'failures/components.html')

def failure_modes(request):
    return render(request, 'failures/failure_modes.html')

def actions(request):
    return render(request, 'failures/actions.html')

def create_systems(request):
    if request.method == "POST":
        formset = SystemFormSet(request.POST, queryset=System.objects.none())
        if formset.is_valid():
            formset.save()
            messages.success(request, "System(s) added successfully!")
            return redirect("failures:system")
    else:
        formset = SystemFormSet(queryset=System.objects.none())
    
    return render(request, "failures/system_list.html", {
        "formset": formset,
    })
    
def edit_system(request):
    all_systems = System.objects.all()
    query = request.GET.get('system_name', '')
    
    if query:
        queryset = System.objects.filter(system_name__icontains=query)
    else:
        queryset = System.objects.none()
    
    if request.method == 'POST':
        formset = SystemEditFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, "System updated successfully!")
            return redirect('failures:system')
        else:
            print(formset.errors)
    else:
        formset = SystemEditFormSet(queryset=queryset)

    return render(request, 'failures/edit_system.html', {
        'formset': formset,
        'all_systems': all_systems,
        'system_name': query,
    })
    
def delete_system(request, pk):
    system = get_object_or_404(System, pk=pk)
    name = system.system_name
    system.delete()
    messages.warning(request, f"System '{name}' has been deleted.")
    return redirect('failures:system')

def upload_systems_csv(request):
    if request.method == "POST":
        csv_file = request.FILES.get('csv_file')
        
        if not csv_file or not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a valid CSV file.')
            return redirect("failures:system_list")

        data_set = csv_file.read().decode('utf-8-sig')
        io_string = io.StringIO(data_set)
        reader = csv.reader(io_string, delimiter=',', quotechar='"')
        
        errors = []
        seen_in_file = set()
        to_save = []
        
        try:
            with transaction.atomic():
                for line_num, row in enumerate(reader, start=1):
                    if len(row) < 3:
                        errors.append(f"Row {line_num}: Invalid format (expected 3 columns).")
                        continue

                    data = {
                        'asset_key': row[0].strip().upper(),
                        'system_name': row[1].strip(),
                        'system_key': row[2].strip().upper()
                    }

                    combined = f"{data['asset_key']}{data['system_key']}"
                    
                    if combined in seen_in_file:
                        errors.append(f"Row {line_num}: Duplicate key in file: {combined}")
                        continue
                    seen_in_file.add(combined)

                    form = SystemForm(data)
                    if form.is_valid():
                        to_save.append(System(**form.cleaned_data))
                    else:
                        for field, error_list in form.errors.items():
                            for error in error_list:
                                if "already exists" in error:
                                    errors.append(f"Row {line_num}: Key already exists in database: {combined}")
                                else:
                                    errors.append(f"Row {line_num}: {field} - {error}")

                if errors:
                    raise Exception("Validation Failed")
                else:
                    System.objects.bulk_create(to_save)
                    messages.success(request, f"Successfully uploaded {len(to_save)} systems.")

        except Exception:
            for error in errors:
                messages.error(request, error)
            return redirect("failures:system_list")

    return redirect("failures:system_list")

def search_system(request):
    asset_key = request.GET.get('asset_key', '')
    system_name = request.GET.get('system_name', '')
    system_key = request.GET.get('system_key', '')
    combined_sys_key = request.GET.get('combined_sys_key', '')
    sort_by = request.GET.get('sort', 'system_name')

    queryset = System.objects.all()
    
    if asset_key:
        queryset = queryset.filter(asset_key__icontains=asset_key)
    if system_name:
        queryset = queryset.filter(system_name__icontains=system_name)
    if system_key:
        queryset = queryset.filter(system_key__icontains=system_key)
    if combined_sys_key:
        queryset = queryset.filter(combined_sys_key__icontains=combined_sys_key)

    queryset = queryset.order_by(sort_by)

    filters = {
        'asset_key': asset_key,
        'system_name': system_name,
        'system_key': system_key,
        'combined_sys_key': combined_sys_key
    }
    
    filter_url = urlencode({k: v for k, v in filters.items() if v})

    context = {
        'system': queryset,
        'asset_key': asset_key,
        'system_name': system_name,
        'system_key': system_key,
        'combined_sys_key': combined_sys_key,
        'sort_by': sort_by,
        'filter_url': filter_url,
    }
    return render(request, 'failures/search_system.html', context)

def export_systems_excel(request):
    asset_key = request.GET.get('asset_key', '')
    system_name = request.GET.get('system_name', '')
    system_key = request.GET.get('system_key', '')
    combined_sys_key = request.GET.get('combined_sys_key', '')

    queryset = System.objects.all()
    if asset_key: queryset = queryset.filter(asset_key__icontains=asset_key)
    if system_name: queryset = queryset.filter(system_name__icontains=system_name)
    if system_key: queryset = queryset.filter(system_key__icontains=system_key)
    if combined_sys_key: queryset = queryset.filter(combined_sys_key__icontains=combined_sys_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Systems"

    headers = ['Asset Key', 'System Name', 'System Key', 'Combined System Key']
    ws.append(headers)

    for obj in queryset:
        ws.append([obj.asset_key, obj.system_name, obj.system_key, obj.combined_sys_key])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Systems_Export.xlsx'
    wb.save(response)
    return response

def add_component(request):
    if request.method == "POST":
        suffix = request.POST.get('component_key', '').upper()
        comp_name = request.POST.get('component_name', '')
        selected_system_ids = request.POST.getlist('selected_systems')
        
        if len(suffix) != 4:
            messages.error(request, "Component key must be exactly 4 characters.")
            return redirect('add_component')

        new_components = []
        errors = []

        for system_id in selected_system_ids:
            try:
                system = System.objects.get(id=system_id)
                full_code = f"{system.combined_sys_key}{suffix}"

                if Component.objects.filter(combined_comp_key=full_code).exists():
                    errors.append(f"The code {full_code} already exists in the database.")
                else:
                    new_components.append(Component(
                        component_name=comp_name,
                        component_key=suffix,
                        combined_sys_key_id=system.id, # Using the ID for the foreign key
                        combined_comp_key=full_code
                    ))
            except System.DoesNotExist:
                continue
                    
        if errors:
            for err in errors:
                messages.error(request, err)
        elif new_components:
            Component.objects.bulk_create(new_components)
            messages.success(request, f"Successfully created {len(new_components)} components.")
            return redirect('failures:components')
        else:
            messages.warning(request, "No systems were selected.")

    systems = System.objects.all()
    return render(request, 'failures/add_component.html', {'systems': systems})

def check_uniqueness(request):
    suffix = request.GET.get('suffix', '').upper()
    asset_keys = request.GET.getlist('asset_keys[]')
    
    codes_to_check = [f"{key}{suffix}" for key in asset_keys]
    
    existing_codes = list(Component.objects.filter(combined_comp_key__in=codes_to_check)
                          .values_list('combined_comp_key', flat=True))
    
    return JsonResponse({'existing_codes': existing_codes})

def mass_upload_components(request):
    if request.method == "POST" and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        
        seen_in_csv = set()
        to_create = []
        errors = []

        try:
            with transaction.atomic():
                reader = csv.reader(io_string)
                for row in reader:
                    if not row or len(row) < 4: continue 
                    
                    name, combined_code, sys_key, comp_key = [item.strip().upper() for item in row]

                    if combined_code in seen_in_csv:
                        errors.append(f"CSV Duplicate: {combined_code} appears twice in file.")
                    
                    elif Component.objects.filter(combined_comp_key=combined_code).exists():
                        errors.append(f"Database Conflict: {combined_code} already exists.")
                    
                    else:
                        try:
                            system = System.objects.get(combined_sys_key=sys_key)
                            seen_in_csv.add(combined_code)
                            to_create.append(Component(
                                component_name=name,
                                combined_comp_key=combined_code,
                                combined_sys_key=system,
                                component_key=comp_key
                            ))
                        except System.DoesNotExist:
                            errors.append(f"System Error: '{sys_key}' not found.")

                if errors:
                    raise Exception("Validation Failed")
                
                Component.objects.bulk_create(to_create)
                messages.success(request, f"Successfully uploaded {len(to_create)} components.")
                return redirect('failures:components')
            
        except Exception:
            for err in errors[:5]:
                messages.error(request, err)
            messages.error(request, "Upload aborted. No data was saved.")
            
    return redirect('failures:add_component')

def edit_component(request):
    component_name = request.GET.get('component_name', '')
    queryset = Component.objects.filter(component_name=component_name) if component_name else Component.objects.none()
    
    ComponentFormSet = modelformset_factory(
        Component, 
        form=ComponentEditForm,
        extra=0,
    )

    if request.method == "POST":
        formset = ComponentFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.combined_comp_key = f"{instance.combined_sys_key.combined_sys_key}{instance.component_key}"
                instance.save()
                messages.success(request, "Component edited successfully!")
            return redirect('failures:components')
        else:
            print(formset.errors) 
    else:
        formset = ComponentFormSet(queryset=queryset)        

    all_components = Component.objects.values('component_name').distinct()
    
    return render(request, 'failures/edit_component.html', {
        'formset': formset,
        'component_name': component_name,
        'all_components': all_components,
    })

def delete_component(request, pk):
    component = get_object_or_404(Component, pk=pk)
    component.delete()
    return redirect('failures:components')

def search_components(request):
    combined_comp_key = request.GET.get('combined_comp_key', '')
    component_name = request.GET.get('component_name', '')
    component_key = request.GET.get('component_key', '')
    combined_sys_key = request.GET.get('combined_sys_key', '')
    sort_by = request.GET.get('sort', 'component_name')

    queryset = Component.objects.all()
    if combined_comp_key: queryset = queryset.filter(combined_comp_key__icontains=combined_comp_key)
    if component_name: queryset = queryset.filter(component_name__icontains=component_name)
    if component_key: queryset = queryset.filter(component_key__icontains=component_key)
    if combined_sys_key: queryset = queryset.filter(combined_sys_key__combined_sys_key__icontains=combined_sys_key)

    queryset = queryset.order_by(sort_by)

    filters = {
        'combined_comp_key': combined_comp_key,
        'component_name': component_name,
        'component_key': component_key,
        'combined_sys_key': combined_sys_key
    }
    
    filter_url = urlencode({k: v for k, v in filters.items() if v})

    context = {
        'component': queryset,
        'combined_comp_key': combined_comp_key,
        'component_name': component_name,
        'component_key': component_key,
        'combined_sys_key': combined_sys_key,
        'sort_by': sort_by,
        'filter_url': filter_url,
    }
    return render(request, 'failures/search_components.html', context)

def export_components_excel(request):
    combined_comp_key = request.GET.get('combined_comp_key', '')
    component_name = request.GET.get('component_name', '')
    component_key = request.GET.get('component_key', '')
    combined_sys_key = request.GET.get('combined_sys_key', '')

    queryset = Component.objects.all()
    if combined_comp_key: queryset = queryset.filter(combined_comp_key__icontains=combined_comp_key)
    if component_name: queryset = queryset.filter(component_name__icontains=component_name)
    if component_key: queryset = queryset.filter(component_key__icontains=component_key)
    if combined_sys_key: queryset = queryset.filter(combined_sys_key__combined_sys_key__icontains=combined_sys_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Components"

    headers = ['Combined Component Key', 'Component Name', 'Component Key', 'Combined System Key']
    ws.append(headers)

    for row_num, comp in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = str(comp.combined_comp_key)
        ws.cell(row=row_num, column=2).value = str(comp.component_name)
        ws.cell(row=row_num, column=3).value = str(comp.component_key)
        ws.cell(row=row_num, column=4).value = str(comp.combined_sys_key)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Components_Export.xlsx'
    wb.save(response)
    return response

def add_failure_mode(request):
    FailureFormSet = modelformset_factory(FailureType, form=FailureTypeForm, extra=1)
    
    if request.method == "POST":
        formset = FailureFormSet(request.POST)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Failure Mode(s) added successfully!")
            return redirect("failures:failure_modes")
    else:
        formset = FailureFormSet(queryset=FailureType.objects.none())
    
    return render(request, "failures/add_failure_mode.html", {
        "formset": formset,
    })
    
def mass_upload_fail_codes(request):
    if request.method == "POST" and request.FILES.get('csv_file'):
        file = request.FILES['csv_file']
        decoded_file = file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(decoded_file))
        
        seen_codes = set()
        to_create = []
        errors = []
        
        try:
            for row in reader:
                    if not row: continue
                    mode, code = [col.strip().upper() for col in row]

                    if code in seen_codes or FailureType.objects.filter(failure_code=code).exists():
                        errors.append(f"Duplicate code: {code}")

                    else:
                        seen_codes.add(code)
                        to_create.append(FailureType(failure_mode=mode, failure_code=code))


            if errors:
                raise Exception("Validation Failed")
            else:
                FailureType.objects.bulk_create(to_create)
                messages.success(request, f"Successfully uploaded {len(to_create)} failure modes.")
                return redirect("failures:failure_modes")

        except Exception:
            for error in errors:
                messages.error(request, error)
            return redirect("failures:add_failure_mode")

    return redirect("failures:add_failure_mode")

def edit_failure_mode(request):
    query = request.GET.get('failure_mode', '').strip()
    queryset = FailureType.objects.filter(failure_mode__icontains=query) if query else FailureType.objects.none()
    
    FailureTypeEditFormSet = modelformset_factory(
        FailureType, 
        form=FailureTypeEditForm, 
        extra=0
    )
    
    if request.method == 'POST':
        formset = FailureTypeEditFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Failure mode updated successfully!")
            return redirect('failures:failure_modes')
        else:
            print("Formset Errors:", formset.errors)
            print("Non-form Errors:", formset.non_form_errors())
    else:
        formset = FailureTypeEditFormSet(queryset=queryset)
        
    return render(request, 'failures/edit_failure_mode.html', {
        'formset': formset,
        'failure_mode': query,
    })
    
def delete_failure_mode(request, pk):
    failuretype = get_object_or_404(FailureType, pk=pk)
    name = failuretype.failure_mode
    failuretype.delete()
    messages.warning(request, f"Failure mode '{name}' has been deleted.")
    return redirect('failures:failure_modes')

def search_failure_modes(request):
    failure_mode = request.GET.get('failure_mode', '')
    failure_code = request.GET.get('failure_code', '')
    sort_by = request.GET.get('sort', 'failure_mode')

    queryset = FailureType.objects.all()
    if failure_mode: queryset = queryset.filter(failure_mode__icontains=failure_mode)
    if failure_code: queryset = queryset.filter(failure_code__icontains=failure_code)

    queryset = queryset.order_by(sort_by)

    filters = {
        'failure_mode': failure_mode,
        'failure_code': failure_code
    }
    
    filter_url = urlencode({k: v for k, v in filters.items() if v})

    context = {
        'failure_modes': queryset,
        'combined_comp_key': failure_mode,
        'component_name': failure_code,
        'sort_by': sort_by,
        'filter_url': filter_url,
    }
    return render(request, 'failures/search_failure_modes.html', context)

def export_failure_modes_excel(request):
    failure_mode = request.GET.get('failure_mode', '')
    failure_code = request.GET.get('failure_code', '')

    queryset = FailureType.objects.all()
    if failure_mode: queryset = queryset.filter(failure_mode__icontains=failure_mode)
    if failure_code: queryset = queryset.filter(failure_code__icontains=failure_code)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failure Modes"

    headers = ['Failure Mode', 'Failure Code']
    ws.append(headers)

    for row_num, comp in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = str(comp.failure_mode)
        ws.cell(row=row_num, column=2).value = str(comp.failure_code)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Failure_Modes_Export.xlsx'
    wb.save(response)
    return response

def add_action(request):
    ActionFormSet = modelformset_factory(Action, form=ActionForm, extra=1)
    
    if request.method == "POST":
        formset = ActionFormSet(request.POST)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Action(s) added successfully!")
            return redirect("failures:actions")
    else:
        formset = ActionFormSet(queryset=Action.objects.none())
    
    return render(request, "failures/add_action.html", {
        "formset": formset,
    })
    
def mass_upload_actions(request):
    if request.method == "POST" and request.FILES.get('csv_file'):
        file = request.FILES['csv_file']
        decoded_file = file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(decoded_file))
        
        seen_codes = set()
        to_create = []
        errors = []
        
        try:
            for row in reader:
                    if not row: continue
                    name, code = [col.strip().upper() for col in row]

                    if code in seen_codes or Action.objects.filter(action_key=code).exists():
                        errors.append(f"Duplicate code: {code}")

                    else:
                        seen_codes.add(code)
                        to_create.append(Action(action_name=name, action_key=code))

            if errors:
                raise Exception("Validation Failed")
            else:
                Action.objects.bulk_create(to_create)
                messages.success(request, f"Successfully uploaded {len(to_create)} actions.")
                return redirect("failures:actions")

        except Exception:
            for error in errors:
                messages.error(request, error)
            return redirect("failures:add_action")

    return redirect("failures:add_action")

def edit_action(request):
    query = request.GET.get('action_name', '').strip()
    all_actions=Action.objects.all()
    queryset = Action.objects.filter(action_name__icontains=query) if query else Action.objects.none()
    
    ActionEditFormSet = modelformset_factory(
        Action, 
        form=ActionEditForm, 
        extra=0
    )
    
    if request.method == 'POST':
        formset = ActionEditFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Action updated successfully!")
            return redirect('failures:actions')
        else:
            print("Formset Errors:", formset.errors)
            print("Non-form Errors:", formset.non_form_errors())
    else:
        formset = ActionEditFormSet(queryset=queryset)
        
    return render(request, 'failures/edit_action.html', {
        'formset': formset,
        'action_name': query,
        'all_actions': all_actions,
    })
    
def delete_action(request, pk):
    actiontype = get_object_or_404(Action, pk=pk)
    name = actiontype.action_name
    actiontype.delete()
    messages.warning(request, f"Action '{name}' has been deleted.")
    return redirect('failures:actions')

def search_actions(request):
    action_name = request.GET.get('action_name', '')
    action_key = request.GET.get('action_key', '')
    sort_by = request.GET.get('sort', 'action_name')

    queryset = Action.objects.all()
    if action_name: queryset = queryset.filter(action_name__icontains=action_name)
    if action_key: queryset = queryset.filter(action_key__icontains=action_key)

    queryset = queryset.order_by(sort_by)

    filters = {
        'action_name': action_name,
        'action_key': action_key
    }
    
    filter_url = urlencode({k: v for k, v in filters.items() if v})

    context = {
        'actions': queryset,
        'action_name': action_name,
        'action_key': action_key,
        'sort_by': sort_by,
        'filter_url': filter_url,
    }
    return render(request, 'failures/search_actions.html', context)

def export_actions_excel(request):
    action_name = request.GET.get('action_name', '')
    action_key = request.GET.get('action_key', '')

    queryset = Action.objects.all()
    if action_name: queryset = queryset.filter(action_name__icontains=action_name)
    if action_key: queryset = queryset.filter(action_key__icontains=action_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actions"

    headers = ['Action Name', 'Action Key']
    ws.append(headers)

    for row_num, act in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = str(act.action_name)
        ws.cell(row=row_num, column=2).value = str(act.action_key)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Actions_Export.xlsx'
    wb.save(response)
    return response