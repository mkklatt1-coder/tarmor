from django.shortcuts import render
from django.views.generic import ListView
from django.db.models import Count, Q, Sum, Count, Avg, F, ExpressionWrapper, fields, Max, Min, FloatField
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from dateutil.rrule import rrule, MONTHLY
from timesheets.models import Timesheet
from .models import FailureFrequency, MTBF
from failures.models import System, Component, FailureType, Action
from equipment.models import AssetType, EQ_Type, Equipment, MachineShiftStatus, ShiftReport
from scheduling.models import WorkWeek, ScheduleSnapshot, Schedule
from work_orders.models import WorkOrder
from timesheets.models import Timesheet
from purchasing.models import Purchase
from meters.models import MeterReading
import openpyxl
import calendar
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from decimal import Decimal

def kpis(request):
    return render(request, 'kpis/kpis.html')

def get_active_week_bounds(request):
    """
    Helper utility function to parse the uniform WorkWeek selection string
    and extract the precise database start and end date variables.
    """
    week_param = request.GET.get('week', '').strip()
    
    clean_week_id = week_param.split(' ').strip() if 'Week ' in week_param else week_param
    
    if clean_week_id:
        selected_week = WorkWeek.objects.filter(id=clean_week_id).first()
    else:
        today = timezone.now().date()
        selected_week = WorkWeek.objects.filter(start_date__lte=today, end_date__gte=today).first()
        
    if not selected_week:
        selected_week = WorkWeek.objects.all().order_by('-week_number').first()

    if selected_week:
        start_date = selected_week.start_date.strftime("%Y-%m-%d")
        end_date = selected_week.end_date.strftime("%Y-%m-%d")
        week_val_string = f"Week {selected_week.id}"
    else:
        today_dt = datetime.now()
        start_date = today_dt.replace(month=1, day=1).strftime("%Y-%m-%d")
        end_date = today_dt.strftime("%Y-%m-%d")
        week_val_string = ""

    return start_date, end_date, week_val_string

def top_failures_report(request):
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    asset_type = request.GET.get('asset_type', '').strip()
    eq_num = request.GET.get('equipment_number', '').strip()
    sys = request.GET.get('fc_system', '').strip()
    comp = request.GET.get('fc_component', '').strip()
    fail = request.GET.get('fc_failure_mode', '').strip()
    action = request.GET.get('fc_action', '').strip()
    
    clean_eq_num = eq_num.split(' - ')[0].strip() if ' - ' in eq_num else eq_num

    if not start_date or not end_date or start_date in ['None', ''] or end_date in ['None', '']:
        today = datetime.now()
        start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")  # Jan 1st
        end_date = today.strftime("%Y-%m-%d")

    s_dt = datetime.strptime(start_date, "%Y-%m-%d")
    e_dt = datetime.strptime(end_date, "%Y-%m-%d")

    qs = WorkOrder.objects.select_related('equipment', 'equipment__Asset_Type', 'equipment__Equipment_Type', 'fc_system', 'fc_component', 'fc_failure_mode').filter(job_status__status_choice="Complete")
    
    if start_date and end_date:
        qs = qs.filter(date_closed__range=[s_dt, e_dt])
    if asset_type:
        qs = qs.filter(equipment__Asset_Type__name__icontains=asset_type)
    if clean_eq_num:
        qs = qs.filter(equipment__Equipment_Number__icontains=clean_eq_num)
    if sys:
        qs = qs.filter(fc_system__system_name__icontains=sys) 
    if comp:
        qs = qs.filter(fc_component__component_name__icontains=comp)
    if fail:
        qs = qs.filter(fc_failure_mode__failure_mode__icontains=fail)
    if action:
        qs = qs.filter(fc_action__action__icontains=fail)

    sort_by = request.GET.get('sort', '-date_closed')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_mapping = {
        'work_order': 'work_order',
        'repair_description': 'repair_description',
        'asset_type': 'equipment__Asset_Type__name',
        'equipment_number': 'equipment__Equipment_Number',
        'equipment_description': 'equipment__Equipment_Description',
        'fc_system': 'fc_system__system_name',
        'fc_component': 'fc_component__component_name',
        'fc_failure_mode': 'fc_failure_mode__failure_mode',
        'date_closed': 'date_closed',
    }

    if clean_sort_key in sort_mapping:
        db_field = sort_mapping[clean_sort_key]
        work_orders_sorted = qs.order_by(f"-{db_field}" if is_descending else db_field)
    else:
        work_orders_sorted = qs.order_by('-date_closed')

    work_orders_limited = work_orders_sorted[:5]

    systems = qs.exclude(fc_system__isnull=True).values('fc_system').annotate(total=Count('id')).order_by('-total')[:5]
    component = qs.exclude(fc_component__isnull=True).values('fc_component').annotate(total=Count('id')).order_by('-total')[:5]
    mode = qs.exclude(fc_failure_mode__isnull=True).values('fc_failure_mode').annotate(total=Count('id')).order_by('-total')[:5]
    action = qs.exclude(fc_action__isnull=True).values('fc_action').annotate(total=Count('id')).order_by('-total')[:5]

    for item in systems:
        item['obj'] = System.objects.filter(pk=item['fc_system']).first()
    for item in component:
        item['obj'] = Component.objects.filter(pk=item['fc_component']).first()
    for item in mode:
        item['obj'] = FailureType.objects.filter(pk=item['fc_failure_mode']).first()
    for item in action:
        item['obj'] = Action.objects.filter(pk=item['fc_action']).first()

    damage_pop = qs.filter(fc_failure_mode__failure_mode="DAMAGE (POP)").count()
    damage_pmp = qs.filter(fc_failure_mode__failure_mode="DAMAGE (PMP)").count()
    damage_nature = qs.filter(fc_failure_mode__failure_mode="DAMAGE (NATURE)").count()

    all_asset_types = AssetType.objects.all().order_by('name')
    if asset_type:
        all_eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_type).distinct().order_by('Equipment_Type')
    else:
        all_eq_types = EQ_Type.objects.all().order_by('Equipment_Type')
        
    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()

    context = {
        'work_orders': work_orders_limited,
        'system_counts': systems,
        'component_counts': component,
        'mode_counts': mode,
        'action_counts': action,
        'damage_pop': damage_pop,
        'damage_pmp': damage_pmp,
        'damage_nature': damage_nature,
        'sort': sort_by,
        'filter_url': filter_url,
        'asset_types': all_asset_types,
        'equipment_types': all_eq_types,
        'all_equipment_suggestions': all_equipment_suggestions,
        'asset_type_val': asset_type,
        'eq_num_val': eq_num,
        'filters': {'start_date': start_date, 'end_date': end_date, 'fc_system': sys, 'fc_component': comp, 'fc_failure_mode': fail},
    }
    return render(request, 'kpis/top_failures.html', context)

def export_top_failures_excel(request):
    start = request.GET.get('start_date')
    end = request.GET.get('end_date')
    asset_type = request.GET.get('asset_type')
    equip_num = request.GET.get('equip_num')
    sys = request.GET.get('fc_system')
    comp = request.GET.get('fc_component')
    fail = request.GET.get('fc_failure_mode')
    action = request.GET.get('fc_action')

    qs = WorkOrder.objects.select_related('equipment').all()

    if start and end:
        qs = qs.filter(date_closed__range=[start, end])
    if asset_type:
        qs = qs.filter(equipment__Asset_Type__icontains=asset_type)
    if equip_num:
        qs = qs.filter(equipment__Equipment_Number__icontains=equip_num)
    if sys:
        qs = qs.filter(fc_system__description__icontains=sys) 
    if comp:
        qs = qs.filter(fc_component__description__icontains=comp)
    if fail:
        qs = qs.filter(fc_failure_mode__description__icontains=fail)
    if action:
        qs = qs.filter(fc_action__description__icontains=action)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Work Order Failures"

    headers = [
        'WO #', 'Repair Description', 'Asset Type', 'Equip #', 'Equip Description', 
        'System', 'Component', 'Failure Mode', 'Action'
    ]
    ws.append(headers)

    for wo in qs:
        asset_type = str(wo.equipment.Asset_Type) if wo.equipment and wo.equipment.Asset_Type else ''
        equip_num = str(wo.equipment.Equipment_Number) if wo.equipment else ''
        equip_desc = str(wo.equipment.Equipment_Description) if wo.equipment else ''
        
        ws.append([
            wo.work_order,
            wo.repair_description,
            asset_type,
            equip_num,
            equip_desc,
            str(wo.fc_system), 
            str(wo.fc_component),
            str(wo.fc_failure_mode),
            str(wo.fc_action)
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Failure_Report.xlsx"'
    wb.save(response)
    return response

def failure_frequency_report(request):
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    asset_type = request.GET.get('asset_type', '').strip()
    eq_type = request.GET.get('equipment_type', '').strip()
    eq_num = request.GET.get('equipment_number', '').strip()

    clean_eq_num = eq_num.split(' - ')[0].strip() if ' - ' in eq_num else eq_num

    if not start_date or not end_date or start_date in ['None', '', '“”'] or end_date in ['None', '', '“”']:
        today = datetime.now()
        start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")

    equipment_qs = Equipment.objects.all().select_related('Asset_Type', 'Equipment_Type')
    if asset_type:
        equipment_qs = equipment_qs.filter(Asset_Type__name__icontains=asset_type)
    if eq_type:
        equipment_qs = equipment_qs.filter(Equipment_Type__Equipment_Type__icontains=eq_type)
    if clean_eq_num:
        equipment_qs = equipment_qs.filter(Equipment_Number__icontains=clean_eq_num)

    report_data = []

    s_dt = datetime.strptime(start_date, "%Y-%m-%d")
    e_dt = datetime.strptime(end_date, "%Y-%m-%d")

    for eq in equipment_qs:
        failures = WorkOrder.objects.filter(
            equipment=eq,
            work_type__work_type__in=["CF", "CP", "WTY"],
            date_created__range=[s_dt, e_dt]
        ).count()

        readings = MeterReading.objects.filter(
            Equipment=eq, 
            Date__range=[s_dt.date(), e_dt.date()]
        ).order_by('Date')
        
        accumulated_hours = 0
        if readings.exists():
            first_val = readings.first().Total_Meter_Value or 0
            last_val = readings.last().Total_Meter_Value or 0
            accumulated_hours = max(last_val - first_val, 0)
            
        frequency = round(failures / float(accumulated_hours), 4) if accumulated_hours > 0 else 0.0

            
        report_data.append({
            'equipment_number': eq.Equipment_Number,
            'equipment_description': eq.Equipment_Description or '',
            'asset_type': eq.Asset_Type.name if eq.Asset_Type else '',
            'equipment_type': eq.Equipment_Type.Equipment_Type if eq.Equipment_Type else '',
            'equipment_hours': accumulated_hours,
            'failure_count': failures,
            'frequency': frequency,
        })
            
    sort_by = request.GET.get('sort', 'equipment_number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    valid_sort_keys = ['equipment_number', 'equipment_description', 'asset_type', 'equipment_type', 'equipment_hours', 'failure_count', 'frequency']
    if clean_sort_key in valid_sort_keys:
        report_data.sort(
            key=lambda x: (x[clean_sort_key] is None, x[clean_sort_key]),
            reverse=is_descending
        )
        
    monthly_data_qs = (
        WorkOrder.objects.filter(
            equipment__in=equipment_qs,
            date_created__range=[s_dt, e_dt],
            work_type__work_type__in=["CF", "CP", "WTY"]
        )
        .annotate(month=TruncMonth('date_created'))
        .values('month')
        .annotate(count=Count('id'))
    )

    failures_lookup = {entry['month'].date() if hasattr(entry['month'], 'date') else entry['month']: entry['count'] for entry in monthly_data_qs}
    
    labels, bar_data, line_data = [], [], []
    total_failures_cumulative = 0
    months_processed = 0

    for dt in rrule(MONTHLY, dtstart=s_dt, until=e_dt):
        labels.append(dt.strftime('%b'))
        m_count = failures_lookup.get(dt.date(), 0)
        bar_data.append(m_count)

        total_failures_cumulative += m_count
        months_processed += 1
        line_data.append(round(total_failures_cumulative / months_processed, 2))

    # Fetch dynamic autocomplete option datasets
    all_asset_types = AssetType.objects.all().order_by('name')
    if asset_type:
        all_eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_type).distinct().order_by('Equipment_Type')
    else:
        all_eq_types = EQ_Type.objects.all().order_by('Equipment_Type')
        
    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()
    
    context = {
        'report_data': report_data,
        'labels': labels,
        'bar_data': bar_data,
        'line_data': line_data,
        'asset_types': all_asset_types,
        'equipment_types': all_eq_types,
        'all_equipment_suggestions': all_equipment_suggestions,
        'filters': {'start_date': start_date, 'end_date': end_date},
        'sort': sort_by,
        'filter_url': filter_url,
        'asset_type_val': asset_type,
        'eq_type_val': eq_type,
        'eq_num_val': eq_num,
    }
    return render(request, 'kpis/failure_frequency.html', context)

def get_freq_linked_eq_types(request):
    asset_name = request.GET.get('asset_type', '').strip()
    if asset_name:
        eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_name).values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')
    else:
        eq_types = EQ_Type.objects.values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')

    html_output = '<datalist id="eq-type-list">'
    for et in eq_types:
        html_output += f'<option value="{et}"></option>'
    html_output += '</datalist>'
    return HttpResponse(html_output, content_type="text/html")

def failure_frequency_chart(request):
    year = datetime.now().year
    
    monthly_data = (
        WorkOrder.objects.filter(
            date_created__year=year,
            work_type__work_type__in=["CF", "CP", "WTY"]
        )
        .annotate(month=TruncMonth('date_created'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    labels = []
    bar_data = []
    line_data = []
    
    total_failures = 0
    months_processed = 0

    for entry in monthly_data:
        month_name = entry['month'].strftime('%b')
        count = entry['count']
        
        labels.append(month_name)
        bar_data.append(count)
        
        total_failures += count
        months_processed += 1
        line_data.append(round(total_failures / months_processed, 2))

    context = {
        'labels': labels,
        'bar_data': bar_data,
        'line_data': line_data,
        'current_year': year,
    }
    return render(request, 'kpis/failure_frequency.html', context)

def export_failure_frequency_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    asset_type = request.GET.get('asset_type', '')
    eq_type = request.GET.get('equipment_type', '')
    eq_num = request.GET.get('equipment_number', '')

    equipment_qs = Equipment.objects.all()
    if asset_type:
        equipment_qs = equipment_qs.filter(Asset_Type__name__icontains=asset_type)
    if eq_type:
        equipment_qs = equipment_qs.filter(Equipment_Type__Equipment_Type__icontains=eq_type)
    if eq_num:
        equipment_qs = equipment_qs.filter(Equipment_Number__icontains=eq_num)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failure Frequency Report"

    headers = [
        'Equip #', 'Equip Description', 'Asset Type', 'Equipment Type', 
        'Hours', 'Failures', 'Frequency'
    ]
    ws.append(headers)

    for eq in equipment_qs:
        accumulated_hours = 0
        failures = 0
        
        if start_date and end_date:
            # Re-run failure logic
            failures = WorkOrder.objects.filter(
                equipment=eq,
                work_type__work_type__in=["CF", "CP", "WTY"],
                date_created__range=[start_date, end_date]
            ).count()

            # Re-run meter logic
            readings = MeterReading.objects.filter(
                Equipment=eq, 
                Date__range=[start_date, end_date]
            ).order_by('Date')
            if readings.exists():
                accumulated_hours = readings.last().Total_Meter_Value - readings.first().Total_Meter_Value

        freq = round(failures / accumulated_hours, 4) if accumulated_hours > 0 else 0

        
        ws.append([
            eq.Equipment_Number,
            eq.Equipment_Description,
            str(eq.Asset_Type),
            str(eq.Equipment_Type),
            accumulated_hours,
            failures,
            freq
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Failure_Frequency.xlsx"'
    
    wb.save(response)
    return response

def mtbf_report(request):
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    asset_type = request.GET.get('asset_type', '').strip()
    eq_type = request.GET.get('equipment_type', '').strip()
    eq_num = request.GET.get('equipment_number', '').strip()

    clean_eq_num = eq_num.split(' - ').strip() if ' - ' in eq_num else eq_num
    
    if not start_date or not end_date or start_date in ['None', ''] or end_date in ['None', '']:
        today = datetime.now()
        start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")  # Jan 1st
        end_date = today.strftime("%Y-%m-%d")

    equipment_qs = Equipment.objects.all().select_related('Asset_Type', 'Equipment_Type')
    if asset_type:
        equipment_qs = equipment_qs.filter(Asset_Type__name__icontains=asset_type)
    if eq_type:
        equipment_qs = equipment_qs.filter(Equipment_Type__Equipment_Type__icontains=eq_type)
    if clean_eq_num:
        equipment_qs = equipment_qs.filter(Equipment_Number__icontains=clean_eq_num)
        
    report_data = []
    for eq in equipment_qs:
        accumulated_hours = 0
        failures = 0
        
        if start_date and end_date:
            failures = WorkOrder.objects.filter(
                equipment=eq,
                work_type__work_type__in=["CF", "CP", "WTY"],
                date_created__range=[start_date, end_date]
            ).count()
        
            readings = MeterReading.objects.filter(
                Equipment=eq, 
                Date__range=[start_date, end_date]
            ).order_by('Date')
            
            if readings.exists():
                accumulated_hours = readings.last().Total_Meter_Value - readings.first().Total_Meter_Value
        
        mtbf = round(accumulated_hours / failures, 1) if failures > 0 else accumulated_hours
        
        
        report_data.append({
            'equipment_number': eq.Equipment_Number,
            'equipment_description': eq.Equipment_Description or '',
            'asset_type': eq.Asset_Type.name if eq.Asset_Type else '',
            'equipment_type': eq.Equipment_Type.Equipment_Type if eq.Equipment_Type else '',
            'hours': accumulated_hours,
            'failures': failures,
            'mtbf': mtbf,
        })
        
    sort_by = request.GET.get('sort', 'equipment_number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_key_mapping = {
        'equipment_number': 'equipment_number',
        'equipment_description': 'equipment_description',
        'asset_type': 'asset_type',
        'equipment_type': 'equipment_type',
        'equipment_hours': 'hours',
        'failure_count': 'failures',
        'mtbf': 'mtbf',
    }

    clean_dict_key = sort_key_mapping.get(clean_sort_key, 'equipment_number')
    if report_data and clean_dict_key in report_data[0]:
        report_data.sort(
            key=lambda x: (x[clean_dict_key] is None, x[clean_dict_key]),
            reverse=is_descending
        )

    monthly_failures_qs = (
        WorkOrder.objects.filter(
            equipment__in=equipment_qs,
            date_created__range=[start_date, end_date],
            work_type__work_type__in=["CF", "CP", "WTY"]
        )
        .annotate(month=TruncMonth('date_created'))
        .values('month')
        .annotate(count=Count('id'))
    )

    failures_dict = {entry['month'].date(): entry['count'] for entry in monthly_failures_qs}

    labels, bar_data, line_data = [], [], []
    total_failures_ytd, total_hours_ytd = 0, 0

    s_dt = datetime.strptime(start_date, "%Y-%m-%d")
    e_dt = datetime.strptime(end_date, "%Y-%m-%d")
    
    for dt in rrule(MONTHLY, dtstart=s_dt, until=e_dt):
        current_month_date = dt.date()
        labels.append(dt.strftime('%b'))
        
        m_failures = failures_dict.get(current_month_date, 0)
        m_hours = 0
        
        for eq in equipment_qs:
            m_readings = MeterReading.objects.filter(
                Equipment=eq, 
                Date__month=dt.month,
                Date__year=dt.year
            ).order_by('Date')
            
            if m_readings.exists():
                m_hours += (m_readings.last().Total_Meter_Value - m_readings.first().Total_Meter_Value)
        
        m_mtbf = round(m_hours / m_failures, 1) if m_failures > 0 else m_hours
        bar_data.append(m_mtbf)
        total_failures_ytd += m_failures
        total_hours_ytd += m_hours
        ytd_mtbf = round(total_hours_ytd / total_failures_ytd, 1) if total_failures_ytd > 0 else total_hours_ytd
        line_data.append(ytd_mtbf)

    all_asset_types = AssetType.objects.all().order_by('name')
    if asset_type:
        all_eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_type).distinct().order_by('Equipment_Type')
    else:
        all_eq_types = EQ_Type.objects.all().order_by('Equipment_Type')
        
    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()

    context = {
        'report_data': report_data,
        'labels': labels,
        'bar_data': bar_data,
        'line_data': line_data,
        'current_year': s_dt.year,
        'filters': {'start_date': start_date, 'end_date': end_date, 'equipment_number': eq_num},
        'sort': sort_by,
        'filter_url': filter_url,
        'asset_types': all_asset_types,
        'equipment_types': all_eq_types,
        'all_equipment_suggestions': all_equipment_suggestions,
        'asset_type_val': asset_type,
        'eq_type_val': eq_type,
        'eq_num_val': eq_num,
    }

    return render(request, 'kpis/mtbf.html', context)

def get_mtbf_linked_eq_types(request):
    asset_name = request.GET.get('asset_type', '').strip()
    if asset_name:
        eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_name).values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')
    else:
        eq_types = EQ_Type.objects.values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')

    html_output = '<datalist id="eq-type-list">'
    for et in eq_types:
        html_output += f'<option value="{et}"></option>'
    html_output += '</datalist>'
    return HttpResponse(html_output, content_type="text/html")

def export_mtbf_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    asset_type = request.GET.get('asset_type', '')
    eq_type = request.GET.get('equipment_type', '')
    eq_num = request.GET.get('equipment_number', '')

    equipment_qs = Equipment.objects.all()
    if asset_type:
        equipment_qs = equipment_qs.filter(Asset_Type__name__icontains=asset_type)
    if eq_type:
        equipment_qs = equipment_qs.filter(Equipment_Type__Equipment_Type__icontains=eq_type)
    if eq_num:
        equipment_qs = equipment_qs.filter(Equipment_Number__icontains=eq_num)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MTBF Report"

    headers = [
        'Equip #', 'Equip Description', 'Asset Type', 'Equipment Type', 
        'Hours', 'Failures', 'MTBF (Hours)'
    ]
    ws.append(headers)

    for eq in equipment_qs:
        accumulated_hours = 0
        failures = 0
        
        if start_date and end_date:
            failures = WorkOrder.objects.filter(
                equipment=eq,
                work_type__work_type__in=["CF", "CP", "WTY"],
                date_created__range=[start_date, end_date]
            ).count()

            readings = MeterReading.objects.filter(
                Equipment=eq, 
                Date__range=[start_date, end_date]
            ).order_by('Date')
            
            if readings.exists():
                accumulated_hours = readings.last().Total_Meter_Value - readings.first().Total_Meter_Value

        mtbf = round(accumulated_hours / failures, 4) if failures > 0 else accumulated_hours

        
        ws.append([
            eq.Equipment_Number,
            eq.Equipment_Description,
            str(eq.Asset_Type),
            str(eq.Equipment_Type),
            accumulated_hours,
            failures,
            mtbf
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="MTBF.xlsx"'
    
    wb.save(response)
    return response

def mttr_report(request):
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    asset_type = request.GET.get('asset_type', '').strip()
    eq_type = request.GET.get('equipment_type', '').strip()
    eq_num = request.GET.get('equipment_number', '').strip()

    clean_eq_num = eq_num.split(' - ').strip() if ' - ' in eq_num else eq_num
    
    if not start_date or not end_date or start_date in ['None', ''] or end_date in ['None', '']:
        today = datetime.now()
        start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")  # Jan 1st
        end_date = today.strftime("%Y-%m-%d")

    equipment_qs = Equipment.objects.all().select_related('Asset_Type', 'Equipment_Type')
    if asset_type:
        equipment_qs = equipment_qs.filter(Asset_Type__name__icontains=asset_type)
    if eq_type:
        equipment_qs = equipment_qs.filter(Equipment_Type__Equipment_Type__icontains=eq_type)
    if clean_eq_num:
        equipment_qs = equipment_qs.filter(Equipment_Number__icontains=clean_eq_num)

    report_data = []
    
    for eq in equipment_qs:
        total_repair_hours = 0
        completed_repairs_count = 0

        work_orders = WorkOrder.objects.filter(
            equipment=eq,
            work_type__work_type__in=["CF", "CP", "WTY"],
            date_created__range=[start_date, end_date],
            date_closed__isnull=False
        )

        for wo in work_orders:
            if wo.machine_oos == 'Yes':
                duration = wo.date_closed - wo.date_created
                repair_time = duration.total_seconds() / 3600
            else:
                repair_time = Timesheet.objects.filter(work_order=wo).aggregate(
                    total=Sum('total_time'))['total'] or 0
            
            total_repair_hours += repair_time
            completed_repairs_count += 1

        mttr_val = round(total_repair_hours / completed_repairs_count, 1) if completed_repairs_count > 0 else 0

        report_data.append({
            'equipment_number': eq.Equipment_Number,
            'equipment_description': eq.Equipment_Description or '',
            'asset_type': eq.Asset_Type.name if eq.Asset_Type else '',
            'equipment_type': eq.Equipment_Type.Equipment_Type if eq.Equipment_Type else '',
            'total_repair_hours': total_repair_hours,
            'repair_count': completed_repairs_count,
            'mttr': mttr_val,
        })

    sort_by = request.GET.get('sort', 'equipment_number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_key_mapping = {
        'equipment_number': 'equipment_number',
        'equipment_description': 'equipment_description',
        'asset_type': 'asset_type',
        'equipment_type': 'equipment_type',
        'total_repair_hours': 'total_repair_hours',
        'repair_count': 'repair_count',
        'mttr': 'mttr',
    }

    clean_dict_key = sort_key_mapping.get(clean_sort_key, 'equipment_number')
    if report_data and clean_dict_key in report_data[0]:
        report_data.sort(
            key=lambda x: (x[clean_dict_key] is None, x[clean_dict_key]),
            reverse=is_descending
        )

    monthly_wo_qs = WorkOrder.objects.filter(
        equipment__in=equipment_qs,
        date_created__range=[start_date, end_date],
        work_type__work_type__in=["CF", "CP", "WTY"],
        date_closed__isnull=False
    ).annotate(month_only=TruncMonth('date_created'))

    labels, bar_data, line_data = [], [], []
    cumulative_hrs, cumulative_count = 0, 0
    
    s_dt = datetime.strptime(start_date, "%Y-%m-%d")
    e_dt = datetime.strptime(end_date, "%Y-%m-%d")
    
    for dt in rrule(MONTHLY, dtstart=s_dt, until=e_dt):
        labels.append(dt.strftime('%b'))

        current_month_wos = monthly_wo_qs.filter(month_only__month=dt.month, month_only__year=dt.year)
        
        m_hrs = 0
        m_count = current_month_wos.count()
        
        for wo in current_month_wos:
            if wo.machine_oos == 'Yes':
                diff = wo.date_closed - wo.date_created
                m_hrs += diff.total_seconds() / 3600
            else:
                m_hrs += Timesheet.objects.filter(work_order=wo).aggregate(
                    total=Sum('total_time'))['total'] or 0
                
        m_avg = round(m_hrs / m_count, 1) if m_count > 0 else 0
        bar_data.append(m_avg)
        
        cumulative_hrs += m_hrs
        cumulative_count += m_count
        ytd_avg = round(cumulative_hrs / cumulative_count, 1) if cumulative_count > 0 else 0
        line_data.append(ytd_avg)

    all_asset_types = AssetType.objects.all().order_by('name')
    if asset_type:
        all_eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_type).distinct().order_by('Equipment_Type')
    else:
        all_eq_types = EQ_Type.objects.all().order_by('Equipment_Type')
        
    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()
        
    context = {
        'report_data': report_data,
        'labels': labels,
        'bar_data': bar_data,
        'line_data': line_data,
        'current_year': s_dt.year,
        'filters': {'start_date': start_date, 'end_date': end_date, 'equipment_number': eq_num},
        'sort': sort_by,
        'filter_url': filter_url,
        'asset_types': all_asset_types,
        'equipment_types': all_eq_types,
        'all_equipment_suggestions': all_equipment_suggestions,
        'asset_type_val': asset_type,
        'eq_type_val': eq_type,
        'eq_num_val': eq_num,
    }
    return render(request, 'kpis/mttr.html', context)

def get_mttr_linked_eq_types(request):
    asset_name = request.GET.get('asset_type', '').strip()
    if asset_name:
        eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_name).values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')
    else:
        eq_types = EQ_Type.objects.values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')

    html_output = '<datalist id="eq-type-list">'
    for et in eq_types:
        html_output += f'<option value="{et}"></option>'
    html_output += '</datalist>'
    return HttpResponse(html_output, content_type="text/html")

def export_mttr_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    asset_type = request.GET.get('asset_type', '')
    eq_type = request.GET.get('equipment_type', '')
    eq_num = request.GET.get('equipment_number', '')

    equipment_qs = Equipment.objects.all()
    if asset_type:
        equipment_qs = equipment_qs.filter(Asset_Type__name__icontains=asset_type)
    if eq_type:
        equipment_qs = equipment_qs.filter(Equipment_Type__Equipment_Type__icontains=eq_type)
    if eq_num:
        equipment_qs = equipment_qs.filter(Equipment_Number__icontains=eq_num)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MTTR Report"

    headers = [
        'Equip #', 'Description', 'Asset Type', 'Equipment Type', 
        'Repair Count', 'Total Repair Hours', 'MTTR (Hours)'
    ]
    ws.append(headers)

    for eq in equipment_qs:
        total_repair_hours = 0
        completed_count = 0

        if start_date and end_date:
            work_orders = WorkOrder.objects.filter(
                equipment=eq,
                work_type__work_type__in=["CF", "CP", "WTY"],
                date_created__range=[start_date, end_date],
                date_closed__isnull=False
            )

            for wo in work_orders:
                if wo.machine_oos == 'yes':
                    duration = wo.date_closed - wo.date_created
                    repair_time = duration.total_seconds() / 3600
                else:
                    repair_time = Timesheet.objects.filter(work_order=wo).aggregate(
                        total=Sum('total_time'))['total'] or 0
                
                total_repair_hours += repair_time
                completed_count += 1

        mttr = round(total_repair_hours / completed_count, 1) if completed_count > 0 else 0

        ws.append([
            eq.Equipment_Number,
            eq.Equipment_Description,
            str(eq.Asset_Type),
            str(eq.Equipment_Type),
            completed_count,
            round(total_repair_hours, 1),
            mttr
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="MTTR_Report.xlsx"'
    
    wb.save(response)
    return response

def availability_utilisation_report(request):
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    asset_type = request.GET.get('asset_type', '')
    eq_type = request.GET.get('equipment_type', '')
    eq_num = request.GET.get('equipment_number', '')
    
    if not start_date or not end_date or start_date in ['None', ''] or end_date in ['None', '']:
        today = datetime.now()
        start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")  # Jan 1st
        end_date = today.strftime("%Y-%m-%d")

    clean_eq_num = eq_num.split(' - ')[0].strip() if ' - ' in eq_num else eq_num

    equipment_qs = Equipment.objects.all().select_related('Asset_Type', 'Equipment_Type')
    if asset_type:
        equipment_qs = equipment_qs.filter(Asset_Type__name__icontains=asset_type)
    if eq_type:
        equipment_qs = equipment_qs.filter(Equipment_Type__Equipment_Type__icontains=eq_type)
    if clean_eq_num:
        equipment_qs = equipment_qs.filter(Equipment_Number__icontains=clean_eq_num)

    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    delta = end_dt - start_dt
    total_days = delta.days + 1
    total_potential_hours = total_days * 24

    report_data = []
    for eq in equipment_qs:
        stats = MachineShiftStatus.objects.filter(
            equipment=eq,
            report__date__range=[start_date, end_date]
        ).aggregate(
            sum_down=Sum('total_down'),
            sum_worked=Sum('total_worked'),
            sum_available=Sum('available')
        )
        
        total_down = stats['sum_down'] or 0
        worked_hrs = stats['sum_worked'] or 0
        available = stats['sum_available'] or 0
                
        availability_pct = round((available / total_potential_hours) * 100, 1) if total_potential_hours > 0 else 0
        utilisation_pct = round((worked_hrs / total_potential_hours) * 100, 1) if total_potential_hours > 0 else 0


        report_data.append({
            'equipment_number': eq.Equipment_Number,
            'equipment_description': eq.Equipment_Description or '',
            'asset_type': eq.Asset_Type.name if eq.Asset_Type else '',
            'equipment_type': eq.Equipment_Type.Equipment_Type if eq.Equipment_Type else '',
            'total_down': total_down,
            'available': max(availability_pct, 0),
            'total_worked': max(utilisation_pct, 0),
        })

    sort_by = request.GET.get('sort', 'equipment_number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    valid_sort_keys = ['equipment_number', 'equipment_description', 'asset_type', 'equipment_type', 'total_down', 'available', 'total_worked']
    if clean_sort_key in valid_sort_keys:
        report_data.sort(
            key=lambda x: (x[clean_sort_key] is None, x[clean_sort_key]),
            reverse=is_descending
        )

    monthly_stats = MachineShiftStatus.objects.filter(
        equipment__in=equipment_qs,
        report__date__range=[start_date, end_date]
    ).annotate(month=TruncMonth('report__date')).values('month').annotate(
        m_available=Sum('available'), m_worked=Sum('total_worked')
    ).order_by('month')

    stats_dict = {entry['month']: entry for entry in monthly_stats}
    labels, avail_chart_data, util_chart_data = [], [], []

    for dt in rrule(MONTHLY, dtstart=start_dt, until=end_dt):
        labels.append(dt.strftime('%b'))
        month_data = stats_dict.get(dt.date(), {})
        days_in_month = calendar.monthrange(dt.year, dt.month)[1]
        m_potential_hrs = equipment_qs.count() * days_in_month * 24
        m_avail_val = month_data.get('m_available') or 0
        m_worked_val = month_data.get('m_worked') or 0
        avail_chart_data.append(round((m_avail_val / m_potential_hrs) * 100, 1) if m_potential_hrs > 0 else 0)
        util_chart_data.append(round((m_worked_val / m_potential_hrs) * 100, 1) if m_potential_hrs > 0 else 0)

    all_asset_types = AssetType.objects.all().order_by('name')

    if asset_type:
        all_eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_type).distinct().order_by('Equipment_Type')
    else:
        all_eq_types = EQ_Type.objects.all().order_by('Equipment_Type')
        
    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()

    context = {
        'report_data': report_data,
        'labels': labels,
        'avail_chart_data': avail_chart_data,
        'util_chart_data': util_chart_data,
        'filters': {'start_date': start_date, 'end_date': end_date, 'equipment_number': eq_num},
        'sort': sort_by,
        'filter_url': filter_url,
        'asset_types': all_asset_types,
        'equipment_types': all_eq_types,
        'all_equipment_suggestions': all_equipment_suggestions,
        'asset_type_val': asset_type,
        'eq_type_val': eq_type,
        'eq_num_val': eq_num,
    }
    return render(request, 'kpis/availability.html', context)

def get_kpi_linked_eq_types(request):
    asset_name = request.GET.get('asset_type', '').strip()
    if asset_name:
        eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_name).values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')
    else:
        eq_types = EQ_Type.objects.values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')

    html_output = '<datalist id="eq-type-list">'
    for et in eq_types:
        html_output += f'<option value="{et}"></option>'
    html_output += '</datalist>'
    return HttpResponse(html_output, content_type="text/html")

def export_au_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    asset_type = request.GET.get('asset_type', '')
    eq_type = request.GET.get('equipment_type', '')
    eq_num = request.GET.get('equipment_number', '')

    equipment_qs = Equipment.objects.all()
    if asset_type:
        equipment_qs = equipment_qs.filter(Asset_Type__name__icontains=asset_type)
    if eq_type:
        equipment_qs = equipment_qs.filter(Equipment_Type__Equipment_Type__icontains=eq_type)
    if eq_num:
        equipment_qs = equipment_qs.filter(Equipment_Number__icontains=eq_num)
        
    try:
        delta = datetime.strptime(end_date, "%Y-%m-%d") - datetime.strptime(start_date, "%Y-%m-%d")
        total_potential_hours = max((delta.days + 1) * 24, 1)
    except:
        total_potential_hours = 24
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "A&U Report"

    headers = ['Equipment #', 'Description', 'Asset Type', 'Downtime (Hrs)', 'Availability %', 'Utilisation %']
    header_fill = PatternFill(start_color="36A2EB", end_color="36A2EB", fill_type="solid")
    
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for eq in equipment_qs:
        stats = MachineShiftStatus.objects.filter(
            equipment=eq,
            report__date__range=[start_date, end_date]
        ).aggregate(
            sum_down=Sum('total_down'),
            sum_worked=Sum('total_worked'),
            sum_available=Sum('available')
        )

        down = stats['sum_down'] or 0
        worked = stats['sum_worked'] or 0
        available_hrs = stats['sum_available'] or 0
        
        avail_pct = round((available_hrs / total_potential_hours) * 100, 1) if total_potential_hours > 0 else 0
        util_pct = round((worked / total_potential_hours) * 100, 1) if total_potential_hours > 0 else 0

        ws.append([
            eq.Equipment_Number,
            eq.Equipment_Description,
            str(eq.Asset_Type),
            down,
            f"{avail_pct}%",
            f"{util_pct}%"
        ])
        
    for col in ws.columns:
        column_letter = col[0].column_letter
        ws.column_dimensions[column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=AU_Report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response

def schedule_compliance_report(request):
    week_param = request.GET.get('week', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    target = float(request.GET.get('target', 85) or 85)
    
    asset_type = request.GET.get('asset_type', '').strip()
    eq_type = request.GET.get('equipment_type', '').strip()
    eq_num = request.GET.get('equipment_number', '').strip()

    clean_eq_num = eq_num.split(' - ')[0].strip() if ' - ' in eq_num else eq_num
    
    clean_week_number = None
    if week_param:
        if 'Week ' in week_param:
            clean_week_number = week_param.split(' ')[1].strip()
        else:
            clean_week_number = week_param

    today = timezone.now().date()
    snapshots = ScheduleSnapshot.objects.select_related(
        'work_order', 'schedule__week', 'work_order__equipment', 'work_order__job_status', 'work_order__work_type'
    ).all()

    if clean_week_number:
        snapshots = snapshots.filter(schedule__week__week_number=clean_week_number)
    if start_date and end_date:
        snapshots = snapshots.filter(schedule__week__start_date__range=[start_date, end_date])
    if asset_type:
        snapshots = snapshots.filter(work_order__equipment__Asset_Type__name__icontains=asset_type)
    if eq_type:
        snapshots = snapshots.filter(work_order__equipment__Equipment_Type__Equipment_Type__icontains=eq_type)
    if clean_eq_num:
        snapshots = snapshots.filter(work_order__equipment__Equipment_Number__icontains=clean_eq_num)

    table_data = []
    compliant_count = 0
    total_count = snapshots.count()

    for ss in snapshots:
        wo = ss.work_order
        closed_date = ss.date_closed_snapshot
        status = wo.job_status.status_choice.lower() if wo.job_status else ""
        
        if any(term in status for term in ["complete", "cancel", "reschedule"]):
            delta = (closed_date - ss.plan_start_snapshot).days if closed_date else 0
        else:
            delta = (today - ss.plan_start_snapshot).days

        is_compliant = closed_date and closed_date <= ss.schedule.week.end_date
        if is_compliant:
            compliant_count += 1

        table_data.append({
            'work_order': wo.work_order,
            'equipment': wo.equipment.Equipment_Number,
            'work_type': wo.work_type.work_type if wo.work_type else '---',
            'job_status': wo.job_status.status_choice if wo.job_status else '---',
            'plan_start': ss.plan_start_snapshot,
            'date_closed': closed_date,
            'delta': delta,
            'is_compliant': is_compliant
        })

    sort_by = request.GET.get('sort', 'work_order')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    valid_sort_keys = ['work_order', 'equipment', 'work_type', 'job_status', 'plan_start', 'date_closed', 'delta']
    if clean_sort_key in valid_sort_keys:
        table_data.sort(
            key=lambda x: (x[clean_sort_key] is None, x[clean_sort_key]),
            reverse=is_descending
        )

    weeks_stats = (
        snapshots.values(
            'schedule__week__week_number',
            'schedule__week__start_date',
            'schedule__week__end_date'
        )
        .annotate(
            total=Count('id'),
            compliant=Count(
                'id', 
                filter=Q(work_order__date_closed__lte=F('schedule__week__end_date'))
            )
        )
        .order_by('schedule__week__start_date')
    )

    labels = []
    bar_data = []
    
    for w in weeks_stats:
        w_num = w['schedule__week__week_number']
        s_date = w['schedule__week__start_date']

        if s_date:
            date_label = s_date.strftime('%b %d')
            labels.append(f"Wk {w_num} ({date_label})")
        else:
            labels.append(f"Wk {w_num}")
            
        total_wos = w['total']
        compliant_wos = w['compliant']
        bar_data.append(round((compliant_wos / total_wos * 100), 1) if total_wos > 0 else 0)

    line_data = [target] * len(labels)

    all_weeks = WorkWeek.objects.all().order_by('week_number')
    all_asset_types = AssetType.objects.all().order_by('name')
    if asset_type:
        all_eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_type).distinct().order_by('Equipment_Type')
    else:
        all_eq_types = EQ_Type.objects.all().order_by('Equipment_Type')
        
    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()

    context = {
        'table_data': table_data,
        'labels': labels,
        'bar_data': bar_data,
        'line_data': line_data,
        'sort': sort_by,
        'filter_url': filter_url,
        'weeks': all_weeks,
        'asset_types': all_asset_types,
        'equipment_types': all_eq_types,
        'all_equipment_suggestions': all_equipment_suggestions,
        'week_val': week_param,
        'start_date_val': start_date,
        'end_date_val': end_date,
        'asset_type_val': asset_type,
        'eq_type_val': eq_type,
        'eq_num_val': eq_num,
        'target': target,
        'overall_compliance': round((compliant_count / total_count * 100), 1) if total_count > 0 else 0
    }
    return render(request, 'kpis/schedule_compliance.html', context)

def get_compliance_linked_eq_types(request):
    asset_name = request.GET.get('asset_type', '').strip()
    if asset_name:
        eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_name).values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')
    else:
        eq_types = EQ_Type.objects.values_list('Equipment_Type', flat=True).distinct().order_by('Equipment_Type')

    html_output = '<datalist id="eq-type-list">'
    for et in eq_types:
        html_output += f'<option value="{et}"></option>'
    html_output += '</datalist>'
    return HttpResponse(html_output, content_type="text/html")

def export_compliance_excel(request):
    week_id = request.GET.get('week')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    asset_type = request.GET.get('asset_type')
    eq_num = request.GET.get('equipment_number')
    today = timezone.now().date()

    snapshots = ScheduleSnapshot.objects.select_related('work_order', 'schedule__week', 'work_order__equipment')

    if week_id:
        snapshots = snapshots.filter(schedule__week_id=week_id)
    if start_date and end_date:
        snapshots = snapshots.filter(schedule__week__start_date__range=[start_date, end_date])
    if asset_type:
        snapshots = snapshots.filter(work_order__equipment__Asset_Type__name__icontains=asset_type)
    if eq_num:
        snapshots = snapshots.filter(work_order__equipment__Equipment_Number__icontains=eq_num)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule Compliance"

    headers = ['Work Order', 'Equipment', 'Status', 'Plan Start', 'Date Closed', 'Delta (Days)', 'Compliant']
    ws.append(headers)

    for ss in snapshots:
        wo = ss.work_order
        closed_date = ss.date_closed_snapshot
        status = wo.job_status.status_choice.lower()
        
        if any(term in status for term in ["complete", "cancel", "reschedule"]):
            delta = (closed_date - ss.plan_start_snapshot).days if closed_date else 0
        else:
            delta = (today - ss.plan_start_snapshot).days

        is_compliant = "Yes" if (closed_date and closed_date <= ss.schedule.week.end_date) else "No"

        ws.append([
            wo.work_order,
            wo.equipment.Equipment_Number,
            wo.job_status.status_choice,
            ss.plan_start_snapshot.strftime('%Y-%m-%d'),
            closed_date.strftime('%Y-%m-%d') if closed_date else "N/A",
            delta,
            is_compliant
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Schedule_Compliance_{today}.xlsx'
    wb.save(response)
    return response

def resource_utilisation_report(request):
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    asset_type = request.GET.get('asset_type', '').strip()
    eq_type = request.GET.get('equipment_type', '').strip()
    eq_num = request.GET.get('equipment_number', '').strip()

    clean_eq_num = eq_num.split(' - ')[0].strip() if ' - ' in eq_num else eq_num

    if not start_date or not end_date or start_date in ['None', ''] or end_date in ['None', '']:
        today = datetime.now()
        start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")

    time_types = ['Prep', 'Travel', 'Troubleshoot', 'Repair']

    wo_qs = WorkOrder.objects.all().select_related(
        'equipment', 'equipment__Asset_Type', 'equipment__Equipment_Type'
    )

    if start_date and end_date:
        wo_qs = wo_qs.filter(date_closed__range=[start_date, end_date])
    if asset_type:
        wo_qs = wo_qs.filter(equipment__Asset_Type__name__icontains=asset_type)
    if eq_type:
        wo_qs = wo_qs.filter(equipment__Equipment_Type__Equipment_Type__icontains=eq_type)
    if clean_eq_num:
        wo_qs = wo_qs.filter(equipment__Equipment_Number__icontains=clean_eq_num)

    table_data = []
    for wo in wo_qs:
        type_hours = {}
        total_actual = 0
        ts_data = Timesheet.objects.filter(work_order=wo)

        for t_type in time_types:
            amount = ts_data.filter(time_type=t_type).aggregate(Sum('total_time'))['total_time__sum'] or 0
            type_hours[t_type] = float(amount)
            total_actual += float(amount)

        table_data.append({
            'work_order': wo.work_order,
            'equipment_num': wo.equipment.Equipment_Number,
            'equipment_name': wo.equipment.Equipment_Description or '',
            'troubleshoot': wo.troubleshoot_description or '',
            'repair': wo.repair_description or '',
            'est_hours': float(wo.est_work_hours or 0),
            'type_hours': type_hours,
            'total_actual': total_actual,
            'over_under': float(wo.est_work_hours or 0) - total_actual,
            'month_key': wo.date_closed.replace(day=1) if wo.date_closed else None
        })

    sort_by = request.GET.get('sort', 'work_order')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_key_mapping = {
        'work_order': 'work_order',
        'equipment_num': 'equipment_num',
        'equipment_name': 'equipment_name',
        'troubleshoot': 'troubleshoot',
        'repair': 'repair',
        'est_hours': 'est_hours',
        'total_actual': 'total_actual',
        'over_under': 'over_under',
    }

    clean_dict_key = sort_key_mapping.get(clean_sort_key, 'work_order')
    
    if clean_dict_key in sort_key_mapping:
        table_data.sort(
            key=lambda x: (x[clean_dict_key] is None, x[clean_dict_key]),
            reverse=is_descending
        )
    elif clean_sort_key in time_types:
        table_data.sort(
            key=lambda x: (x['type_hours'].get(clean_sort_key, 0) is None, x['type_hours'].get(clean_sort_key, 0)),
            reverse=is_descending
        )

    labels = []
    est_line_data = []
    type_series = {t_type: [] for t_type in time_types}

    s_dt = datetime.strptime(start_date, '%Y-%m-%d')
    e_dt = datetime.strptime(end_date, '%Y-%m-%d')

    for dt in rrule(MONTHLY, dtstart=s_dt, until=e_dt):
        labels.append(dt.strftime('%b %Y'))

        month_items = [item for item in table_data if item['month_key'] and item['month_key'].date() == dt.date()]
        month_est = sum(item['est_hours'] for item in month_items)
        est_line_data.append(month_est)

        for t_type in time_types:
            month_type_total = sum(item['type_hours'].get(t_type, 0) for item in month_items)
            type_series[t_type].append(month_type_total)

    colors = {'Prep': '#36a2eb', 'Travel': '#ffce56', 'Troubleshoot': '#4bc0c0', 'Repair': '#9966ff'}
    stacked_datasets = []
    for t_type in time_types:
        stacked_datasets.append({
            'label': t_type,
            'data': type_series[t_type],
            'backgroundColor': colors.get(t_type, '#cccccc'),
            'type': 'bar',
            'stack': 'actual'
        })

    all_asset_types = AssetType.objects.all().order_by('name')
    if asset_type:
        all_eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_type).distinct().order_by('Equipment_Type')
    else:
        all_eq_types = EQ_Type.objects.all().order_by('Equipment_Type')
        
    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()

    context = {
        'table_data': table_data,
        'time_types': time_types,
        'labels': labels,
        'est_line_data': est_line_data,
        'stacked_datasets': stacked_datasets,
        'sort': sort_by,
        'filter_url': filter_url,
        'asset_types': all_asset_types,
        'equipment_types': all_eq_types,
        'all_equipment_suggestions': all_equipment_suggestions,
        'start_date_val': start_date,
        'end_date_val': end_date,
        'asset_type_val': asset_type,
        'eq_type_val': eq_type,
        'eq_num_val': eq_num,
    }
    return render(request, 'kpis/resource_utilisation.html', context)

def export_res_util_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    asset_type = request.GET.get('asset_type', '')
    eq_type = request.GET.get('equipment_type', '')
    eq_num = request.GET.get('equipment_number', '')

    time_types = ['Prep', 'Travel', 'Troubleshoot', 'Repair']

    wo_qs = WorkOrder.objects.all().select_related(
        'equipment', 'equipment__Asset_Type', 'equipment__Equipment_Type'
    )

    if start_date and end_date:
        wo_qs = wo_qs.filter(date_created__range=[start_date, end_date])
    if asset_type:
        wo_qs = wo_qs.filter(equipment__Asset_Type__name__icontains=asset_type)
    if eq_type:
        wo_qs = wo_qs.filter(equipment__Equipment_Type__Equipment_Type__icontains=eq_type)
    if eq_num:
        wo_qs = wo_qs.filter(equipment__Equipment_Number__icontains=eq_num)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resource Utilisation"

    headers = ['Work Order', 'Equipment #', 'Make/Model', 'Troubleshoot', 'Repair', 'Est. Hours']
    headers.extend(time_types)
    headers.append('Over / Under')
    ws.append(headers)

    for wo in wo_qs:
        row = [
            wo.work_order,
            wo.equipment.Equipment_Number,
            f"{wo.equipment.Make} {wo.equipment.Model}",
            wo.troubleshoot_description or "",
            wo.repair_description or "",
            wo.est_work_hours or 0,
        ]

        total_actual = 0
        for t_type in time_types:
            amount = Timesheet.objects.filter(
                work_order=wo, 
                time_type=t_type
            ).aggregate(Sum('total_time'))['total_time__sum'] or 0
            row.append(amount)
            total_actual += amount
        
        row.append(total_actual)
        over_under = (wo.est_work_hours or 0) - total_actual
        row.append(over_under)
        ws.append(row)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Resource_Utilisation_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def cost_per_hour_report(request):
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    asset_type = request.GET.get('asset_type', '').strip()
    eq_type = request.GET.get('equipment_type', '').strip()
    eq_num = request.GET.get('equipment_number', '').strip()

    clean_eq_num = eq_num.split(' - ')[0].strip() if ' - ' in eq_num else eq_num

    if not start_date or not end_date or start_date in ['None', ''] or end_date in ['None', '']:
        today = datetime.now()
        start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")  # Jan 1st
        end_date = today.strftime("%Y-%m-%d")

    s_dt = datetime.strptime(start_date, '%Y-%m-%d')
    e_dt = datetime.strptime(end_date, '%Y-%m-%d')

    equipment_qs = Equipment.objects.all().select_related('Asset_Type', 'Equipment_Type', 'Garage')
    if asset_type:
        equipment_qs = equipment_qs.filter(Asset_Type__name__icontains=asset_type)
    if eq_type:
        equipment_qs = equipment_qs.filter(Equipment_Type__Equipment_Type__icontains=eq_type)
    if clean_eq_num:
        equipment_qs = equipment_qs.filter(Equipment_Number__icontains=clean_eq_num)

    table_data = []

    for eq in equipment_qs:
        work_orders = WorkOrder.objects.filter(
            equipment=eq,
            date_closed__range=[s_dt, e_dt]
        )
        
        if not work_orders.exists():
            continue

        readings = MeterReading.objects.filter(
            Equipment=eq, 
            Date__range=[s_dt.date(), e_dt.date()]
        ).order_by('Date')

        meter_delta = 0.0
        if readings.exists():
            meter_delta = float((readings.last().Total_Meter_Value or 0) - (readings.first().Total_Meter_Value or 0))

        shop_rate = eq.Garage.Shop_Rate if eq.Garage else Decimal('0.00')
        ts_hours = Timesheet.objects.filter(work_order__in=work_orders).aggregate(Sum('total_time'))['total_time__sum'] or 0
        labour_cost = float(shop_rate) * float(ts_hours)
        parts_cost = float(Purchase.objects.filter(wo_cc__in=work_orders).aggregate(Sum('grand_total'))['grand_total__sum'] or 0)
        
        total_cost = labour_cost + parts_cost
        cost_per_hr = total_cost / meter_delta if meter_delta > 0 else 0.0

        latest_wo = work_orders.order_by('-date_closed').first()
        month_key = latest_wo.date_closed.replace(day=1) if latest_wo and latest_wo.date_closed else None

        table_data.append({
            'equipment_num': eq.Equipment_Number,
            'description': eq.Equipment_Description or '',
            'labour_cost': round(labour_cost, 2),
            'parts_cost': round(parts_cost, 2),
            'total_cost': round(total_cost, 2),
            'total_hours': meter_delta,
            'cost_per_hr': round(cost_per_hr, 2),
            'month_key': month_key
        })

    sort_by = request.GET.get('sort', 'equipment_num')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_key_mapping = {
        'equipment_num': 'equipment_num',
        'description': 'description',
        'labour_cost': 'labour_cost',
        'parts_cost': 'parts_cost',
        'total_cost': 'total_cost',
        'total_hours': 'total_hours',
        'cost_per_hr': 'cost_per_hr',
    }

    clean_dict_key = sort_key_mapping.get(clean_sort_key, 'equipment_num')
    if clean_dict_key in sort_key_mapping and table_data:
        table_data.sort(
            key=lambda x: (x[clean_dict_key] is None, x[clean_sort_key.lstrip('-')]),
            reverse=is_descending
        )

    labels, labour_series, parts_series, cost_hr_series = [], [], [], []

    s_dt = datetime.strptime(start_date, '%Y-%m-%d')
    e_dt = datetime.strptime(end_date, '%Y-%m-%d')

    for dt in rrule(MONTHLY, dtstart=s_dt, until=e_dt):
        labels.append(dt.strftime('%b %Y'))
        month_items = [item for item in table_data if item['month_key'] and item['month_key'].date() == dt.date()]
        
        m_labour = sum(item['labour_cost'] for item in month_items)
        m_parts = sum(item['parts_cost'] for item in month_items)
        m_hours = sum(item['total_hours'] for item in month_items)
        
        m_total_cost = m_labour + m_parts
        m_cph = m_total_cost / m_hours if m_hours > 0 else 0
            
        labour_series.append(round(m_labour, 2))
        parts_series.append(round(m_parts, 2))
        cost_hr_series.append(round(m_cph, 2))

    all_asset_types = AssetType.objects.all().order_by('name')
    if asset_type:
        all_eq_types = EQ_Type.objects.filter(equipment__Asset_Type__name__icontains=asset_type).distinct().order_by('Equipment_Type')
    else:
        all_eq_types = EQ_Type.objects.all().order_by('Equipment_Type')
        
    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()

    context = {
        'table_data': table_data,
        'labels': labels,
        'labour_series': labour_series,
        'parts_series': parts_series,
        'cost_hr_series': cost_hr_series,
        'sort': sort_by,
        'filter_url': filter_url,
        'asset_types': all_asset_types,
        'equipment_types': all_eq_types,
        'all_equipment_suggestions': all_equipment_suggestions,
        'start_date_val': start_date,
        'end_date_val': end_date,
        'asset_type_val': asset_type,
        'eq_type_val': eq_type,
        'eq_num_val': eq_num,
    }
    return render(request, 'kpis/cost_per_hour.html', context)

def export_cost_report_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    eq_num = request.GET.get('equipment_number', '')

    wo_qs = WorkOrder.objects.all().select_related(
        'equipment', 'equipment__Garage'
    )

    if start_date and end_date:
        wo_qs = wo_qs.filter(date_closed__range=[start_date, end_date])
    if eq_num:
        wo_qs = wo_qs.filter(equipment__Equipment_Number__icontains=eq_num)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cost Per Hour Report"

    headers = [
        'Equipment #', 'Description', 'Meter Hours', 
        'Labour Cost', 'Parts Cost', 'Total Cost', 'Cost / Hr'
    ]
    ws.append(headers)

    for wo in wo_qs:
        eq = wo.equipment

        readings = MeterReading.objects.filter(
            Equipment=eq, 
            Date__range=[start_date, end_date]
        ).order_by('Date')

        if readings.exists():
            start_meter = readings.first().Total_Meter_Value
            end_meter = readings.last().Total_Meter_Value
            meter_delta = float(end_meter - start_meter)
        else:
            meter_delta = 0

        shop_rate = eq.Garage.Shop_Rate if eq.Garage else Decimal('0.00')
        ts_hours = Timesheet.objects.filter(work_order=wo).aggregate(Sum('total_time'))['total_time__sum'] or 0
        labour_cost = float(shop_rate) * float(ts_hours)

        parts_cost = float(Purchase.objects.filter(wo_cc=wo).aggregate(Sum('grand_total'))['grand_total__sum'] or 0)
        
        total_wo_cost = labour_cost + parts_cost
        
        cost_per_hr = total_wo_cost / meter_delta if meter_delta > 0 else 0

        ws.append([
            eq.Equipment_Number,
            eq.Equipment_Description,
            round(meter_delta, 1),
            round(labour_cost, 2),
            round(parts_cost, 2),
            round(total_wo_cost, 2),
            round(cost_per_hr, 2)
        ])

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Cost_Per_Hour_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response