from django.contrib import messages
from django.forms.models import inlineformset_factory
from django.db import transaction
from django.http import HttpResponse
from django.utils.http import url_has_allowed_host_and_scheme
import openpyxl
from .forms import (QualityMaintenanceCreateForm, QualityMaintenanceEditForm, QualityMaintenanceStepFormSet, QualityMaintenanceDocument,
    QualityMaintenanceSearchForm, QualityMaintenanceEditLookupForm, QualityMaintenanceStepForm, QualityMaintenancePlanForm)
from .models import QualityMaintenanceDocument, QualityMaintenanceInstance, QualityMaintenanceDocumentStep, QualityMaintenancePlan
from equipment.models import Equipment, Meter
from work_orders.models import WorkOrder
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count
from django.shortcuts import get_object_or_404, redirect, render
from .services import evaluate_plan_for_forecast, evaluate_qm_for_work_order_creation

def planning(request):
    return render(request, 'planning/planning.html')

def create_qm(request):
    StepFormSetClass = inlineformset_factory(
        QualityMaintenanceDocument,
        QualityMaintenanceDocumentStep,
        form=QualityMaintenanceStepForm,
        extra=2,
        can_delete=True
    )
    if request.method == 'POST':
        form = QualityMaintenanceCreateForm(request.POST, request.FILES)
        formset = StepFormSetClass(request.POST, request.FILES)

        form_valid = form.is_valid()
        formset_valid = formset.is_valid()

        if form_valid and formset_valid:
            step_type = form.cleaned_data.get('step_type')
            qm_type = form.cleaned_data.get('qm_type')

            if step_type == 'MULTI':
                valid_step_count = 0
                calendar_groups = set()

                for f in formset.forms:
                    if f.cleaned_data and not f.cleaned_data.get('DELETE', False):
                        cleaned = f.cleaned_data
                        valid_step_count += 1

                        unit = cleaned.get('interval_unit')

                        if qm_type == 'CALENDAR':
                            if not unit:
                                f.add_error('interval_unit', 'Required for calendar steps.')
                            elif unit in ['DAY', 'WEEK']:
                                calendar_groups.add('DAY_BASED')
                            elif unit in ['MONTH', 'YEAR']:
                                calendar_groups.add('MONTH_BASED')

                        elif qm_type == 'METER' and unit:
                            f.add_error('interval_unit', 'Meter steps should not have calendar units.')


                if valid_step_count == 0:
                    form.add_error(None, 'At least one sequence step is required for a multi-step document.')
                
                if qm_type == 'CALENDAR' and len(calendar_groups) > 1:
                    form.add_error(None, 'Cannot mix Day/Week intervals with Month/Year intervals in one document.')

        if not form.errors:
                with transaction.atomic():
                    qm = form.save()
                    formset.instance = qm
                    formset.save()
                messages.success(request, f'QM Document {qm.qm_number} created successfully.')
                return redirect('planning:edit_qm_record', pk=qm.pk)    
    else:
        form = QualityMaintenanceCreateForm()
        formset = StepFormSetClass()

    return render(request, 'planning/create_qm.html', {
        'qmform': form,
        'step_formset': formset,
        'preview_number': QualityMaintenanceDocument.get_next_number()
    })
    
def lookup_qm_for_edit(request):
    all_qm_suggestions = QualityMaintenanceDocument.objects.all().only('qm_number', 'description').order_by('qm_number')

    if request.method == 'POST':
        form = QualityMaintenanceEditLookupForm(request.POST)
        if form.is_valid():
            qm_number_input = form.cleaned_data.get('qm_number').strip()
            
            target_qm = QualityMaintenanceDocument.objects.filter(qm_number__iexact=qm_number_input).first()
            
            if target_qm:
                return redirect('planning:edit_qm_record', pk=target_qm.pk)
            else:
                form.add_error('qm_number', f"No Quality Maintenance Document found with number '{qm_number_input}'.")
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = QualityMaintenanceEditLookupForm()

    return render(request, 'planning/edit_qm.html', {
        'lookup_form': form,
        'all_qm_suggestions': all_qm_suggestions,
    })
    
def edit_qm_record(request, pk):
    qm = get_object_or_404(QualityMaintenanceDocument, pk=pk)
    
    StepFormSetClass = inlineformset_factory(
        QualityMaintenanceDocument,
        QualityMaintenanceDocumentStep,
        form=QualityMaintenanceStepForm,
        extra=0,
        can_delete=True
    )

    if request.method == 'POST':
        form = QualityMaintenanceEditForm(request.POST, request.FILES, instance=qm)
        formset = StepFormSetClass(request.POST, request.FILES, instance=qm)
        
        if form.is_valid() and formset.is_valid():
            # 2. Structural Validation
            step_type = form.cleaned_data.get('step_type')
            qm_type = form.cleaned_data.get('qm_type')
            
            if step_type == 'MULTI':
                valid_steps = 0
                calendar_groups = set()
                
                for f in formset.forms:
                    if f.cleaned_data and not f.cleaned_data.get('DELETE', False):
                        valid_steps += 1
                        unit = f.cleaned_data.get('interval_unit')
                        
                        if qm_type == 'CALENDAR':
                            if not unit:
                                f.add_error('interval_unit', 'Required for calendar steps.')
                            elif unit in ['DAY', 'WEEK']:
                                calendar_groups.add('DAY_BASED')
                            elif unit in ['MONTH', 'YEAR']:
                                calendar_groups.add('MONTH_BASED')
                        elif qm_type == 'METER' and unit:
                            f.add_error('interval_unit', 'Meter steps should not have units.')

                if valid_steps == 0:
                    form.add_error(None, 'Multi-step documents require at least one step.')
                if qm_type == 'CALENDAR' and len(calendar_groups) > 1:
                    form.add_error(None, 'Cannot mix Day/Week intervals with Month/Year intervals.')

            if not form.errors:
                with transaction.atomic():
                    qm = form.save()
                    formset.save()
                
                messages.success(request, f'QM Document {qm.qm_number} updated successfully.')
                
                if 'save_exit' in request.POST:
                    return redirect('planning:planning')
                return redirect('planning:edit_qm_record', pk=qm.pk)
    else:
        form = QualityMaintenanceEditForm(instance=qm)
        formset = StepFormSetClass(instance=qm)

    return render(request, 'planning/edit_qm_record.html', {
        'qmform': form,
        'step_formset': formset,
        'qm': qm,
    })
    
def search_qm(request):
    form = QualityMaintenanceSearchForm(request.GET or None)

    results = QualityMaintenanceDocument.objects.annotate(
        plan_count=Count('plans')
    ).all()

    if form.is_valid():
        qm_number = form.cleaned_data.get('qm_number')
        description = form.cleaned_data.get('description')
        qm_type = form.cleaned_data.get('qm_type')
        step_type = form.cleaned_data.get('step_type')
        active_input = form.cleaned_data.get('active')

        if qm_number:
            results = results.filter(qm_number__icontains=qm_number)
        if description:
            results = results.filter(description__icontains=description)
        if qm_type:
            results = results.filter(qm_type=qm_type)
        if step_type:
            results = results.filter(step_type=step_type)
        if active_input == 'Yes':
            results = results.filter(active=True)
        elif active_input == 'No':
            results = results.filter(active=False)

    sort_by = request.GET.get('sort', 'qm_number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_mapping = {
        'qm_number': 'qm_number',
        'description': 'description',
        'qm_type': 'qm_type',
        'step_type': 'step_type',
        'plan_count': 'plan_count',
    }

    if clean_sort_key in sort_mapping:
        db_field = sort_mapping[clean_sort_key]
        order_field = f"-{db_field}" if is_descending else db_field
        results = results.order_by(order_field)
    else:
        results = results.order_by('qm_number')

    all_qm_numbers = QualityMaintenanceDocument.objects.values_list('qm_number', flat=True).distinct().order_by('qm_number')
    all_qm_descriptions = QualityMaintenanceDocument.objects.values_list('description', flat=True).distinct().order_by('description')

    return render(request, 'planning/search_qm.html', {
        'search_form': form,
        'results': results,
        'all_qm_numbers': all_qm_numbers,
        'all_qm_descriptions': all_qm_descriptions,
        'sort': sort_by,
    })
    
def search_plan_orders(request):
    work_orders = WorkOrder.objects.select_related('equipment', 'job_status').filter(
        job_status__status_choice='Planning')
    
    sort_by = request.GET.get('sort', '-date_created')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_mapping = {
        'work_order': 'work_order',
        'equipment': 'equipment__Equipment_Number',
        'date_created': 'date_created',
        'plan_start_date': 'plan_start_date',
        'priority': 'priority',
    }

    if clean_sort_key in sort_mapping:
        db_field = sort_mapping[clean_sort_key]
        order_field = f"-{db_field}" if is_descending else db_field
        work_orders = work_orders.order_by(order_field)
    else:
        work_orders = work_orders.order_by('-date_created')

    filter_params = request.GET.copy()
    if 'sort' in filter_params:
        filter_params.pop('sort')
    filter_url = filter_params.urlencode()

    return render(request, 'planning/search_plan_orders.html', {
        'work_orders': work_orders,
        'job_status': 'Planning',
        'sort': sort_by,
        'filter_url': filter_url,
    })
    
def export_plan_wos_excel(request):
    work_orders = WorkOrder.objects.select_related('equipment', 'job_status').filter(
        job_status__status_choice='planning')
    
    sort_by = request.GET.get('sort', '-date_created')
    work_orders = work_orders.order_by(sort_by)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Work Orders In Planning"

    ws.append(['Work Order', 'Eq Num', 'Eq Desc', 'Status'])

    for wo in work_orders:
        ws.append([
            str(wo.work_order), 
            wo.equipment.Equipment_Number,
            wo.equipment.Equipment_Description,
            str(wo.job_status)
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plan_Work_Export.xlsx"'
    wb.save(response)
    return response

def create_qm_work_order_now(request, pk):
    qm = get_object_or_404(QualityMaintenanceDocument, pk=pk)
    work_order, created, message = evaluate_qm_for_work_order_creation(qm)
    
    if created:
        messages.success(request, message)
    else:
        messages.warning(request, message)
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('planning:edit_qm_record', pk=plan.pk)

def forecast_dashboard(request):
    today = timezone.localdate()
    fortnight = today + timedelta(days=14)
    preview_items = []

    if request.GET.get('run_check'):
        active_plans = QualityMaintenancePlan.objects.filter(active=True).select_related('equipment', 'document')
        for plan in active_plans:
            preview = evaluate_plan_for_forecast(plan)
            if preview:
                preview_items.append(preview)
    
    if request.method == "POST" and 'commit_forecast' in request.POST:
        selected_ids = request.POST.getlist('selected_plans')
        created_count = 0
        for item in preview_items:
            wo, created, msg = evaluate_qm_for_work_order_creation(item['plan'])
            if created:
                created_count += 1
        
        messages.success(request, f"Committed {created_count} new Work Orders to Planning.")
        return redirect('planning:planning')

    return render(request, 'planning/forecast.html', {
        'preview_items': preview_items,
        'today': today,
        'fortnight': fortnight
    })

def create_plan(request):
    plan = None
    
    if request.method == 'POST':
        form = QualityMaintenancePlanForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Plan created successfully.")
            return redirect('planning:planning')
    else:
        form = QualityMaintenancePlanForm()

    all_equipment = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')
    all_documents = QualityMaintenanceDocument.objects.all().only('qm_number', 'description').order_by('qm_number')

    current_meters = []
    if plan and plan.equipment:
        current_meters = Meter.objects.filter(equipment=plan.equipment).values_list('meter_type', flat=True).distinct()

    context = {
        'form': form,
        'all_equipment': all_equipment,
        'all_documents': all_documents,
        'current_meters': current_meters,
    }
    return render(request, 'planning/create_plan.html', context)

def edit_plan(request, pk=None):
    plan = None
    forecast = {}
    
    eq_query = request.GET.get('eq_search', '').strip()
    qm_query = request.GET.get('qm_search', '').strip()

    if ' - ' in eq_query:
        eq_query = eq_query.split(' - ')[0].strip()
        
    if ' - ' in qm_query:
        qm_query = qm_query.split(' - ')[0].strip()

    if pk:
        plan = get_object_or_404(QualityMaintenancePlan, pk=pk)
    elif eq_query and qm_query:
        plan = QualityMaintenancePlan.objects.filter(
            equipment__Equipment_Number=eq_query,
            document__qm_number=qm_query
        ).first()

        if not plan:
            messages.error(request, f"No plan found for Unit {eq_query} and Document {qm_query}")


    if request.method == 'POST':
        form = QualityMaintenancePlanForm(request.POST, instance=plan)
        if form.is_valid():
            temp_plan = form.save(commit=False)
            try:
                forecast = temp_plan.get_next_due()
            except Exception as e:
                forecast = {'error': f"Could not calculate forecast: {e}"}
                
            plan = form.save()
            messages.success(request, "Plan saved successfully.")
            return redirect('planning:planning')
    else:
        form = QualityMaintenancePlanForm(instance=plan)

    if plan:
        try:
            forecast = plan.get_next_due()
        except Exception as e:
            forecast = {'error': str(e)}
    elif form.is_bound or (eq_query and qm_query):
        target_eq = Equipment.objects.filter(Equipment_Number=eq_query).first()
        target_doc = QualityMaintenanceDocument.objects.filter(qm_number=qm_query).first()

        if target_eq and target_doc:
            mock_plan = QualityMaintenancePlan(equipment=target_eq, document=target_doc)
            try:
                forecast = mock_plan.get_next_due()
            except Exception:
                forecast = {}

    all_equipment = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')
    all_documents = QualityMaintenanceDocument.objects.all().only('qm_number', 'description').order_by('qm_number')
    
    current_meters = []
    if plan and plan.equipment:
        current_meters = Meter.objects.filter(equipment=plan.equipment).values_list('meter_type', flat=True).distinct()


    context = {
        'form': form,
        'plan': plan,
        'forecast': forecast,
        'all_equipment': all_equipment,
        'all_documents': all_documents,
        'current_meters': current_meters,
        'eq_search_val': request.GET.get('eq_search', ''),
        'qm_search_val': request.GET.get('qm_search', ''),
    }
    return render(request, 'planning/edit_plan.html', context)

def get_linked_meters(request):
    eq_num = request.GET.get('equipment', '').strip()
    meters = []

    if eq_num:
        eq_unit = Equipment.objects.filter(Equipment_Number=eq_num).first()
        if eq_unit:
            current_meters = (
                Meter.objects.filter(equipment=eq_unit)
                .values_list('meter_type', flat=True)
                .distinct()
                .order_by('meter_type')
            )

    html_output = '<datalist id="meter-type-suggestions">'
    if current_meters:
        for mtr in current_meters:
            html_output += f'<option value="{mtr}"></option>'

    else:
        # Stable system fallbacks
        html_output += '<option value="Hours"></option>'
        html_output += '<option value="Kilometers"></option>'
    html_output += '</datalist>'
    
    return HttpResponse(html_output)

def search_plans(request):
    plans = QualityMaintenancePlan.objects.select_related('document', 'equipment').all()
    
    unit_query = request.GET.get('unit_number', '').strip()
    qm_query = request.GET.get('qm_number', '').strip()

    clean_unit = unit_query.split(' - ')[0].strip() if ' - ' in unit_query else unit_query
    clean_qm = qm_query.split(' - ')[0].strip() if ' - ' in qm_query else qm_query
        
    if clean_unit:
        plans = plans.filter(equipment__Equipment_Number__icontains=clean_unit)
    if clean_qm:
        plans = plans.filter(document__qm_number__icontains=clean_qm)

    enriched_results = []
    for p in plans:
        forecast = p.get_next_due()
        enriched_results.append({
            'plan': p,
            'equipment_number': p.equipment.Equipment_Number,
            'eq_desc': p.equipment.Equipment_Description or '',
            'qm_number': p.document.qm_number,
            'doc_desc': p.document.description or '',
            'next_due_date': forecast.get('next_due_date'),
            'next_due_meter': forecast.get('next_due_meter'),
            'active_status': 1 if p.active else 0,
        })

    sort_by = request.GET.get('sort', 'equipment_number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    if enriched_results and hasattr(enriched_results[0], clean_sort_key) or clean_sort_key in ['equipment_number', 'eq_desc', 'qm_number', 'doc_desc', 'next_due_date', 'next_due_meter', 'active_status']:
        
        enriched_results.sort(
            key=lambda x: (x[clean_sort_key] is None, x[clean_sort_key]),
            reverse=is_descending
        )

    all_equipment = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')
    all_documents = QualityMaintenanceDocument.objects.all().only('qm_number', 'description').order_by('qm_number')

    return render(request, 'planning/search_plans.html', {
        'results': enriched_results,
        'all_equipment': all_equipment,
        'all_documents': all_documents,
        'sort': sort_by,
        'unit_query_val': unit_query,
        'qm_query_val': qm_query,
    })

def export_plans_excel(request):
    unit_query = request.GET.get('unit_number', '').strip()
    qm_query = request.GET.get('qm_number', '').strip()
    
    clean_unit = unit_query.split(' - ')[0].strip() if ' - ' in unit_query else unit_query
    clean_qm = qm_query.split(' - ')[0].strip() if ' - ' in qm_query else qm_query
    
    plans = QualityMaintenancePlan.objects.select_related('document', 'equipment').all()
    
    if clean_unit:
        plans = plans.filter(equipment__Equipment_Number__icontains=clean_unit)
    if clean_qm:
        plans = plans.filter(document__qm_number__icontains=clean_qm)
        
    plans = plans.order_by('equipment__Equipment_Number', 'document__qm_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Plans"

    headers = [
        "Equipment Number", "Equipment Description", 
        "QM Document", "Doc Description", 
        "Next Due Date", "Next Due Meter", "Status"
    ]
    ws.append(headers)

    for p in plans:
        forecast = p.get_next_due()
        
        eq_num = p.equipment.Equipment_Number
        eq_desc = p.equipment.Equipment_Description or "---"
        qm_num = p.document.qm_number
        doc_desc = p.document.description or "---"
        
        due_date = forecast.get('next_due_date')
        if due_date and hasattr(due_date, 'strftime'):
            due_date = due_date.strftime('%Y-%m-%d')
        elif due_date:
            due_date = str(due_date)
        else:
            due_date = "---"
            
        due_meter = forecast.get('next_due_meter')
        if due_meter is None:
            due_meter = "---"
            
        status_text = "Active" if p.active else "Paused"

        ws.append([
            eq_num,
            eq_desc,
            qm_num,
            doc_desc,
            due_date,
            due_meter,
            status_text
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Maintenance_Plans_Export.xlsx"'
    wb.save(response)
    
    return response

def export_qm_excel(request):
    form = QualityMaintenanceSearchForm(request.GET or None)

    results = QualityMaintenanceDocument.objects.annotate(
        plan_count=Count('plans')
    ).all()

    if form.is_valid():
        qm_number = form.cleaned_data.get('qm_number')
        description = form.cleaned_data.get('description')
        qm_type = form.cleaned_data.get('qm_type')
        step_type = form.cleaned_data.get('step_type')
        active_input = form.cleaned_data.get('active')

        if qm_number:
            results = results.filter(qm_number__icontains=qm_number)
        if description:
            results = results.filter(description__icontains=description)
        if qm_type:
            results = results.filter(qm_type=qm_type)
        if step_type:
            results = results.filter(step_type=step_type)
        if active_input == 'Yes':
            results = results.filter(active=True)
        elif active_input == 'No':
            results = results.filter(active=False)

    results = results.order_by('qm_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QM Documents"

    headers = ["QM Number", "Description", "Type", "Step Type", "Units Assigned", "Active Status"]
    ws.append(headers)

    for doc in results:
        qm_type_display = doc.get_qm_type_display() if hasattr(doc, 'get_qm_type_display') else str(doc.qm_type)
        step_type_display = doc.get_step_type_display() if hasattr(doc, 'get_step_type_display') else str(doc.step_type)
        status_text = "Active" if doc.active else "Inactive"

        ws.append([
            doc.qm_number,
            doc.description or "---",
            qm_type_display,
            step_type_display,
            doc.plan_count,
            status_text
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="QM_Documents_Export.xlsx"'
    wb.save(response)
    
    return response