from django.shortcuts import render, redirect, get_object_or_404
from .models import InventoryItem
from django.contrib import messages
from django.utils import timezone
from .forms import InventoryItemForm, AlternativeFormSet
from django.http import HttpResponse
import openpyxl
from django.db.models import Sum, Count, Max, Q
from purchasing.models import PurchaseLine, Purchase

def inventory(request):
    return render(request, 'inventory/inventory.html')

def add_inventory_item(request, pk=None):
    instance = get_object_or_404(InventoryItem, pk=pk) if pk else None
    
    if request.method == "POST":
        form = InventoryItemForm(request.POST, instance=instance)
        existing_alts = instance.alternatives.all() if instance else InventoryItem.objects.none()
        formset = AlternativeFormSet(request.POST, queryset=existing_alts, prefix='alt')

        if form.is_valid() and formset.is_valid():
            main_item = form.save()
            new_alternatives = formset.save()
            for alt in new_alternatives:
                main_item.alternatives.add(alt)
            messages.success(request, "Inventory item added successfully!")
            return redirect('inventory:inventory')
        else:
            messages.error(request, "There was an error saving the inventory item(s). Please check the fields below.")
            
    else:
        form = InventoryItemForm(instance=instance)
        existing_alts = instance.alternatives.all() if instance else InventoryItem.objects.none()
        formset = AlternativeFormSet(queryset=existing_alts, prefix='alt')

    return render(request, 'inventory/add_inventory_item.html', {
        'new_inv_form': form,
        'inv_formset': formset,
        'is_edit': bool(pk)
    })
    
def edit_inventory_item(request):
    part_num = request.GET.get('part_number')
    instance = None
    
    if part_num:
        instance = InventoryItem.objects.filter(part_number=part_num).first()
    
    if part_num and not instance:
        messages.error(request, f"Part number '{part_num}' not found.")
        return render(request, 'inventory/edit_inventory_item.html', {
            'new_inv_form': InventoryItemForm(),
            'inv_formset': AlternativeFormSet(queryset=InventoryItem.objects.none(), prefix='alt'),
            'instance': None
        })

    if request.method == "POST":
        part_num_from_post = request.POST.get('part_number') 
        instance = InventoryItem.objects.filter(part_number=part_num_from_post).first()
        form = InventoryItemForm(request.POST, instance=instance)
        existing_alts = instance.alternatives.all() if instance else InventoryItem.objects.none()
        formset = AlternativeFormSet(request.POST, queryset=existing_alts, prefix='alt')

        if form.is_valid() and formset.is_valid():
            main_item = form.save()
            new_alts = formset.save()
            for alt in new_alts:
                main_item.alternatives.add(alt)
            messages.success(request, "Inventory item updated successfully!")    
            return redirect('inventory:inventory')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = InventoryItemForm(instance=instance)
        existing_alts = instance.alternatives.all() if instance else InventoryItem.objects.none()
        formset = AlternativeFormSet(queryset=existing_alts, prefix='alt')

    return render(request, 'inventory/edit_inventory_item.html', {
        'new_inv_form': form,
        'inv_formset': formset,
        'instance': instance,
        'part_number': part_num,
    })

def delete_inventory_group(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)

    if request.method == "POST":
        item.delete()
        messages.success(request, "Part and all associated data deleted.")
        return redirect('inventory:inventory')
    
    return redirect('inventory:edit_inventory_item')

def search_inventory(request):
    part_number = request.GET.get('part_number', '')
    supplier = request.GET.get('supplier', '')
    manufacturer = request.GET.get('manufacturer', '')
    controlled_product = request.GET.get('controlled_product', '')
   
    clean_part = part_number.split(' - ')[0].strip() if ' - ' in part_number else part_number
    inventory = InventoryItem.objects.all().distinct()

    if clean_part:
        inventory = inventory.filter(part_number__icontains=clean_part)
    if supplier:
        inventory = inventory.filter(supplier__supplier_name__icontains=supplier)
    if manufacturer:
        inventory = inventory.filter(manufacturer__icontains=manufacturer)
    if controlled_product:
        inventory = inventory.filter(controlled_product__icontains=controlled_product)

    sort_by = request.GET.get('sort', 'part_number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_mapping = {
        'part_number': 'part_number',
        'part_description': 'part_description',
        'supplier': 'supplier__supplier_name',
        'manufacturer': 'manufacturer',
        'qty': 'qty',
        'uom': 'uom',
        'unit_price': 'unit_price',
        'stock': 'stock',
        'controlled_product': 'controlled_product',
    }

    if clean_sort_key in sort_mapping:
        db_field = sort_mapping[clean_sort_key]
        order_field = f"-{db_field}" if is_descending else db_field
        inventory = inventory.order_by(order_field)
    else:
        inventory = inventory.order_by('part_number')
    
    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()

    all_items = InventoryItem.objects.all().only('part_number', 'part_description').order_by('part_number')
    all_suppliers = InventoryItem.objects.exclude(supplier__supplier_name__isnull=True).values_list('supplier__supplier_name', flat=True).distinct().order_by('supplier__supplier_name')
    all_manufacturers = InventoryItem.objects.exclude(manufacturer__isnull=True).values_list('manufacturer', flat=True).distinct().order_by('manufacturer')
    all_controlled = InventoryItem.objects.exclude(controlled_product__isnull=True).values_list('controlled_product', flat=True).distinct().order_by('controlled_product')

    context = {
        'inventory': inventory,
        'filter_url': filter_url,
        'sort': sort_by,
        'all_items': all_items,
        'all_suppliers': all_suppliers,
        'all_manufacturers': all_manufacturers,
        'all_controlled': all_controlled,
        'part_number_val': part_number,
        'supplier_val': supplier,
        'manufacturer_val': manufacturer,
        'controlled_product_val': controlled_product,
    }

    return render(request, 'inventory/search_inventory.html', context)

def export_inventory_excel(request):
    part_number = request.GET.get('part_number', '')
    supplier = request.GET.get('supplier', '')
    manufacturer = request.GET.get('manufacturer', '')
    controlled_product = request.GET.get('controlled_product', '')
   
    inventory = InventoryItem.objects.all()

    if part_number:
        inventory = inventory.filter(part_number__icontains=part_number)
    if supplier:
        inventory = inventory.filter(supplier__supplier_name__icontains=supplier)
    if manufacturer:
        inventory = inventory.filter(manufacturer__icontains=manufacturer)
    if controlled_product:
        inventory = inventory.filter(controlled_product__icontains=controlled_product)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    columns = ['part_number', 'part_description', 'supplier', 'manufacturer', 'qty', 'uom', 
                'unit_price','stock', 'controlled_product']

    ws.append(columns)

    for inv in inventory:
        supplier_display = inv.supplier.supplier_name if inv.supplier else ""
        ws.append([
            inv.part_number,
            inv.part_description,
            supplier_display,
            inv.manufacturer,
            inv.qty,
            inv.uom,
            inv.unit_price,
            inv.stock,
            inv.controlled_product,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Inventory_Export.xlsx"'
    wb.save(response)
    return response

def export_manage_inventory_excel(request):
    manage_inventory = InventoryItem.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manage_Inventory"

    columns = [
        "Part Number", 
        "Part Description", 
        "Bin Location", 
        "Qty On Hand", 
        "Min Qty", 
        "Max Qty", 
        "UoM", 
        "Last Transaction Date", 
        "Last Transaction Number"]
        
    ws.append(columns)

    for manage in manage_inventory:
        last_date = manage.last_transaction_date.strftime('%Y-%m-%d %H:%M') if manage.last_transaction_date else ""
    
        ws.append([
            str(manage.part_number),
            str(manage.part_description),
            str(manage.bin_location or ""),
            manage.qty_onhand,
            manage.min_qty,
            manage.max_qty,
            str(manage.uom or ""),
            last_date,
            str(manage.last_transaction_number or ""),
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Manage_Inventory_Export.xlsx"'
    wb.save(response)
    return response

def manage_inventory(request):
    if request.method == "POST":

        item_id = request.POST.get('item_id')
        try:
            if item_id:
                item = get_object_or_404(InventoryItem, id=item_id)
            
                item.bin_location = request.POST.get('bin_location')
                item.qty_onhand = int(request.POST.get('qty_onhand') or 0)
                item.min_qty = int(request.POST.get('min_qty') or 0)
                item.max_qty = int(request.POST.get('max_qty') or 0)
                item.uom = request.POST.get('uom')
                
                now = timezone.now()
                item.last_transaction_date = now
                item.last_transaction_number = f"M{now.strftime('%Y%m%d')}"
                
                print(f"DEBUG: Saving Item {item_id} - New Bin: {item.bin_location}")
                item.save()
                messages.success(request, "Inventory item updated successfully!")
                return redirect('inventory:manage_inventory')
            else:
                messages.error(request, "Error: No item ID provided for update.")
        except Exception as e:
            messages.error(request, f"Error updating inventory item: {str(e)}")
            return redirect('inventory:manage_inventory')

    edit_id = request.GET.get('edit')
    edit_item = None
    if edit_id:
        edit_item = get_object_or_404(InventoryItem, id=edit_id)

    sort_by = request.GET.get('sort', 'part_number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_mapping = {
        'part_number': 'part_number',
        'part_description': 'part_description',
        'bin_location': 'bin_location',
        'qty_onhand': 'qty_onhand',
        'min_qty': 'min_qty',
        'max_qty': 'max_qty',
        'uom': 'uom',
        'last_transaction_date': 'last_transaction_date',
        'last_transaction_number': 'last_transaction_number',
    }

    if clean_sort_key in sort_mapping:
        db_field = sort_mapping[clean_sort_key]
        order_field = f"-{db_field}" if is_descending else db_field
        inventory_list = InventoryItem.objects.all().order_by(order_field)
    else:
        inventory_list = InventoryItem.objects.all().order_by('part_number')

    return render(request, 'inventory/manage_inventory.html', {
        'manage_inventory': inventory_list,
        'edit_item': edit_item,
        'sort': sort_by,
    })

def consumption_report(request):
    part_no = request.GET.get('part_number_input', '').strip()
    supplier = request.GET.get('supplier', '').strip()
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    sort_by = request.GET.get('sort', 'part_number_input')
    from_date = request.GET.get('from_date')

    clean_part = part_no.split(' - ')[0].strip() if ' - ' in part_no else part_no

    if from_date in [None, '', 'None']:
        from_date = None
        
    to_date = request.GET.get('to_date')
    if to_date in [None, '', 'None']:
        to_date = None

    lines = PurchaseLine.objects.select_related('purchase', 'supplier').all()

    if part_no:
        lines = lines.filter(part_number_input__icontains=part_no)
    if supplier:
        lines = lines.filter(supplier__supplier_name__icontains=supplier)
    if from_date:
        lines = lines.filter(purchase__date__gte=from_date)
    if to_date:
        lines = lines.filter(purchase__date__lte=to_date)
    
    report_data = lines.values(
        'part_number_input', 
        'manufacturer', 
        'part_description', 
        'uom', 
        'supplier__supplier_name',
        'inventory_item__unit_price'

    ).annotate(
        total_qty=Sum('qty'),
        total_spent=Sum('total_price'),
        order_count=Count('id'),
        last_purch_date=Max('purchase__date')
    )

    sort_mapping = {
        'part_number_input': 'part_number_input',
        '-part_number_input': '-part_number_input',
        'manufacturer': 'manufacturer',
        '-manufacturer': '-manufacturer',
        'supplier': 'supplier',
        '-supplier': '-supplier',
        'qty': 'total_qty',
        '-qty': '-total_qty',
        'unit_price': 'inventory_item__unit_price',
        '-unit_price': '-inventory_item__unit_price',
        'spent': 'total_spent',
        '-spent': '-total_spent',
        'order_count': 'order_count',
        '-order_count': '-order_count',
        'last_purch_date': 'last_purch_date',
        '-last_purch_date': '-last_purch_date',
    }
    final_sort = sort_mapping.get(sort_by, 'part_number_input')
    report_data = report_data.order_by(final_sort)

    seen_parts = set()
    part_options = []
    for item in report_data:
        p_num = item.get('part_number_input')
        if p_num and p_num not in seen_parts:
            seen_parts.add(p_num)
            part_options.append({
                'part_number_input': p_num,
                'part_description': item.get('part_description') or 'No Description'
            })

    part_options.sort(key=lambda x: x['part_number_input'])
    supp_options = PurchaseLine.objects.values_list('supplier__supplier_name', flat=True).distinct().order_by('supplier__supplier_name')

    context = {
        'report_data': report_data,
        'part_number_input': part_no,
        'supplier': supplier,
        'from_date': from_date,
        'to_date': to_date,
        'sort': sort_by,
        'part_number_options': part_options,
        'supplier_options': supp_options,
    }
    return render(request, 'inventory/consumption.html', context)

def export_consumption_excel(request):
    part_no = request.GET.get('part_number_input', '').strip()
    supplier = request.GET.get('supplier', '').strip()
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    lines = PurchaseLine.objects.select_related('purchase', 'supplier').all()

    if part_no:
        lines = lines.filter(part_number_input__icontains=part_no)
    if supplier:
        lines = lines.filter(supplier__supplier_name__icontains=supplier)
    if from_date and from_date != 'None':
        lines = lines.filter(purchase__date__gte=from_date)
    if to_date and to_date != 'None':
        lines = lines.filter(purchase__date__lte=to_date)

    report_data = lines.values(
        'part_number_input', 'manufacturer', 'part_description', 
        'uom', 'supplier__supplier_name', 'inventory_item__unit_price'
    ).annotate(
        total_qty=Sum('qty'),
        total_spent=Sum('total_price'),
        order_count=Count('id'),
        last_purch_date=Max('purchase__date')
    ).order_by('part_number_input')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parts Consumption"

    headers = ['Part Number', 'Manufacturer', 'Description', 'Supplier', 'Qty', 'UOM', 'Unit Price', 'Total Spent', 'Annual Count', 'Last Purchased']
    ws.append(headers)

    for item in report_data:
        ws.append([
            item['part_number_input'],
            item['manufacturer'],
            item['part_description'],
            item['supplier__supplier_name'],
            item['total_qty'],
            item['uom'],
            float(item['inventory_item__unit_price'] or 0),
            float(item['total_spent'] or 0),
            item['order_count'],
            item['last_purch_date'].strftime('%Y-%m-%d') if item['last_purch_date'] else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Consumption_Report.xlsx"'
    wb.save(response)
    return response