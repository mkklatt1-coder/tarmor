from django.shortcuts import render, redirect, get_object_or_404
from django.forms import modelformset_factory, inlineformset_factory, BaseInlineFormSet
from django.views.decorators.http import require_POST
from .models import (ShortTermCM, MagPlug, FilterRating, ValveSet, ValveSetReading, CylinderTemp, CylinderTempReading, BucketLip, LipMeasurement, BoxLiner, RimInspection,
                     CycleTime, CycleTimeMeasurement, TireInformation, TireFailure, TireChangeInfo, TireChange, TireInspection, TireInspectionReading)
from .forms import (ActionItemForm, MagPlugFormSet, MagPlugSearchForm, MagPlugForm, FilterRatingFormSet, CylinderTempForm, CylinderTempReadingFormSet,
                    FilterRatingForm, FilterRatingSearchForm, ValveSetReadingFormSet, ValveSetForm, BucketLipForm, LipMeasurementFormSet, TireFailureTypeForm,
                    BoxLinerForm, LinerMeasurementForm, LinerMeasurementFormSet, CycleTimeForm, CycleTimeMeasurementFormSet, TireInformationForm, TireChangeInfoForm, 
                    TireInformationFormSet, TireChangeInfoFormSet, TireChangeForm, TireInspectionForm, TireInspectionReadingForm, TireInspectionReadingFormSet,
                    RimInspectionForm, RimInspectionFormSet)
from equipment.models import Equipment, Meter, EQ_Type
from purchasing.models import Purchase
from work_orders.models import WorkOrder
from django.http import JsonResponse, HttpResponse
from work_orders.models import WorkOrder
from django.contrib import messages
import openpyxl, json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import Avg, Q, Min, Max, Sum
from django.db.models.functions import TruncMonth
from django.db.models import Prefetch
from datetime import datetime, date
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from decimal import Decimal

def condition_monitoring(request):
    return render(request, 'condition_monitoring/condition_monitoring.html', {'content_type': 'text/html'})

def mag_plugs(request):
    return render(request, 'condition_monitoring/mag_plugs.html', {'content_type': 'text/html'})

def filters(request):
    return render(request, 'condition_monitoring/filters.html', {'content_type': 'text/html'})

def valve_settings(request):
    return render(request, 'condition_monitoring/valve.settings.html', {'content_type': 'text/html'})

def cylinder_temps(request):
    return render(request, 'condition_monitoring/cylinder_temps.html', {'content_type': 'text/html'})

def bucket_lips(request):
    return render(request, 'condition_monitoring/bucket_lips.html', {'content_type': 'text/html'})

def box_liners(request):
    return render(request, 'condition_monitoring/box_liners.html', {'content_type': 'text/html'})

def cycle_times(request):
    return render(request, 'condition_monitoring/cycle_times.html', {'content_type': 'text/html'})

def tires(request):
    return render(request, 'condition_monitoring/tires.html', {'content_type': 'text/html'})

def tracker_view(request):
    ActionItemFormSet = modelformset_factory(
        ShortTermCM, 
        form=ActionItemForm, 
        extra=1, 
        can_delete=True
    )
    
    if request.method == 'POST':
        formset = ActionItemFormSet(request.POST)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Action items updated successfully.')
            return redirect('condition_monitoring:condition_monitoring')
    else:
        formset = ActionItemFormSet(queryset=ShortTermCM.objects.all().order_by('-date'))
        
    return render(request, 'condition_monitoring/tracker.html', {'formset': formset})

def get_equipment_details(request, pk):
    equipment = Equipment.objects.filter(pk=pk).values('Equipment_Description').first()
    if not equipment:
        return JsonResponse({'Equipment_Description': ''})
    return JsonResponse(equipment)

def get_workorder_details(request, pk):
    wo = WorkOrder.objects.filter(pk=pk).values('troubleshoot_description', 'repair_description').first()
    if not wo:
        return JsonResponse({'troubleshoot_description': '', 'repair_description': ''})
    return JsonResponse(wo)

def add_mag_plugs(request):
    if request.method == 'POST':
        formset = MagPlugFormSet(request.POST)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Mag Plug entries updated successfully.')
            return redirect('condition_monitoring:mag_plugs')
    else:
        formset = MagPlugFormSet(queryset=MagPlug.objects.none())
        
    return render(request, 'condition_monitoring/add_mag_plugs.html', {'formset': formset})

def get_meters_for_equipment(request):
    """API endpoint to fetch meters belonging to an equipment ID."""
    equipment_id = request.GET.get('equipment_id')
    
    meters = Meter.objects.filter(equipment_id=equipment_id).values('id', 'name')
    
    return JsonResponse(list(meters), safe=False)

def search_mag_plugs(request):
    search_form = MagPlugSearchForm(request.GET or None)
    queryset = MagPlug.objects.all().select_related('equipment', 'work_order', 'meter')

    if search_form.is_valid():
        if search_form.cleaned_data.get('date'):
            queryset = queryset.filter(date=search_form.cleaned_data['date'])
            
        eq_input = search_form.cleaned_data.get('equipment')
        if eq_input and ' - ' in eq_input:
            eq_num = eq_input.split(' - ')[0].strip()
            queryset = queryset.filter(equipment__Equipment_Number=eq_num)
            
        wo_input = search_form.cleaned_data.get('work_order')
        if wo_input and ' - ' in wo_input:
            wo_num = wo_input.split(' - ')[0].strip()
            queryset = queryset.filter(work_order__work_order=wo_num)
            
        if search_form.cleaned_data.get('compartment'):
            queryset = queryset.filter(compartment=search_form.cleaned_data['compartment'])

    sort_by = request.GET.get('sort', '-date')
    allowed_sort_fields = [
        'date', '-date', 
        'equipment__Equipment_Number', '-equipment__Equipment_Number',
        'work_order__work_order', '-work_order__work_order',
        'compartment', '-compartment', 
        'meter__meter_type', '-meter__meter_type',
        'meter_reading', '-meter_reading', 
        'plug_rating', '-plug_rating'
    ]
    if sort_by in allowed_sort_fields:
        queryset = queryset.order_by(sort_by)

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()

    context = {
        'search_form': search_form,
        'records': queryset,
        'sort': sort_by,
        'filter_url': filter_url,
        'all_equipment': Equipment.objects.all(),
        'all_work_orders': WorkOrder.objects.all(),
    }
    return render(request, 'condition_monitoring/search_mag_plugs.html', context)

def edit_mag_plugs(request, pk):
    record = get_object_or_404(MagPlug, pk=pk)
    
    MagPlugEditFormSet = modelformset_factory(MagPlug, form=MagPlugForm, extra=0)
    
    queryset = MagPlug.objects.filter(pk=pk)

    if request.method == 'POST':
        formset = MagPlugEditFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, f"Magnetic plug entry updated successfully.")
            return redirect('condition_monitoring:search_mag_plugs')
    else:
        formset = MagPlugEditFormSet(queryset=queryset)
        
    return render(request, 'condition_monitoring/edit_mag_plugs.html', {'formset': formset})

def export_mag_plugs_excel(request):
    search_form = MagPlugSearchForm(request.GET or None)
    queryset = MagPlug.objects.all().select_related('equipment', 'work_order', 'meter')

    if search_form.is_valid():
        if search_form.cleaned_data.get('date'):
            queryset = queryset.filter(date=search_form.cleaned_data['date'])
            
        eq_input = search_form.cleaned_data.get('equipment')
        if eq_input and ' - ' in eq_input:
            eq_num = eq_input.split(' - ')[0].strip()
            queryset = queryset.filter(equipment__Equipment_Number=eq_num)
            
        wo_input = search_form.cleaned_data.get('work_order')
        if wo_input and ' - ' in wo_input:
            wo_num = wo_input.split(' - ')[0].strip()
            queryset = queryset.filter(work_order__work_order=wo_num)
            
        if search_form.cleaned_data.get('compartment'):
            queryset = queryset.filter(compartment=search_form.cleaned_data['compartment'])

    sort_by = request.GET.get('sort', '-date')
    queryset = queryset.order_by(sort_by)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mag Plug Ratings"

    headers = [
        'Date', 'Equipment Number', 'Equipment Description', 
        'Work Order', 'Compartment', 'Meter Type', 
        'Meter Reading', 'Plug Rating', 'Comments'
    ]
    ws.append(headers)

    bold_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = bold_font

    for record in queryset:
        ws.append([
            record.date.strftime('%Y-%m-%d') if record.date else '',
            record.equipment.Equipment_Number if record.equipment else '',
            record.equipment.Equipment_Description if record.equipment else '',
            record.work_order.work_order if record.work_order else '',
            record.get_compartment_display() if record.compartment else '',
            record.meter.meter_type if record.meter else '',
            record.meter_reading if record.meter_reading is not None else '',
            record.get_plug_rating_display() if record.plug_rating else '',
            record.comments or ''
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="magnetic_plug_ratings.xlsx"'
    
    wb.save(response)
    return response

def mag_plug_trends(request):
    today = datetime.today()
    default_start = f"{today.year}-01-01"
    default_end = today.strftime('%Y-%m-%d')

    filters = {
        'start_date': request.GET.get('start_date', default_start),
        'end_date': request.GET.get('end_date', default_end),
        'equipment': request.GET.get('equipment', ''),
        'work_order': request.GET.get('work_order', ''),
        'compartment': request.GET.get('compartment', ''),
    }

    queryset = MagPlug.objects.all()

    if filters['start_date']:
        queryset = queryset.filter(date__gte=filters['start_date'])
    if filters['end_date']:
        queryset = queryset.filter(date__lte=filters['end_date'])
        
    if filters['equipment'] and ' - ' in filters['equipment']:
        eq_num = filters['equipment'].split(' - ')[0].strip()
        queryset = queryset.filter(equipment__Equipment_Number=eq_num)
    elif filters['equipment']:
        queryset = queryset.filter(equipment__Equipment_Number__icontains=filters['equipment'])
        
    if filters['work_order'] and ' - ' in filters['work_order']:
        wo_num = filters['work_order'].split(' - ')[0].strip()
        queryset = queryset.filter(work_order__work_order=wo_num)
    elif filters['work_order']:
        queryset = queryset.filter(work_order__work_order__icontains=filters['work_order'])
        
    if filters['compartment']:
        queryset = queryset.filter(compartment=filters['compartment'])

    trend_data = (
        queryset.annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(avg_rating=Avg('plug_rating'))
        .order_by('month')
    )

    raw_data_dict = {}
    for entry in trend_data:
        if entry['month'] and entry['avg_rating'] is not None:
            month_str = entry['month'].strftime('%Y-%m')
            raw_data_dict[month_str] = round(float(entry['avg_rating']), 2)

    context = {
        'filters': filters,
        'raw_data_dict': raw_data_dict,
        'current_year': today.year,
        'all_equipment': Equipment.objects.all(),
        'all_work_orders': WorkOrder.objects.all(),
        'compartment_choices': MagPlug._meta.get_field('compartment').choices,
    }
    return render(request, 'condition_monitoring/mag_plug_trends.html', context)

def add_filters(request):
    if request.method == 'POST':
        formset = FilterRatingFormSet(request.POST)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Filter rating entries updated successfully.')
            return redirect('condition_monitoring:filters')
    else:
        formset = FilterRatingFormSet(queryset=FilterRating.objects.none())
        
    return render(request, 'condition_monitoring/add_filters.html', {'formset': formset})

def search_filters(request):
    search_form = FilterRatingSearchForm(request.GET or None)
    queryset = FilterRating.objects.all().select_related('equipment', 'work_order', 'meter')

    if search_form.is_valid():
        if search_form.cleaned_data.get('date'):
            queryset = queryset.filter(date=search_form.cleaned_data['date'])
            
        eq_input = search_form.cleaned_data.get('equipment')
        if eq_input and ' - ' in eq_input:
            eq_num = eq_input.split(' - ')[0].strip()
            queryset = queryset.filter(equipment__Equipment_Number=eq_num)
            
        wo_input = search_form.cleaned_data.get('work_order')
        if wo_input and ' - ' in wo_input:
            wo_num = wo_input.split(' - ')[0].strip()
            queryset = queryset.filter(work_order__work_order=wo_num)
            
        if search_form.cleaned_data.get('compartment'):
            queryset = queryset.filter(compartment=search_form.cleaned_data['compartment'])

    sort_by = request.GET.get('sort', '-date')
    allowed_sort_fields = [
        'date', '-date', 
        'equipment__Equipment_Number', '-equipment__Equipment_Number',
        'work_order__work_order', '-work_order__work_order',
        'compartment', '-compartment', 
        'meter__meter_type', '-meter__meter_type',
        'meter_reading', '-meter_reading', 
        'filter_rating', '-filter_rating'
    ]
    if sort_by in allowed_sort_fields:
        queryset = queryset.order_by(sort_by)

    params = request.GET.copy()
    if 'sort' in params:
        del params['sort']
    filter_url = params.urlencode()

    context = {
        'search_form': search_form,
        'records': queryset,
        'sort': sort_by,
        'filter_url': filter_url,
        'all_equipment': Equipment.objects.all(),
        'all_work_orders': WorkOrder.objects.all(),
    }
    return render(request, 'condition_monitoring/search_filters.html', context)

def edit_filters(request, pk):
    record = get_object_or_404(FilterRating, pk=pk)

    FilterRatingEditFormSet = modelformset_factory(FilterRating, form=FilterRatingForm, extra=0)

    queryset = FilterRating.objects.filter(pk=pk)

    if request.method == 'POST':
        formset = FilterRatingEditFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, f"Filter rating entry updated successfully.")
            return redirect('condition_monitoring:search_filters')
    else:
        formset = FilterRatingEditFormSet(queryset=queryset)
        
    return render(request, 'condition_monitoring/edit_filters.html', {'formset': formset})

def export_filter_ratings_excel(request):
    search_form = FilterRatingSearchForm(request.GET or None)
    queryset = FilterRating.objects.all().select_related('equipment', 'work_order', 'meter')

    if search_form.is_valid():
        if search_form.cleaned_data.get('date'):
            queryset = queryset.filter(date=search_form.cleaned_data['date'])
            
        eq_input = search_form.cleaned_data.get('equipment')
        if eq_input and ' - ' in eq_input:
            eq_num = eq_input.split(' - ')[0].strip()
            queryset = queryset.filter(equipment__Equipment_Number=eq_num)
            
        wo_input = search_form.cleaned_data.get('work_order')
        if wo_input and ' - ' in wo_input:
            wo_num = wo_input.split(' - ')[0].strip()
            queryset = queryset.filter(work_order__work_order=wo_num)
            
        if search_form.cleaned_data.get('compartment'):
            queryset = queryset.filter(compartment=search_form.cleaned_data['compartment'])

    sort_by = request.GET.get('sort', '-date')
    queryset = queryset.order_by(sort_by)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filter Ratings"

    headers = [
        'Date', 'Equipment Number', 'Equipment Description', 
        'Work Order', 'Compartment', 'Meter Type', 
        'Meter Reading', 'Filter Rating', 'Comments'
    ]
    ws.append(headers)

    bold_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = bold_font

    for record in queryset:
        ws.append([
            record.date.strftime('%Y-%m-%d') if record.date else '',
            record.equipment.Equipment_Number if record.equipment else '',
            record.equipment.Equipment_Description if record.equipment else '',
            record.work_order.work_order if record.work_order else '',
            record.get_compartment_display() if record.compartment else '',
            record.meter.meter_type if record.meter else '',
            record.meter_reading if record.meter_reading is not None else '',
            record.get_filter_rating_display() if record.filter_rating else '',
            record.comments or ''
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="filter_ratings.xlsx"'
    
    wb.save(response)
    return response

def filter_rating_trends(request):
    today = datetime.today()
    default_start = f"{today.year}-01-01"
    default_end = today.strftime('%Y-%m-%d')

    filters = {
        'start_date': request.GET.get('start_date', default_start),
        'end_date': request.GET.get('end_date', default_end),
        'equipment': request.GET.get('equipment', ''),
        'work_order': request.GET.get('work_order', ''),
        'compartment': request.GET.get('compartment', ''),
    }

    queryset = FilterRating.objects.all()

    if filters['start_date']:
        queryset = queryset.filter(date__gte=filters['start_date'])
    if filters['end_date']:
        queryset = queryset.filter(date__lte=filters['end_date'])
        
    if filters['equipment'] and ' - ' in filters['equipment']:
        eq_num = filters['equipment'].split(' - ')[0].strip()
        queryset = queryset.filter(equipment__Equipment_Number=eq_num)
    elif filters['equipment']:
        queryset = queryset.filter(equipment__Equipment_Number__icontains=filters['equipment'])
        
    if filters['work_order'] and ' - ' in filters['work_order']:
        wo_num = filters['work_order'].split(' - ')[0].strip()
        queryset = queryset.filter(work_order__work_order=wo_num)
    elif filters['work_order']:
        queryset = queryset.filter(work_order__work_order__icontains=filters['work_order'])
        
    if filters['compartment']:
        queryset = queryset.filter(compartment=filters['compartment'])

    trend_data = (
        queryset.annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(avg_rating=Avg('filter_rating'))
        .order_by('month')
    )

    raw_data_dict = {}
    for entry in trend_data:
        if entry['month'] and entry['avg_rating'] is not None:
            month_str = entry['month'].strftime('%Y-%m')
            raw_data_dict[month_str] = round(float(entry['avg_rating']), 2)

    context = {
        'filters': filters,
        'raw_data_dict': raw_data_dict,
        'current_year': today.year,
        'all_equipment': Equipment.objects.all(),
        'all_work_orders': WorkOrder.objects.all(),
        'compartment_choices': FilterRating._meta.get_field('compartment').choices,
    }
    return render(request, 'condition_monitoring/filter_trends.html', context)

def add_valve_sets(request):
    if request.method == 'POST':
        form = ValveSetForm(request.POST)
        formset = ValveSetReadingFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            valve_set = form.save()
            formset.instance = valve_set
            formset.save()
            messages.success(request, 'Valve set created successfully.')
            return redirect('condition_monitoring:valve_settings')
    else:
        form = ValveSetForm()
        formset = ValveSetReadingFormSet()
    return render(request, 'condition_monitoring/add_valve_sets.html', {
        'form': form,
        'formset': formset,
    })

def edit_valve_sets(request, pk):
    valve_set = get_object_or_404(ValveSet, pk=pk)
    if request.method == 'POST':
        form = ValveSetForm(request.POST, instance=valve_set)
        formset = ValveSetReadingFormSet(request.POST, instance=valve_set)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Valve set updated successfully.')
            return redirect('condition_monitoring:search_valve_sets')
    else:
        form = ValveSetForm(instance=valve_set)
        formset = ValveSetReadingFormSet(instance=valve_set)
    return render(request, 'condition_monitoring/edit_valve_sets.html', {
        'form': form,
        'formset': formset,
        'valve_set': valve_set,
    })

@require_POST
def delete_valve_set(request, pk):
    valve_set = get_object_or_404(ValveSet, pk=pk)
    valve_set.delete()
    messages.success(request, 'Valve set deleted successfully.')
    return redirect('condition_monitoring:valve_settings')

def search_valve_sets(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    valve_sets = ValveSet.objects.select_related(
        'equipment',
        'work_order',
        'meter'
    ).prefetch_related(
        'readings'
    )
    if eq_num_val:
        valve_sets = valve_sets.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    
    if wo_num_val:
        valve_sets = valve_sets.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
        
    if selected_meter_type:
        valve_sets = valve_sets.filter(meter__meter_type=selected_meter_type)

    sort_mapping = {
        'date': 'date',
        '-date': '-date',
        'equipment_number': 'equipment__Equipment_Number',
        '-equipment_number': '-equipment__Equipment_Number',
        'equipment_desc': 'equipment__Equipment_Description',
        '-equipment_desc': '-equipment__Equipment_Description',
        'work_order': 'work_order__work_order',
        '-work_order': '-work_order__work_order',
        'wo_desc': 'work_order__troubleshoot_description',
        '-wo_desc': '-work_order__troubleshoot_description',
        'meter': 'meter__meter_type',
        '-meter': '-meter__meter_type',
        'meter_reading': 'meter_reading',
        '-meter_reading': '-meter_reading',
        'valve_setting': 'readings__valve_setting',
        '-valve_setting': '-readings__valve_setting',
    }
    
    db_sort_field = sort_mapping.get(sort_by, '-date')
    valve_sets = valve_sets.order_by(db_sort_field).distinct()
    meter_type_choices = [choice[0] for choice in Meter._meta.get_field('meter_type').choices]

    return render(request, 'condition_monitoring/search_valve_sets.html', {
        'valve_sets': valve_sets,
        'meter_choices': meter_type_choices,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'selected_meter': selected_meter_type,
        'sort': sort_by,
    })

def valve_sets_trend(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    sort_by = request.GET.get('sort', 'date')

    today = date.today()
    if not start_date_str:
        start_date = date(today.year, 1, 1)
        start_date_str = start_date.strftime('%Y-%m-%d')
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
        end_date_str = end_date.strftime('%Y-%m-%d')
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    readings = ValveSetReading.objects.select_related(
        'valve_set',
        'valve_set__equipment',
        'valve_set__work_order'
    ).filter(
        valve_set__date__gte=start_date,
        valve_set__date__lte=end_date
    )

    if eq_num_val:
        readings = readings.filter(
            Q(valve_set__equipment__Equipment_Number__icontains=eq_num_val) |
            Q(valve_set__equipment__Equipment_Description__icontains=eq_num_val)
        )
    if wo_num_val:
        readings = readings.filter(
            Q(valve_set__work_order__work_order__icontains=wo_num_val) |
            Q(valve_set__work_order__troubleshoot_description__icontains=wo_num_val)
        )

    sort_mapping = {
        'date': 'valve_set__date',
        '-date': '-valve_set__date',
        'equipment_number': 'valve_set__equipment__Equipment_Number',
        '-equipment_number': '-valve_set__equipment__Equipment_Number',
        'work_order': 'valve_set__work_order__work_order',
        '-work_order': '-valve_set__work_order__work_order',
        'valve_setting': 'valve_setting',
        '-valve_setting': '-valve_setting',
    }
    db_sort_field = sort_mapping.get(sort_by, 'valve_set__date')
    readings = readings.order_by(db_sort_field).distinct()

    chart_months = []
    current_bucket = date(start_date.year, start_date.month, 1)
    target_end_bucket = date(end_date.year, end_date.month, 1)

    while current_bucket <= target_end_bucket:
        chart_months.append(current_bucket.strftime('%Y-%m'))
        current_bucket += relativedelta(months=1)

    valve_monthly_values = defaultdict(lambda: defaultdict(list))
    for r in readings:
        if r.valve_set.date:
            month_str = r.valve_set.date.strftime('%Y-%m')
            valve_label = f"Cyl #{r.cylinder_number} {r.int_exh} V{r.valve_number}"
            if r.valve_setting is not None:
                valve_monthly_values[valve_label][month_str].append(float(r.valve_setting))

    chart_datasets = []
    colors = ['#FF6384', '#36A2EB', '#4BC0C0', '#FFCE56', '#9966FF', '#FF9F40', '#32CD32', '#008080']

    for index, (valve_label, monthly_map) in enumerate(valve_monthly_values.items()):
        data_points = []
        for m_str in chart_months:
            vals = monthly_map.get(m_str, [])
            data_points.append(sum(vals) / len(vals) if vals else None)
        
        color = colors[index % len(colors)]
        chart_datasets.append({
            'label': valve_label,
            'data': data_points,
            'borderColor': color,
            'backgroundColor': color,
            'fill': False,
            'tension': 0.1,
            'spanGaps': True
        })

    return render(request, 'condition_monitoring/valve_sets_trend.html', {
        'readings': readings,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'sort': sort_by,
        'labels_json': json.dumps(chart_months),
        'datasets_json': json.dumps(chart_datasets),
    })

def export_valve_sets_excel(request):
    equipment_id = request.GET.get('equipment')
    work_order_id = request.GET.get('work_order')
    meter_id = request.GET.get('meter')
    valve_sets = ValveSet.objects.select_related(
        'equipment',
        'work_order',
        'meter'
    ).prefetch_related(
        'readings'
    )
    if equipment_id:
        valve_sets = valve_sets.filter(equipment_id=equipment_id)
    if work_order_id:
        valve_sets = valve_sets.filter(work_order_id=work_order_id)
    if meter_id:
        valve_sets = valve_sets.filter(meter_id=meter_id)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Valve Sets'
    headers = [
        'Date',
        'Equipment',
        'Equipment Description',
        'Work Order',
        'WO Troubleshoot Description',
        'Meter',
        'Meter Type',
        'Meter Reading',
        'Cylinder #',
        'Int/Exh',
        'Valve #',
        'Valve Setting',
        'Comments',
    ]
    ws.append(headers)
    header_fill = PatternFill(
        start_color='D9EAF7',
        end_color='D9EAF7',
        fill_type='solid'
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for valve_set in valve_sets:
        readings = valve_set.readings.all()
        if readings:
            for reading in readings:
                ws.append([
                    valve_set.date,
                    str(valve_set.equipment) if valve_set.equipment else '',
                    getattr(valve_set.equipment, 'description', '') if valve_set.equipment else '',
                    str(valve_set.work_order) if valve_set.work_order else '',
                    getattr(valve_set.work_order, 'troubleshoot_description', '') if valve_set.work_order else '',
                    str(valve_set.meter) if valve_set.meter else '',
                    getattr(valve_set.meter, 'meter_type', '') if valve_set.meter else '',
                    valve_set.meter_reading if valve_set.meter_reading is not None else '',
                    reading.cylinder_number or '',
                    reading.int_exh or '',
                    reading.valve_number or '',
                    float(reading.valve_setting) if reading.valve_setting is not None else '',
                    valve_set.comments or '',
                ])
        else:
            ws.append([
                valve_set.date,
                str(valve_set.equipment) if valve_set.equipment else '',
                getattr(valve_set.equipment, 'description', '') if valve_set.equipment else '',
                str(valve_set.work_order) if valve_set.work_order else '',
                getattr(valve_set.work_order, 'troubleshoot_description', '') if valve_set.work_order else '',
                str(valve_set.meter) if valve_set.meter else '',
                getattr(valve_set.meter, 'meter_type', '') if valve_set.meter else '',
                valve_set.meter_reading if valve_set.meter_reading is not None else '',
                '',
                '',
                '',
                '',
                valve_set.comments or '',
            ])
    for column_cells in ws.columns:
        max_length = 0
        column_number = column_cells[0].column
        column_letter = get_column_letter(column_number)
        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    ws.freeze_panes = 'A2'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="valve_sets.xlsx"'
    wb.save(response)
    return response

def add_cylinder_temps(request):
    if request.method == 'POST':
        form = CylinderTempForm(request.POST)
        formset = CylinderTempReadingFormSet(request.POST, prefix='readings')
        if form.is_valid() and formset.is_valid():
            cylinder_temp = form.save()
            formset.instance = cylinder_temp
            formset.save()
            messages.success(request, 'Cylinder temps added successfully.')
            return redirect('condition_monitoring:cylinder_temps')
    else:
        form = CylinderTempForm()
        formset = CylinderTempReadingFormSet(prefix='readings')
        
    return render(request, 'condition_monitoring/add_cylinder_temps.html', {
        'form': form,
        'formset': formset,
    })

def search_cylinder_temps(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    cyl_temps = CylinderTemp.objects.select_related(
        'equipment',
        'work_order',
        'meter'
    ).prefetch_related(
        'cyl_temp_readings'
    )

    if eq_num_val:
        cyl_temps = cyl_temps.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    
    if wo_num_val:
        cyl_temps = cyl_temps.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
        
    if selected_meter_type:
        cyl_temps = cyl_temps.filter(meter__meter_type=selected_meter_type)

    sort_mapping = {
        'date': 'date',
        '-date': '-date',
        'equipment_number': 'equipment__Equipment_Number',
        '-equipment_number': '-equipment__Equipment_Number',
        'equipment_desc': 'equipment__Equipment_Description',
        '-equipment_desc': '-equipment__Equipment_Description',
        'work_order': 'work_order__work_order',
        '-work_order': '-work_order__work_order',
        'wo_desc': 'work_order__troubleshoot_description',
        '-wo_desc': '-work_order__troubleshoot_description',
        'meter': 'meter__meter_type',
        '-meter': '-meter__meter_type',
        'meter_reading': 'meter_reading',
        '-meter_reading': '-meter_reading',
        'temp_reading': 'cyl_temp_readings__temp_reading',
        '-temp_reading': '-cyl_temp_readings__temp_reading',
    }
    
    db_sort_field = sort_mapping.get(sort_by, '-date')
    cyl_temps = cyl_temps.order_by(db_sort_field).distinct()
    meter_type_choices = [choice[0] for choice in Meter._meta.get_field('meter_type').choices]

    return render(request, 'condition_monitoring/search_cylinder_temps.html', {
        'cyl_temps': cyl_temps,
        'meter_choices': meter_type_choices,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'selected_meter': selected_meter_type,
        'sort': sort_by,
    })

def edit_cylinder_temps(request, pk):
    cyl_temp = get_object_or_404(CylinderTemp, pk=pk)
    if request.method == 'POST':
        form = CylinderTempForm(request.POST, instance=cyl_temp)
        formset = CylinderTempReadingFormSet(request.POST, instance=cyl_temp)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Cylinder temps updated successfully.')
            return redirect('condition_monitoring:search_cylinder_temps')
    else:
        form = CylinderTempForm(instance=cyl_temp)
        formset = CylinderTempReadingFormSet(instance=cyl_temp)
    return render(request, 'condition_monitoring/edit_cylinder_temps.html', {
        'form': form,
        'formset': formset,
        'cyl_temp': cyl_temp,
    })

@require_POST
def delete_cylinder_temps(request, pk):
    cyl_temp = get_object_or_404(CylinderTemp, pk=pk)
    cyl_temp.delete()
    messages.success(request, 'Cylinder temps deleted successfully.')
    return redirect('condition_monitoring:search_cylinder_temps')

def export_cylinder_temps_excel(request):
    equipment_id = request.GET.get('equipment')
    work_order_id = request.GET.get('work_order')
    meter_id = request.GET.get('meter')

    cyl_temps = CylinderTemp.objects.select_related(
        'equipment', 'work_order', 'meter'
    ).prefetch_related(
        'cyl_temp_readings'
    )

    if equipment_id:
        cyl_temps = cyl_temps.filter(equipment_id=equipment_id)
    if work_order_id:
        cyl_temps = cyl_temps.filter(work_order_id=work_order_id)
    if meter_id:
        cyl_temps = cyl_temps.filter(meter_id=meter_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cylinder Temperatures"

    headers = [
        'Date', 'Equipment #', 'Equipment Description', 
        'Work Order', 'WO Troubleshoot Desc', 'Meter Type', 
        'Meter Reading', 'Cyl #', 'Temp Reading', 'UOM', 'Comments'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for temp in cyl_temps:
        readings = temp.cyl_temp_readings.all()
        
        if readings.exists():
            for reading in readings:
                ws.append([
                    temp.date.strftime('%Y-%m-%d') if temp.date else '',
                    str(temp.equipment.Equipment_Number) if temp.equipment else '',
                    temp.equipment.Equipment_Description if temp.equipment else '',
                    str(temp.work_order.work_order) if temp.work_order else '',
                    temp.work_order.troubleshoot_description if temp.work_order else '',
                    temp.meter.meter_type if temp.meter else '',
                    temp.meter_reading if temp.meter_reading is not None else '',
                    reading.cylinder_number if reading.cylinder_number is not None else '',
                    reading.temp_reading if reading.temp_reading is not None else '',
                    reading.uom or '',
                    temp.comments or ''
                ])
        else:
            ws.append([
                temp.date.strftime('%Y-%m-%d') if temp.date else '',
                str(temp.equipment.Equipment_Number) if temp.equipment else '',
                temp.equipment.Equipment_Description if temp.equipment else '',
                str(temp.work_order.work_order) if temp.work_order else '',
                temp.work_order.troubleshoot_description if temp.work_order else '',
                temp.meter.meter_type if temp.meter else '',
                temp.meter_reading if temp.meter_reading is not None else '',
                'No readings', '', '',
                temp.comments or ''
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cylinder_temperatures.xlsx"'
    wb.save(response)
    
    return response

def cylinder_temps_trend(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    sort_by = request.GET.get('sort', 'date')

    today = date.today()
    if not start_date_str:
        start_date = date(today.year, 1, 1)
        start_date_str = start_date.strftime('%Y-%m-%d')
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
        end_date_str = end_date.strftime('%Y-%m-%d')
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    readings = CylinderTempReading.objects.select_related(
        'cylinder_temp',
        'cylinder_temp__equipment',
        'cylinder_temp__work_order'
    ).filter(
        cylinder_temp__date__gte=start_date,
        cylinder_temp__date__lte=end_date
    )

    if eq_num_val:
        readings = readings.filter(
            Q(cylinder_temp__equipment__Equipment_Number__icontains=eq_num_val) |
            Q(cylinder_temp__equipment__Equipment_Description__icontains=eq_num_val)
        )
    if wo_num_val:
        readings = readings.filter(
            Q(cylinder_temp__work_order__work_order__icontains=wo_num_val) |
            Q(cylinder_temp__work_order__troubleshoot_description__icontains=wo_num_val)
        )

    sort_mapping = {
        'date': 'cylinder_temp__date',
        '-date': '-cylinder_temp__date',
        'equipment_number': 'cylinder_temp__equipment__Equipment_Number',
        '-equipment_number': '-cylinder_temp__equipment__Equipment_Number',
        'work_order': 'cylinder_temp__work_order__work_order',
        '-work_order': '-cylinder_temp__work_order__work_order',
        'temp_reading': 'temp_reading',
        '-temp_reading': '-temp_reading',
    }
    db_sort_field = sort_mapping.get(sort_by, 'cylinder_temp__date')
    readings = readings.order_by(db_sort_field).distinct()

    chart_months = []
    current_bucket = date(start_date.year, start_date.month, 1)
    target_end_bucket = date(end_date.year, end_date.month, 1)

    while current_bucket <= target_end_bucket:
        chart_months.append(current_bucket.strftime('%Y-%m'))
        current_bucket += relativedelta(months=1)

    cyl_monthly_values = defaultdict(lambda: defaultdict(list))
    for r in readings:
        if r.cylinder_temp.date:
            month_str = r.cylinder_temp.date.strftime('%Y-%m')
            cyl_num = f"Cyl #{r.cylinder_number}"
            if r.temp_reading is not None:
                cyl_monthly_values[cyl_num][month_str].append(float(r.temp_reading))

    chart_datasets = []
    colors = ['#FF6384', '#36A2EB', '#4BC0C0', '#FFCE56', '#9966FF', '#FF9F40', '#32CD32', '#008080']

    for index, (cyl_label, monthly_map) in enumerate(cyl_monthly_values.items()):
        data_points = []
        for m_str in chart_months:
            vals = monthly_map.get(m_str, [])
            data_points.append(sum(vals) / len(vals) if vals else None)
        
        color = colors[index % len(colors)]
        chart_datasets.append({
            'label': cyl_label,
            'data': data_points,
            'borderColor': color,
            'backgroundColor': color,
            'fill': False,
            'tension': 0.1,
            'spanGaps': True
        })

    return render(request, 'condition_monitoring/cyl_temp_trends.html', {
        'readings': readings,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'sort': sort_by,
        'labels_json': json.dumps(chart_months),
        'datasets_json': json.dumps(chart_datasets),
    })

def add_lip_measurement(request):
    if request.method == 'POST':
        form = BucketLipForm(request.POST)
        formset = LipMeasurementFormSet(request.POST, prefix='readings')
        if form.is_valid() and formset.is_valid():
            bucket_lip = form.save()
            formset.instance = bucket_lip
            formset.save()
            messages.success(request, 'Bucket lip measurements added successfully.')
            return redirect('condition_monitoring:bucket_lips')
    else:
        form = BucketLipForm()
        formset = LipMeasurementFormSet(prefix='readings')
        
    return render(request, 'condition_monitoring/add_lip_measurement.html', {
        'form': form,
        'formset': formset,
    })

def search_lip_measurements(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    lip_meas = BucketLip.objects.select_related(
        'equipment',
        'work_order',
        'meter'
    ).prefetch_related(
        'lip_measurements'
    )

    if eq_num_val:
        lip_meas = lip_meas.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    
    if wo_num_val:
        lip_meas = lip_meas.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
        
    if selected_meter_type:
        lip_meas = lip_meas.filter(meter__meter_type=selected_meter_type)

    sort_mapping = {
        'date': 'date',
        '-date': '-date',
        'equipment_number': 'equipment__Equipment_Number',
        '-equipment_number': '-equipment__Equipment_Number',
        'equipment_desc': 'equipment__Equipment_Description',
        '-equipment_desc': '-equipment__Equipment_Description',
        'work_order': 'work_order__work_order',
        '-work_order': '-work_order__work_order',
        'wo_desc': 'work_order__troubleshoot_description',
        '-wo_desc': '-work_order__troubleshoot_description',
        'meter': 'meter__meter_type',
        '-meter': '-meter__meter_type',
        'meter_reading': 'meter_reading',
        '-meter_reading': '-meter_reading',
        'lip_measurement': 'lip_measurements__lip_measurement',
        '-lip_measurement': '-lip_measurements__lip_measurement',
    }
    
    db_sort_field = sort_mapping.get(sort_by, '-date')
    lip_meas = lip_meas.order_by(db_sort_field).distinct()
    meter_type_choices = [choice[0] for choice in Meter._meta.get_field('meter_type').choices]

    return render(request, 'condition_monitoring/search_lip_measurements.html', {
        'lip_measurements': lip_meas,
        'meter_choices': meter_type_choices,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'selected_meter': selected_meter_type,
        'sort': sort_by,
    })

def edit_lip_measurement(request, pk):
    bucket_lip = get_object_or_404(BucketLip, pk=pk)
    
    if request.method == 'POST':
        form = BucketLipForm(request.POST, instance=bucket_lip)
        formset = LipMeasurementFormSet(request.POST, instance=bucket_lip)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Lip measurements updated successfully.')
            return redirect('condition_monitoring:search_lip_measurements')
    else:
        form = BucketLipForm(instance=bucket_lip)
        formset = LipMeasurementFormSet(instance=bucket_lip)

    return render(request, 'condition_monitoring/edit_lip_measurement.html', {
        'form': form,
        'formset': formset,
        'lip_measurement': bucket_lip,
    })

@require_POST
def delete_lip_measurement(request, pk):
    bucket_lip = get_object_or_404(BucketLip, pk=pk)
    bucket_lip.delete()
    messages.success(request, 'Lip measurements deleted successfully.')
    return redirect('condition_monitoring:search_lip_measurements')

def export_lip_measurements_excel(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    lip_meas = BucketLip.objects.select_related(
        'equipment',
        'work_order',
        'meter'
    ).prefetch_related(
        'lip_measurements'
    )

    if eq_num_val:
        lip_meas = lip_meas.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    
    if wo_num_val:
        lip_meas = lip_meas.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
        
    if selected_meter_type:
        lip_meas = lip_meas.filter(meter__meter_type=selected_meter_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bucket Lip Measurements"

    headers = [
        'Date', 'Equipment #', 'Equipment Description', 
        'Work Order', 'WO Troubleshoot Desc', 'Meter Type', 
        'Meter Reading', 'Left Side', 'Centre', 'Right Side', 'Comments',
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for lip in lip_meas:
        readings = lip.lip_measurements.all()
        
        if readings.exists():
            for reading in readings:
                ws.append([
                    lip.date.strftime('%Y-%m-%d') if lip.date else '',
                    str(lip.equipment.Equipment_Number) if lip.equipment else '',
                    lip.equipment.Equipment_Description if lip.equipment else '',
                    str(lip.work_order.work_order) if lip.work_order else '',
                    lip.work_order.troubleshoot_description if lip.work_order else '',
                    lip.meter.meter_type if lip.meter else '',
                    lip.meter_reading if lip.meter_reading is not None else '',
                    reading.left_side if reading.left_side is not None else '',
                    reading.centre if reading.centre is not None else '',
                    reading.right_side if reading.right_side is not None else '',
                    lip.comments or ''
                ])
        else:
            ws.append([
                lip.date.strftime('%Y-%m-%d') if lip.date else '',
                str(lip.equipment.Equipment_Number) if lip.equipment else '',
                lip.equipment.Equipment_Description if lip.equipment else '',
                str(lip.work_order.work_order) if lip.work_order else '',
                lip.work_order.troubleshoot_description if lip.work_order else '',
                lip.meter.meter_type if lip.meter else '',
                lip.meter_reading if lip.meter_reading is not None else '',
                'No readings', '', '',
                lip.comments or ''
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bucket_lip_measurements.xlsx"'
    wb.save(response)
    
    return response

def lip_measurement_trends(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    sort_by = request.GET.get('sort', 'date')

    today = date.today()
    if not start_date_str:
        start_date = date(today.year, 1, 1)
        start_date_str = start_date.strftime('%Y-%m-%d')
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
        end_date_str = end_date.strftime('%Y-%m-%d')
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    readings = LipMeasurement.objects.select_related(
        'bucket_lip',
        'bucket_lip__equipment',
        'bucket_lip__work_order'
    ).filter(
        bucket_lip__date__gte=start_date,
        bucket_lip__date__lte=end_date
    )

    if eq_num_val:
        readings = readings.filter(
            Q(bucket_lip__equipment__Equipment_Number__icontains=eq_num_val) |
            Q(bucket_lip__equipment__Equipment_Description__icontains=eq_num_val)
        )
    if wo_num_val:
        readings = readings.filter(
            Q(bucket_lip__work_order__work_order__icontains=wo_num_val) |
            Q(bucket_lip__work_order__troubleshoot_description__icontains=wo_num_val)
        )

    sort_mapping = {
        'date': 'bucket_lip__date',
        '-date': '-bucket_lip__date',
        'equipment_number': 'bucket_lip__equipment__Equipment_Number',
        '-equipment_number': '-bucket_lip__equipment__Equipment_Number',
        'work_order': 'bucket_lip__work_order__work_order',
        '-work_order': '-bucket_lip__work_order__work_order',
    }
    db_sort_field = sort_mapping.get(sort_by, 'bucket_lip__date')
    readings = readings.order_by(db_sort_field).distinct()

    chart_months = []
    current_bucket = date(start_date.year, start_date.month, 1)
    target_end_bucket = date(end_date.year, end_date.month, 1)

    while current_bucket <= target_end_bucket:
        chart_months.append(current_bucket.strftime('%Y-%m'))
        current_bucket += relativedelta(months=1)

    pos_monthly_values = defaultdict(lambda: defaultdict(list))
    
    for r in readings:
        if r.bucket_lip.date:
            m_str = r.bucket_lip.date.strftime('%Y-%m')
            if r.left_side is not None:
                pos_monthly_values['Left Side'][m_str].append(float(r.left_side))
            if r.centre is not None:
                pos_monthly_values['Centre'][m_str].append(float(r.centre))
            if r.right_side is not None:
                pos_monthly_values['Right Side'][m_str].append(float(r.right_side))

    chart_datasets = []
    colors = ['#36A2EB', '#4BC0C0', '#FF9F40']
    
    for index, pos_label in enumerate(['Left Side', 'Centre', 'Right Side']):
        data_points = []
        monthly_map = pos_monthly_values[pos_label]
        
        for m_str in chart_months:
            vals = monthly_map.get(m_str, [])
            data_points.append(sum(vals) / len(vals) if vals else None)
            
        chart_datasets.append({
            'label': f"{pos_label} Avg",
            'data': data_points,
            'borderColor': colors[index],
            'backgroundColor': colors[index],
            'fill': False,
            'tension': 0.1,
            'spanGaps': True
        })

    return render(request, 'condition_monitoring/lip_measurement_trends.html', {
        'readings': readings,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'sort': sort_by,
        'labels_json': json.dumps(chart_months),
        'datasets_json': json.dumps(chart_datasets),
    })

def add_liner_measurement(request):
    if request.method == 'POST':
        form = BoxLinerForm(request.POST)
        formset = LinerMeasurementFormSet(request.POST, prefix='readings')
        if form.is_valid() and formset.is_valid():
            box_liner = form.save()
            formset.instance = box_liner
            formset.save()
            messages.success(request, 'Box liner measurements added successfully.')
            return redirect('condition_monitoring:box_liners')
    else:
        form = BoxLinerForm()
        formset = LinerMeasurementFormSet(prefix='readings')
        
    return render(request, 'condition_monitoring/add_liner_measurement.html', {
        'form': form,
        'formset': formset,
    })

def search_liner_measurements(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    liner_meas = BoxLiner.objects.select_related(
        'equipment',
        'work_order',
        'meter'
    ).prefetch_related(
        'liner_measurements'
    )

    if eq_num_val:
        liner_meas = liner_meas.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    
    if wo_num_val:
        liner_meas = liner_meas.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
        
    if selected_meter_type:
        liner_meas = liner_meas.filter(meter__meter_type=selected_meter_type)

    sort_mapping = {
        'date': 'date',
        '-date': '-date',
        'equipment_number': 'equipment__Equipment_Number',
        '-equipment_number': '-equipment__Equipment_Number',
        'equipment_desc': 'equipment__Equipment_Description',
        '-equipment_desc': '-equipment__Equipment_Description',
        'work_order': 'work_order__work_order',
        '-work_order': '-work_order__work_order',
        'wo_desc': 'work_order__troubleshoot_description',
        '-wo_desc': '-work_order__troubleshoot_description',
        'meter': 'meter__meter_type',
        '-meter': '-meter__meter_type',
        'meter_reading': 'meter_reading',
        '-meter_reading': '-meter_reading',
        'liner_measurement': 'liner_measurements__liner_measurement',
        '-liner_measurement': '-liner_measurements__liner_measurement',
    }
    
    db_sort_field = sort_mapping.get(sort_by, '-date')
    liner_meas = liner_meas.order_by(db_sort_field).distinct()
    meter_type_choices = [choice[0] for choice in Meter._meta.get_field('meter_type').choices]

    return render(request, 'condition_monitoring/search_liner_measurements.html', {
        'liner_measurements': liner_meas,
        'meter_choices': meter_type_choices,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'selected_meter': selected_meter_type,
        'sort': sort_by,
    })

def edit_liner_measurement(request, pk):
    box_liner = get_object_or_404(BoxLiner, pk=pk)
    
    if request.method == 'POST':
        form = BoxLinerForm(request.POST, instance=box_liner)
        formset = LinerMeasurementFormSet(request.POST, instance=box_liner)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Liner measurements updated successfully.')
            return redirect('condition_monitoring:search_liner_measurements')
    else:
        form = BoxLinerForm(instance=box_liner)
        formset = LinerMeasurementFormSet(instance=box_liner)

    return render(request, 'condition_monitoring/edit_liner_measurement.html', {
        'form': form,
        'formset': formset,
        'liner_measurement': box_liner,
    })

@require_POST
def delete_liner_measurement(request, pk):
    box_liner = get_object_or_404(BoxLiner, pk=pk)
    box_liner.delete()
    messages.success(request, 'Liner measurements deleted successfully.')
    return redirect('condition_monitoring:search_liner_measurements')

def export_liner_measurements_excel(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    liner_meas = BoxLiner.objects.select_related(
        'equipment',
        'work_order',
        'meter'
    ).prefetch_related(
        'liner_measurements'
    )

    if eq_num_val:
        liner_meas = liner_meas.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    
    if wo_num_val:
        liner_meas = liner_meas.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
        
    if selected_meter_type:
        liner_meas = liner_meas.filter(meter__meter_type=selected_meter_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Box Liner Measurements"

    headers = [
        'Date', 'Equipment #', 'Equipment Description', 
        'Work Order', 'WO Troubleshoot Desc', 'Meter Type', 
        'Meter Reading', 'Position', 'Measurement', 'Comments',
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for liner in liner_meas:
        readings = liner.liner_measurements.all()
        
        if readings.exists():
            for reading in readings:
                ws.append([
                    liner.date.strftime('%Y-%m-%d') if liner.date else '',
                    str(liner.equipment.Equipment_Number) if liner.equipment else '',
                    liner.equipment.Equipment_Description if liner.equipment else '',
                    str(liner.work_order.work_order) if liner.work_order else '',
                    liner.work_order.troubleshoot_description if liner.work_order else '',
                    liner.meter.meter_type if liner.meter else '',
                    liner.meter_reading if liner.meter_reading is not None else '',
                    reading.left_side if reading.left_side is not None else '',
                    reading.centre if reading.centre is not None else '',
                    reading.right_side if reading.right_side is not None else '',
                    liner.comments or ''
                ])
        else:
            ws.append([
                liner.date.strftime('%Y-%m-%d') if liner.date else '',
                str(liner.equipment.Equipment_Number) if liner.equipment else '',
                liner.equipment.Equipment_Description if liner.equipment else '',
                str(liner.work_order.work_order) if liner.work_order else '',
                liner.work_order.troubleshoot_description if liner.work_order else '',
                liner.meter.meter_type if liner.meter else '',
                liner.meter_reading if liner.meter_reading is not None else '',
                'No readings', '', '',
                liner.comments or ''
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bucket_liner_measurements.xlsx"'
    wb.save(response)
    
    return response

def liner_measurement_trends(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    sort_by = request.GET.get('sort', 'date')

    today = date.today()
    if not start_date_str:
        start_date = date(today.year, 1, 1)
        start_date_str = start_date.strftime('%Y-%m-%d')
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
        end_date_str = end_date.strftime('%Y-%m-%d')
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    readings = LinerMeasurement.objects.select_related(
        'box_liner',
        'box_liner__equipment',
        'box_liner__work_order'
    ).filter(
        box_liner__date__gte=start_date,
        box_liner__date__lte=end_date
    )

    if eq_num_val:
        readings = readings.filter(
            Q(box_liner__equipment__Equipment_Number__icontains=eq_num_val) |
            Q(box_liner__equipment__Equipment_Description__icontains=eq_num_val)
        )
    if wo_num_val:
        readings = readings.filter(
            Q(box_liner__work_order__work_order__icontains=wo_num_val) |
            Q(box_liner__work_order__troubleshoot_description__icontains=wo_num_val)
        )

    sort_mapping = {
        'date': 'box_liner__date',
        '-date': '-box_liner__date',
        'equipment_number': 'box_liner__equipment__Equipment_Number',
        '-equipment_number': '-box_liner__equipment__Equipment_Number',
        'work_order': 'box_liner__work_order__work_order',
        '-work_order': '-box_liner__work_order__work_order',
    }
    db_sort_field = sort_mapping.get(sort_by, 'box_liner__date')
    readings = readings.order_by(db_sort_field).distinct()

    chart_months = []
    current_bucket = date(start_date.year, start_date.month, 1)
    target_end_bucket = date(end_date.year, end_date.month, 1)

    while current_bucket <= target_end_bucket:
        chart_months.append(current_bucket.strftime('%Y-%m'))
        current_bucket += relativedelta(months=1)

    pos_monthly_values = defaultdict(lambda: defaultdict(list))
    
    for r in readings:
        if r.box_liner.date:
            m_str = r.box_liner.date.strftime('%Y-%m')
            pos_label = str(r.position).strip() if r.position else "Unknown Position"
            if r.pos_reading is not None:
                pos_monthly_values[pos_label][m_str].append(float(r.pos_reading))

    chart_datasets = []
    colors = [
        '#36A2EB', '#4BC0C0', '#FF9F40', '#FF6384', '#9966FF', 
        '#FFCE56', '#32CD32', '#008080', '#E6194B', '#3CB44B', 
        '#FFE119', '#4363D8', '#F58231'
    ]
    
    for index, pos_label in enumerate(sorted(pos_monthly_values.keys())):
        data_points = []
        monthly_map = pos_monthly_values[pos_label]
        
        for m_str in chart_months:
            vals = monthly_map.get(m_str, [])
            data_points.append(sum(vals) / len(vals) if vals else None)
            
        color = colors[index % len(colors)]
            
        chart_datasets.append({
            'label': f"{pos_label} Avg",
            'data': data_points,
            'borderColor': color,
            'backgroundColor': color,
            'fill': False,
            'tension': 0.1,
            'spanGaps': True
        })

    return render(request, 'condition_monitoring/liner_measurement_trends.html', {
        'readings': readings,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'sort': sort_by,
        'labels_json': json.dumps(chart_months),
        'datasets_json': json.dumps(chart_datasets),
    })

def add_cycle_time(request):
    if request.method == 'POST':
        form = CycleTimeForm(request.POST)
        formset = CycleTimeMeasurementFormSet(request.POST, prefix='readings')
        if form.is_valid() and formset.is_valid():
            cycle_time = form.save()
            formset.instance = cycle_time
            formset.save()
            messages.success(request, 'Cycle time measurements added successfully.')
            return redirect('condition_monitoring:cycle_times')
    else:
        form = CycleTimeForm()
        formset = CycleTimeMeasurementFormSet(prefix='readings')
        
    return render(request, 'condition_monitoring/add_cycle_time.html', {
        'form': form,
        'formset': formset,
    })

def search_cycle_times(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    cycle_time = CycleTime.objects.select_related(
        'equipment',
        'work_order',
        'meter'
    ).prefetch_related(
        'cycle_time_measurements'
    )

    if eq_num_val:
        cycle_time = cycle_time.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    
    if wo_num_val:
        cycle_time = cycle_time.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
        
    if selected_meter_type:
        cycle_time = cycle_time.filter(meter__meter_type=selected_meter_type)

    sort_mapping = {
        'date': 'date',
        '-date': '-date',
        'equipment_number': 'equipment__Equipment_Number',
        '-equipment_number': '-equipment__Equipment_Number',
        'equipment_desc': 'equipment__Equipment_Description',
        '-equipment_desc': '-equipment__Equipment_Description',
        'work_order': 'work_order__work_order',
        '-work_order': '-work_order__work_order',
        'wo_desc': 'work_order__troubleshoot_description',
        '-wo_desc': '-work_order__troubleshoot_description',
        'meter': 'meter__meter_type',
        '-meter': '-meter__meter_type',
        'meter_reading': 'meter_reading',
        '-meter_reading': '-meter_reading',
        'cycle_time_measurement': 'cycle_time_measurements__cycle_time_measurement',
        '-cycle_time_measurement': '-cycle_time_measurements__cycle_time_measurement',
    }
    
    db_sort_field = sort_mapping.get(sort_by, '-date')
    cycle_time = cycle_time.order_by(db_sort_field).distinct()
    meter_type_choices = [choice[0] for choice in Meter._meta.get_field('meter_type').choices]

    return render(request, 'condition_monitoring/search_cycle_times.html', {
        'cycle_times': cycle_time,
        'meter_choices': meter_type_choices,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'selected_meter': selected_meter_type,
        'sort': sort_by,
    })

def edit_cycle_time(request, pk):
    cycle_time = get_object_or_404(CycleTime, pk=pk)
    
    if request.method == 'POST':
        form = CycleTimeForm(request.POST, instance=cycle_time)
        formset = CycleTimeMeasurementFormSet(request.POST, instance=cycle_time)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Cycle time updated successfully.')
            return redirect('condition_monitoring:search_cycle_times')
    else:
        form = CycleTimeForm(instance=cycle_time)
        formset = CycleTimeMeasurementFormSet(instance=cycle_time)

    return render(request, 'condition_monitoring/edit_cycle_time.html', {
        'form': form,
        'formset': formset,
        'cycle_time': cycle_time,
    })

@require_POST
def delete_cycle_time(request, pk):
    cycle_time = get_object_or_404(CycleTime, pk=pk)
    cycle_time.delete()
    messages.success(request, 'Cycle time deleted successfully.')
    return redirect('condition_monitoring:search_cycle_times')

def export_cycle_times_excel(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    cycle_time_meas = CycleTime.objects.select_related(
        'equipment',
        'work_order',
        'meter'
    ).prefetch_related(
        'cycle_time_measurements'
    )

    if eq_num_val:
        cycle_time_meas = cycle_time_meas.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    
    if wo_num_val:
        cycle_time_meas = cycle_time_meas.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
        
    if selected_meter_type:
        cycle_time_meas = cycle_time_meas.filter(meter__meter_type=selected_meter_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cycle Times"

    headers = [
        'Date', 'Equipment #', 'Equipment Description', 
        'Work Order', 'WO Troubleshoot Desc', 'Meter Type', 
        'Meter Reading', 'System', 'Position', 'Measurement', 'Comments',
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for cycle_time in cycle_time_meas:
        readings = cycle_time.cycle_time_measurements.all()
        
        if readings.exists():
            for reading in readings:
                ws.append([
                    cycle_time.date.strftime('%Y-%m-%d') if cycle_time.date else '',
                    str(cycle_time.equipment.Equipment_Number) if cycle_time.equipment else '',
                    cycle_time.equipment.Equipment_Description if cycle_time.equipment else '',
                    str(cycle_time.work_order.work_order) if cycle_time.work_order else '',
                    cycle_time.work_order.troubleshoot_description if cycle_time.work_order else '',
                    cycle_time.meter.meter_type if cycle_time.meter else '',
                    cycle_time.meter_reading if cycle_time.meter_reading is not None else '',
                    reading.system if reading.system is not None else '',
                    reading.position if reading.position is not None else '',
                    reading.time if reading.time is not None else '',
                    cycle_time.comments or ''
                ])
        else:
            ws.append([
                cycle_time.date.strftime('%Y-%m-%d') if cycle_time.date else '',
                str(cycle_time.equipment.Equipment_Number) if cycle_time.equipment else '',
                cycle_time.equipment.Equipment_Description if cycle_time.equipment else '',
                str(cycle_time.work_order.work_order) if cycle_time.work_order else '',
                cycle_time.work_order.troubleshoot_description if cycle_time.work_order else '',
                cycle_time.meter.meter_type if cycle_time.meter else '',
                cycle_time.meter_reading if cycle_time.meter_reading is not None else '',
                'No readings', '', '',
                cycle_time.comments or ''
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cycle_time_measurements.xlsx"'
    wb.save(response)
    
    return response

def cycle_time_trends(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    sort_by = request.GET.get('sort', 'date')

    today = date.today()
    if not start_date_str:
        start_date = date(today.year, 1, 1)
        start_date_str = start_date.strftime('%Y-%m-%d')
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

    if not end_date_str:
        end_date = today
        end_date_str = end_date.strftime('%Y-%m-%d')
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    readings = CycleTimeMeasurement.objects.select_related(
        'cycle_time',
        'cycle_time__equipment',
        'cycle_time__work_order'
    ).filter(
        cycle_time__date__gte=start_date,
        cycle_time__date__lte=end_date
    )

    if eq_num_val:
        readings = readings.filter(
            Q(cycle_time__equipment__Equipment_Number__icontains=eq_num_val) |
            Q(cycle_time__equipment__Equipment_Description__icontains=eq_num_val)
        )
    if wo_num_val:
        readings = readings.filter(
            Q(cycle_time__work_order__work_order__icontains=wo_num_val) |
            Q(cycle_time__work_order__troubleshoot_description__icontains=wo_num_val)
        )

    sort_mapping = {
        'date': 'cycle_time__date',
        '-date': '-cycle_time__date',
        'equipment_number': 'cycle_time__equipment__Equipment_Number',
        '-equipment_number': '-cycle_time__equipment__Equipment_Number',
        'work_order': 'cycle_time__work_order__work_order',
        '-work_order': '-cycle_time__work_order__work_order',
    }
    db_sort_field = sort_mapping.get(sort_by, 'cycle_time__date')
    readings = readings.order_by(db_sort_field).distinct()

    chart_months = []
    current_bucket = date(start_date.year, start_date.month, 1)
    target_end_bucket = date(end_date.year, end_date.month, 1)

    while current_bucket <= target_end_bucket:
        chart_months.append(current_bucket.strftime('%Y-%m'))
        current_bucket += relativedelta(months=1)

    pos_monthly_values = defaultdict(lambda: defaultdict(list))
    
    for r in readings:
        if r.cycle_time.date:
            m_str = r.cycle_time.date.strftime('%Y-%m')
            
            sys_name = str(r.system).strip() if hasattr(r, 'system') and r.system else "Unknown System"
            pos_name = str(r.position).strip() if hasattr(r, 'position') and r.position else "Unknown Position"
            
            trend_line_label = f"{sys_name} - {pos_name}"
            
            if r.time is not None:
                pos_monthly_values[trend_line_label][m_str].append(float(r.time))

    chart_datasets = []
    colors = [
        '#36A2EB', '#4BC0C0', '#FF9F40', '#FF6384', '#9966FF', 
        '#FFCE56', '#32CD32', '#008080', '#E6194B', '#3CB44B', 
        '#FFE119', '#4363D8', '#F58231'
    ]
    
    for index, line_label in enumerate(sorted(pos_monthly_values.keys())):
        data_points = []
        monthly_map = pos_monthly_values[line_label]
        
        for m_str in chart_months:
            vals = monthly_map.get(m_str, [])
            data_points.append(sum(vals) / len(vals) if vals else None)
            
        color = colors[index % len(colors)]
            
        chart_datasets.append({
            'label': f"{line_label} Avg",
            'data': data_points,
            'borderColor': color,
            'backgroundColor': color,
            'fill': False,
            'tension': 0.1,
            'spanGaps': True
        })

    return render(request, 'condition_monitoring/cycle_time_trends.html', {
        'readings': readings,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'sort': sort_by,
        'labels_json': json.dumps(chart_months),
        'datasets_json': json.dumps(chart_datasets),
    })

def tire_information(request, pk=None):
    if pk:
        active_instance = get_object_or_404(TireInformation, pk=pk)
    else:
        active_instance = None

    if request.method == 'POST':
        form = TireInformationForm(request.POST, instance=active_instance)
        if form.is_valid():
            form.save()
            if active_instance:
                messages.success(request, 'Tire Information entry updated successfully.')
            else:
                messages.success(request, 'Tire Information entry added successfully.')
            return redirect('condition_monitoring:tire_information')
    else:
        form = TireInformationForm(instance=active_instance)

    sort_by = request.GET.get('sort', 'asset_type')
    
    sort_mapping = {
        'asset_type': 'asset_type__name',
        '-asset_type': '-asset_type__name',
        'equipment_type': 'equipment_type__Equipment_Type',
        '-equipment_type': '-equipment_type__Equipment_Type',
        'make': 'make',
        '-make': '-make',
        'model': 'model',
        '-model': '-model',
        'tire_size': 'tire_size',
        '-tire_size': '-tire_size',
        'tire_cost': 'tire_cost',
        '-tire_cost': '-tire_cost',
    }
    
    db_sort_field = sort_mapping.get(sort_by, 'asset_type__name')

    tire_info_history = TireInformation.objects.all().select_related(
        'asset_type', 'equipment_type'
    ).order_by(db_sort_field)

    return render(request, 'condition_monitoring/tire_information.html', {
        'form': form,
        'tire_info': tire_info_history,
        'is_editing': active_instance is not None,
        'sort': sort_by,
    })

@require_POST
def delete_tire_information(request, pk):
    tire_info_record = get_object_or_404(TireInformation, pk=pk)
    tire_info_record.delete()
    messages.success(request, 'Tire Information entry deleted successfully.')
    return redirect('condition_monitoring:tire_information')

def get_equipment_types(request):
    asset_type_id = request.GET.get('asset_type_id')
    
    if not asset_type_id:
        return JsonResponse([], safe=False)
    eq_types = EQ_Type.objects.filter(
        Asset_Type_id=asset_type_id
    ).values('id', 'Equipment_Type')
    
    return JsonResponse(list(eq_types), safe=False)

def tire_fail_types(request):
    if request.method == 'POST':
        form = TireFailureTypeForm(request.POST)
        action = request.POST.get('action_type', 'add')
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Tire failure type updated successfully.')
            
            if action == 'save':
                return redirect('condition_monitoring:tires')
                
            return redirect('condition_monitoring:tire_fail_types')
    else:
        form = TireFailureTypeForm()

    sort_by = request.GET.get('sort', 'failure_mode')
    sort_mapping = {
        'failure_mode': 'failure_mode',
        '-failure_mode': '-failure_mode',
    }
    db_sort_field = sort_mapping.get(sort_by, 'failure_mode')
    failures = TireFailure.objects.all().order_by(db_sort_field)

    return render(request, 'condition_monitoring/tire_fail_types.html', {
        'form': form,
        'failures': failures,
        'sort': sort_by,
    })

@require_POST
def delete_selected_tire_failures(request):
    selected_ids = request.POST.getlist('selected_failures')
    if selected_ids:
        deleted_count, _ = TireFailure.objects.filter(id__in=selected_ids).delete()
        messages.success(request, f'Successfully deleted {deleted_count} failure type(s).')
    else:
        messages.warning(request, 'No items were selected for deletion.')
    return redirect('condition_monitoring:tire_fail_types')

def get_purchase_orders_by_wo(request):
    wo_number = request.GET.get('work_order', '').strip()
    if not wo_number:
        return JsonResponse([], safe=False)
    pos = Purchase.objects.filter(wo_cc=wo_number).values('id', 'purchase_number')
    return JsonResponse(list(pos), safe=False)

def get_po_cost(request):
    po_id = request.GET.get('po_id')
    if not po_id:
        return JsonResponse({'cost': '0.00'})
    purchase = get_object_or_404(Purchase, id=po_id)
    return JsonResponse({'cost': str(purchase.grand_total)})

def tire_change(request, pk=None):
    if pk:
        tire_change_parent = get_object_or_404(TireChange, pk=pk)
    else:
        tire_change_parent = None

    if request.method == 'POST':
        form = TireChangeForm(request.POST, instance=tire_change_parent)
        formset = TireChangeInfoFormSet(request.POST, instance=tire_change_parent, prefix='readings')
        
        if form.is_valid() and formset.is_valid():
            parent_instance = form.save()
            formset.instance = parent_instance
            formset.save() 
            
            for child in parent_instance.tire_change_info.all():
                if child.purchase_order:
                    child.tire_cost = child.purchase_order.grand_total
                else:
                    child.tire_cost = Decimal('0.00')
                child.save()

            messages.success(request, 'Tire change recorded and individual PO tire costs synchronized successfully.')
            return redirect('condition_monitoring:tires')
        else:
            print("--- parent form errors ---", form.errors)
            print("--- child formset errors ---", formset.errors)
            messages.error(request, 'Please correct the errors detailed below.')
    else:
        form = TireChangeForm(instance=tire_change_parent)
        extra_rows = 1 if tire_change_parent is None else 0
        
        formset = TireChangeInfoFormSet(
            instance=tire_change_parent, 
            prefix='readings',
            queryset=TireChangeInfo.objects.filter(tire_change=tire_change_parent) if tire_change_parent else TireChangeInfo.objects.none()
        )
        formset.extra = extra_rows
        
    return render(request, 'condition_monitoring/tire_change.html', {
        'form': form,
        'formset': formset,
    })

def search_tire_changes(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    po_num_val = request.GET.get('purchase_order', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    tire_changes = TireChange.objects.select_related(
        'equipment', 'work_order', 'meter'
    ).prefetch_related(
        'tire_change_info', 
        'tire_change_info__purchase_order',
        'tire_change_info__reason_for_failure',
        'tire_change_info__scrap_reason'
    )

    if eq_num_val:
        tire_changes = tire_changes.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    if wo_num_val:
        tire_changes = tire_changes.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
    if selected_meter_type:
        tire_changes = tire_changes.filter(meter__meter_type=selected_meter_type)

    if po_num_val:
        tire_changes = tire_changes.filter(
            tire_change_info__purchase_order__purchase_number__icontains=po_num_val
        )

    if sort_by == 'tire_id_off':
        tire_changes = tire_changes.annotate(sort_val=Min('tire_change_info__tire_id_off')).order_by('sort_val')
    elif sort_by == '-tire_id_off':
        tire_changes = tire_changes.annotate(sort_val=Max('tire_change_info__tire_id_off')).order_by('-sort_val')
        
    elif sort_by == 'position':
        tire_changes = tire_changes.annotate(sort_val=Min('tire_change_info__position')).order_by('sort_val')
    elif sort_by == '-position':
        tire_changes = tire_changes.annotate(sort_val=Max('tire_change_info__position')).order_by('-sort_val')
        
    elif sort_by == 'tire_id_on':
        tire_changes = tire_changes.annotate(sort_val=Min('tire_change_info__tire_id_on')).order_by('sort_val')
    elif sort_by == '-tire_id_on':
        tire_changes = tire_changes.annotate(sort_val=Max('tire_change_info__tire_id_on')).order_by('-sort_val')
        
    elif sort_by == 'rim_id_on':
        tire_changes = tire_changes.annotate(sort_val=Min('tire_change_info__rim_id_on')).order_by('sort_val')
    elif sort_by == '-rim_id_on':
        tire_changes = tire_changes.annotate(sort_val=Max('tire_change_info__rim_id_on')).order_by('-sort_val')
        
    else:
        sort_mapping = {
            'date': 'date',
            '-date': '-date',
            'equipment_number': 'equipment__Equipment_Number',
            '-equipment_number': '-equipment__Equipment_Number',
            'equipment_desc': 'equipment__Equipment_Description',
            '-equipment_desc': '-equipment__Equipment_Description',
            'work_order': 'work_order__work_order',
            '-work_order': '-work_order__work_order',
            'wo_desc': 'work_order__troubleshoot_description',
            '-wo_desc': '-work_order__troubleshoot_description',
            'meter': 'meter__meter_type',
            '-meter': '-meter__meter_type',
        }
        db_sort_field = sort_mapping.get(sort_by, '-date')
        tire_changes = tire_changes.order_by(db_sort_field)

    meter_type_choices = [choice[0] for choice in Meter._meta.get_field('meter_type').choices]
    tire_changes = tire_changes.distinct()

    return render(request, 'condition_monitoring/search_tire_changes.html', {
        'tire_changes': tire_changes,
        'meter_choices': meter_type_choices,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'all_purchase_order_suggestions': Purchase.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'po_num_val': po_num_val,
        'selected_meter': selected_meter_type,
        'sort': sort_by,
    })

def export_tire_changes_excel(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    po_num_val = request.GET.get('purchase_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    tire_changes = TireChange.objects.select_related(
        'equipment', 'work_order', 'meter'
    ).prefetch_related(
        'tire_change_info', 
        'tire_change_info__purchase_number',
        'tire_change_info__reason_for_failure',
        'tire_change_info__scrap_reason'
    )

    if eq_num_val:
        tire_changes = tire_changes.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    if wo_num_val:
        tire_changes = tire_changes.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
    if selected_meter_type:
        tire_changes = tire_changes.filter(meter__meter_type=selected_meter_type)
    if po_num_val:
        tire_changes = tire_changes.filter(
            tire_change_info__purchase_order__purchase_order__icontains=po_num_val
        )

    if sort_by == 'tire_id_off':
        tire_changes = tire_changes.annotate(sort_val=Min('tire_change_info__tire_id_off')).order_by('sort_val')
    elif sort_by == '-tire_id_off':
        tire_changes = tire_changes.annotate(sort_val=Max('tire_change_info__tire_id_off')).order_by('-sort_val')
    elif sort_by == 'position':
        tire_changes = tire_changes.annotate(sort_val=Min('tire_change_info__position')).order_by('sort_val')
    elif sort_by == '-position':
        tire_changes = tire_changes.annotate(sort_val=Max('tire_change_info__position')).order_by('-sort_val')
    elif sort_by == 'tire_id_on':
        tire_changes = tire_changes.annotate(sort_val=Min('tire_change_info__tire_id_on')).order_by('sort_val')
    elif sort_by == '-tire_id_on':
        tire_changes = tire_changes.annotate(sort_val=Max('tire_change_info__tire_id_on')).order_by('-sort_val')
    elif sort_by == 'rim_id_on':
        tire_changes = tire_changes.annotate(sort_val=Min('tire_change_info__rim_id_on')).order_by('sort_val')
    elif sort_by == '-rim_id_on':
        tire_changes = tire_changes.annotate(sort_val=Max('tire_change_info__rim_id_on')).order_by('-sort_val')
    else:
        sort_mapping = {
            'date': 'date', '-date': '-date',
            'equipment_number': 'equipment__Equipment_Number', '-equipment_number': '-equipment__Equipment_Number',
            'equipment_desc': 'equipment__Equipment_Description', '-equipment_desc': '-equipment__Equipment_Description',
            'work_order': 'work_order__work_order', '-work_order': '-work_order__work_order',
            'wo_desc': 'work_order__troubleshoot_description', '-wo_desc': '-work_order__troubleshoot_description',
            'meter': 'meter__meter_type', '-meter': '-meter__meter_type',
        }
        db_sort_field = sort_mapping.get(sort_by, '-date')
        tire_changes = tire_changes.order_by(db_sort_field)

    tire_changes = tire_changes.distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tire Change Log Index"

    headers = [
        'Date', 'Equipment #', 'Equipment Description', 'Work Order', 'WO Troubleshoot Desc', 
        'Meter Type', 'Meter Reading', 'Tire ID Off', 'Position', 'Tread Depth Off', 'Rim ID Off', 
        'Tire ID On', 'Tread Depth On', 'Inflation Pressure', 'Rim ID On', 'Purchase Number', 'Tire Cost',
        'Failure Reason', 'Scrapped?', 'Recapped?', 'Scrap Reason', 'Comments'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for change in tire_changes:
        readings = change.tire_change_info.all()
        
        if readings.exists():
            for reading in readings:
                ws.append([
                    change.date.strftime('%Y-%m-%d') if change.date else '',
                    str(change.equipment.Equipment_Number) if change.equipment else '',
                    change.equipment.Equipment_Description if change.equipment else '',
                    str(change.work_order.work_order) if change.work_order else '',
                    change.work_order.troubleshoot_description if change.work_order else '',
                    change.meter.meter_type if change.meter else '',
                    change.meter_reading if change.meter_reading is not None else '',
                    reading.tire_id_off or '',
                    reading.position or '',
                    reading.tread_depth_off if reading.tread_depth_off is not None else '',
                    reading.rim_id_off or '',
                    reading.tire_id_on or '',
                    reading.tread_depth_on if reading.tread_depth_on is not None else '',
                    reading.inflation_pressure if reading.inflation_pressure is not None else '',
                    reading.rim_id_on or '',
                    str(reading.purchase_order.purchase_number) if reading.purchase_order else '',
                    reading.tire_cost if reading.tire_cost is not None else '',
                    reading.reason_for_failure.failure_mode if reading.reason_for_failure else '',
                    reading.scrapped or '',
                    reading.recapped or '',
                    reading.scrap_reason.failure_mode if reading.scrap_reason else '',
                    change.comments or ''
                ])
        else:
            ws.append([
                change.date.strftime('%Y-%m-%d') if change.date else '',
                str(change.equipment.Equipment_Number) if change.equipment else '',
                change.equipment.Equipment_Description if change.equipment else '',
                str(change.work_order.work_order) if change.work_order else '',
                change.work_order.troubleshoot_description if change.work_order else '',
                change.meter.meter_type if change.meter else '',
                change.meter_reading if change.meter_reading is not None else '',
                'No row details entries recorded.', '', '', '', '', '', '', '', '', '', '', '', '',
                change.comments or ''
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="tire_change_records.xlsx"'
    wb.save(response)
    
    return response

def tire_inspections(request, pk=None):
    if pk:
        inspection_parent = get_object_or_404(TireInspection, pk=pk)
    else:
        inspection_parent = None

    if request.method == 'POST':
        form = TireInspectionForm(request.POST, instance=inspection_parent)
        formset = TireInspectionReadingFormSet(request.POST, instance=inspection_parent, prefix='readings')
        
        if form.is_valid() and formset.is_valid():
            parent_instance = form.save()
            formset.instance = parent_instance
            formset.save()
            messages.success(request, 'Tire inspection recorded successfully.')
            return redirect('condition_monitoring:tire_inspection_history')
        else:
            print("--- parent form errors ---", form.errors)
            print("--- child formset errors ---", formset.errors)
            messages.error(request, 'Please correct the errors detailed below.')
    else:
        form = TireInspectionForm(instance=inspection_parent)
        extra_rows = 1 if inspection_parent is None else 0
        
        formset = TireInspectionReadingFormSet(instance=inspection_parent, prefix='readings')
        formset.extra = extra_rows
        
    return render(request, 'condition_monitoring/tire_inspections.html', {
        'form': form,
        'formset': formset,
    })

def tire_inspection_history(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    inspections = TireInspection.objects.select_related(
        'equipment', 'work_order', 'meter'
    ).prefetch_related('tire_inspection_readings')

    if eq_num_val:
        inspections = inspections.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    if wo_num_val:
        inspections = inspections.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
    if selected_meter_type:
        inspections = inspections.filter(meter__meter_type=selected_meter_type)

    if sort_by == 'tire_id':
        inspections = inspections.annotate(sort_val=Min('tire_inspection_readings__tire_id')).order_by('sort_val')
    elif sort_by == '-tire_id':
        inspections = inspections.annotate(sort_val=Max('tire_inspection_readings__tire_id')).order_by('-sort_val')
    elif sort_by == 'position':
        inspections = inspections.annotate(sort_val=Min('tire_inspection_readings__position')).order_by('sort_val')
    elif sort_by == '-position':
        inspections = inspections.annotate(sort_val=Max('tire_inspection_readings__position')).order_by('-sort_val')
    else:
        
        sort_mapping = {
            'date': 'date', '-date': '-date',
            'equipment_number': 'equipment__Equipment_Number', '-equipment_number': '-equipment__Equipment_Number',
            'equipment_desc': 'equipment__Equipment_Description', '-equipment_desc': '-equipment__Equipment_Description',
            'work_order': 'work_order__work_order', '-work_order': '-work_order__work_order',
            'wo_desc': 'work_order__troubleshoot_description', '-wo_desc': '-work_order__troubleshoot_description',
            'meter': 'meter__meter_type', '-meter': '-meter__meter_type',
        }
        db_sort_field = sort_mapping.get(sort_by, '-date')
        inspections = inspections.order_by(db_sort_field)

    inspections = inspections.distinct()
    meter_type_choices = [choice for choice in Meter._meta.get_field('meter_type').choices]

    return render(request, 'condition_monitoring/tire_inspection_history.html', {
        'tire_inspections': inspections,
        'meter_choices': meter_type_choices,
        'all_equipment_suggestions': Equipment.objects.all(),
        'all_work_order_suggestions': WorkOrder.objects.all(),
        'eq_num_val': eq_num_val,
        'wo_num_val': wo_num_val,
        'selected_meter': selected_meter_type,
        'sort': sort_by,
    })

def export_inspection_history_excel(request):
    eq_num_val = request.GET.get('equipment_number', '').strip()
    wo_num_val = request.GET.get('work_order_number', '').strip()
    selected_meter_type = request.GET.get('meter', '').strip()
    sort_by = request.GET.get('sort', '-date')

    inspections = TireInspection.objects.select_related(
        'equipment', 'work_order', 'meter'
    ).prefetch_related('tire_inspection_readings')

    if eq_num_val:
        inspections = inspections.filter(
            Q(equipment__Equipment_Number__icontains=eq_num_val) |
            Q(equipment__Equipment_Description__icontains=eq_num_val)
        )
    if wo_num_val:
        inspections = inspections.filter(
            Q(work_order__work_order__icontains=wo_num_val) |
            Q(work_order__troubleshoot_description__icontains=wo_num_val)
        )
    if selected_meter_type:
        inspections = inspections.filter(meter__meter_type=selected_meter_type)

    if sort_by == 'tire_id':
        inspections = inspections.annotate(sort_val=Min('tire_inspection_readings__tire_id')).order_by('sort_val')
    elif sort_by == '-tire_id':
        inspections = inspections.annotate(sort_val=Max('tire_inspection_readings__tire_id')).order_by('-sort_val')
    elif sort_by == 'position':
        inspections = inspections.annotate(sort_val=Min('tire_inspection_readings__position')).order_by('sort_val')
    elif sort_by == '-position':
        inspections = inspections.annotate(sort_val=Max('tire_inspection_readings__position')).order_by('-sort_val')
    else:
        sort_mapping = {
            'date': 'date', '-date': '-date',
            'equipment_number': 'equipment__Equipment_Number', '-equipment_number': '-equipment__Equipment_Number',
            'equipment_desc': 'equipment__Equipment_Description', '-equipment_desc': '-equipment__Equipment_Description',
            'work_order': 'work_order__work_order', '-work_order': '-work_order__work_order',
            'wo_desc': 'work_order__troubleshoot_description', '-wo_desc': '-work_order__troubleshoot_description',
            'meter': 'meter__meter_type', '-meter': '-meter__meter_type',
        }
        inspections = inspections.order_by(sort_mapping.get(sort_by, '-date'))

    inspections = inspections.distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tire Inspections Log"

    headers = [
        'Date', 'Equipment #', 'Equipment Description', 'Work Order', 'WO Troubleshoot Desc', 
        'Meter Type', 'Meter Reading', 'Tire Identifier', 'Position', 'Tread Depth', 
        'Tire Diameter', 'Inflation Pressure', 'Comments'
    ]
    ws.append(headers)

    for cell in ws:
        cell.font = openpyxl.styles.Font(bold=True)

    for inspect in inspections:
        readings = inspect.tire_inspection_readings.all()
        if readings.exists():
            for reading in readings:
                ws.append([
                    inspect.date.strftime('%Y-%m-%d') if inspect.date else '',
                    str(inspect.equipment.Equipment_Number) if inspect.equipment else '',
                    inspect.equipment.Equipment_Description if inspect.equipment else '',
                    str(inspect.work_order.work_order) if inspect.work_order else '',
                    inspect.work_order.troubleshoot_description if inspect.work_order else '',
                    inspect.meter.meter_type if inspect.meter else '',
                    inspect.meter_reading if inspect.meter_reading is not None else '',
                    reading.tire_id or '',
                    reading.position or '',
                    reading.tread_depth if reading.tread_depth is not None else '',
                    reading.inflation_pressure if reading.inflation_pressure is not None else '',
                    inspect.comments or ''
                ])
        else:
            ws.append([
                inspect.date.strftime('%Y-%m-%d') if inspect.date else '',
                str(inspect.equipment.Equipment_Number) if inspect.equipment else '',
                inspect.equipment.Equipment_Description if inspect.equipment else '',
                str(inspect.work_order.work_order) if inspect.work_order else '',
                inspect.work_order.troubleshoot_description if inspect.work_order else '',
                inspect.meter.meter_type if inspect.meter else '',
                inspect.meter_reading if inspect.meter_reading is not None else '',
                'No structural inspection data entries.', '', '', '',
                inspect.comments or ''
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col.column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="tire_inspection_history.xlsx"'
    wb.save(response)
    return response

def rim_inspection_panel(request, pk=None):
    if pk:
        active_instance = get_object_or_404(RimInspection, pk=pk)
    else:
        active_instance = None

    if request.method == 'POST':
        form = RimInspectionForm(request.POST, instance=active_instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Rim inspection log committed successfully.')
            return redirect('condition_monitoring:rim_inspections')
    else:
        form = RimInspectionForm(instance=active_instance)

    sort_by = request.GET.get('sort', '-date_tested')
    
    sort_mapping = {
        'date_tested': 'date_tested',
        '-date_tested': '-date_tested',
        'rim_id': 'rim_id',
        '-rim_id': '-rim_id',
        'pass_fail': 'pass_fail',
        '-pass_fail': '-pass_fail',
        'failure_reason': 'failure_reason',
        '-failure_reason': '-failure_reason',
        'last_test_date': 'last_test_date',
        '-last_test_date': '-last_test_date',
        'number_of_tests': 'number_of_tests',
        '-number_of_tests': '-number_of_tests',
        'next_test_date': 'next_test_date',
        '-next_test_date': '-next_test_date',
    }
    
    db_sort_field = sort_mapping.get(sort_by, '-date_tested')

    all_history = RimInspection.objects.all().order_by(db_sort_field, '-id')

    selected_color = request.GET.get('status_color', '')
    filtered_history = []
    today = date.today()
    amber_threshold = today + relativedelta(months=6)

    for item in all_history:
        item_color = ''
        if item.next_test_date and item.next_test_date.lower() != 'scrap':
            try:
                due_dt = datetime.strptime(item.next_test_date, '%Y-%m-%d').date()
                if due_dt < today:
                    item_color = 'red'
                elif due_dt <= amber_threshold:
                    item_color = 'amber'
                else:
                    item_color = 'clear'
            except ValueError:
                pass

        if not selected_color:
            filtered_history.append(item)
        elif selected_color == 'red' and item_color == 'red':
            filtered_history.append(item)
        elif selected_color == 'amber' and item_color == 'amber':
            filtered_history.append(item)
        elif selected_color == 'clear' and (item_color == 'clear' or not item.next_test_date):
            filtered_history.append(item)

    return render(request, 'condition_monitoring/rim_inspections.html', {
        'form': form,
        'history': filtered_history,
        'selected_color': selected_color,
        'is_editing': active_instance is not None,
        'sort': sort_by,
    })

@require_POST
def delete_rim_inspection(request, pk):
    inspection = get_object_or_404(RimInspection, pk=pk)
    inspection.delete()
    
    messages.success(request, 'Rim inspection record deleted successfully.')
    return redirect('condition_monitoring:rim_inspections')

def tires_dashboard(request):
    today = date.today()
    today_str = today.strftime('%Y-%m-%d')
    six_months_out_str = (today + relativedelta(months=6)).strftime('%Y-%m-%d')

    # ==========================================
    # 1. RIM COMPLIANCE LISTS (OVERDUE & UPCOMING)
    # ==========================================
    # Exclude failed/scrapped rows from active date tracking lists
    rims_base = RimInspection.objects.exclude(next_test_date__iexact='scrap').exclude(next_test_date__isnull=True)
    
    rims_overdue = rims_base.filter(next_test_date__lt=today_str).order_by('next_test_date')
    rims_upcoming = rims_base.filter(next_test_date__gte=today_str, next_test_date__lte=six_months_out_str).order_by('next_test_date')

    # ==========================================
    # 2. PIE CHART: SCRAP RIMS BY FAILURE REASON
    # ==========================================
    scrap_rims_query = RimInspection.objects.filter(pass_fail__iexact='Fail')
    rim_scrap_map = defaultdict(int)
    for rim in scrap_rims_query:
        reason = rim.failure_reason or "Unspecified"
        rim_scrap_map[reason] += 1
        
    rim_scrap_labels = list(rim_scrap_map.keys())
    rim_scrap_counts = list(rim_scrap_map.values())

    # ==========================================
    # 3. PIE CHART: SCRAP TIRES BY FAILURE MODE
    # ==========================================
    scrap_tires_query = TireChangeInfo.objects.filter(scrapped__iexact='Yes').select_related('scrap_reason')
    tire_scrap_map = defaultdict(int)
    for tire in scrap_tires_query:
        mode = tire.scrap_reason.failure_mode if tire.scrap_reason else "Unspecified"
        tire_scrap_map[mode] += 1

    tire_scrap_labels = list(tire_scrap_map.keys())
    tire_scrap_counts = list(tire_scrap_map.values())

    # ==========================================
    # 4. WEAR PROFILE, FINANCES, & VALUE LOSSES
    # ==========================================
    all_changes = TireChangeInfo.objects.select_related('tire_change', 'tire_change__equipment').all()
    
    # Pre-load configuration matching matrices into memory dictionary map to optimize lookup speeds
    ti_records = TireInformation.objects.all()
    ti_map = {}
    for ti in ti_records:
        key = (ti.asset_type_id, ti.equipment_type_id, str(ti.make).strip().lower(), str(ti.model).strip().lower())
        ti_map[key] = ti

    spend_map = defaultdict(float)
    loss_map = defaultdict(float)
    wear_profiles = []

    for r in all_changes:
        eq = r.tire_change.equipment if r.tire_change else None
        if not eq:
            continue
            
        eq_num = eq.Equipment_Number
        
        # Pull matching baseline properties specs
        ti_key = (eq.Asset_Type_id, eq.Equipment_Type_id, str(eq.Make).strip().lower(), str(eq.Model).strip().lower())
        matched_ti = ti_map.get(ti_key)

        # A. Track total spendings across equipment profiles
        if r.tire_cost:
            spend_map[eq_num] += float(r.tire_cost)

        # Verify numerical attributes values exist
        tread_off = r.tread_depth_off
        meter_rd = r.tire_change.meter_reading if r.tire_change else 0

        if matched_ti and tread_off is not None:
            tread_new = matched_ti.tread_depth_new
            cost_new = float(matched_ti.tire_cost or 0)

            # B. Wear Profile Rate Calculation: 32nds worn down divided by total active hours
            if meter_rd > 0 and tread_new and tread_new > tread_off:
                worn_32nds = tread_new - tread_off
                wear_rate_per_hour = worn_32nds / meter_rd
                # Express as Rate per 1000 operational machine hours for legibility
                wear_profiles.append({
                    'tire_id': r.tire_id_off or "Unknown",
                    'equipment': eq_num,
                    'position': r.position or "Unknown",
                    'rate': round(wear_rate_per_hour * 1000, 4)
                })

            # C. Financial Scrap Losses Calculation: Discarded tread value ratio matching your formula
            if r.scrapped and r.scrapped.lower() == 'yes' and tread_new > 0 and cost_new > 0:
                dollar_value_lost = tread_off * (cost_new / tread_new)
                loss_map[eq_num] += dollar_value_lost

    # Format structured bar chart metrics lists
    financial_labels = sorted(list(set(list(spend_map.keys()) + list(loss_map.keys()))))
    spend_data = [spend_map[lbl] for lbl in financial_labels]
    loss_data = [loss_map[lbl] for lbl in financial_labels]

    return render(request, 'condition_monitoring/tires_dashboard.html', {
        'rims_overdue': rims_overdue,
        'rims_upcoming': rims_upcoming,
        'wear_profiles': sorted(wear_profiles, key=lambda x: x['rate'], reverse=True)[:10], # Top 10 worst wearers
        
        # Safe structuralized JSON conversion strings arrays payload
        'tire_scrap_labels': json.dumps(tire_scrap_labels),
        'tire_scrap_counts': json.dumps(tire_scrap_counts),
        'rim_scrap_labels': json.dumps(rim_scrap_labels),
        'rim_scrap_counts': json.dumps(rim_scrap_counts),
        'financial_labels': json.dumps(financial_labels),
        'spend_data': json.dumps(spend_data),
        'loss_data': json.dumps(loss_data),
    })