from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from .models import AssetType, Equipment, EQ_Type, Meter, ComponentType, Component, ComponentHistory, ShiftReport, MachineShiftStatus
from .forms import (
    EQTypeForm,
    EqEditForm,
    EqUploadForm,
    MeterFormSet,
    CompUploadForm,
    CompChangeForm,
    ShiftReportForm,
    AssetTypeForm,
    ComponentTypeForm,
)
import json, pandas as pd, openpyxl, io
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.db.models import Q
from django.db import IntegrityError

def equipment(request):
    return render(request, 'equipment/equipment.html', {'content_type': 'text/html'})

def tables(request):
    return render(request, 'equipment/tables.html', {'content_type': 'text/html'})

def shift_report_list(request):
    reports = ShiftReport.objects.all().order_by('-date')
    return render(request, 'equipment/shift_report_list.html', {'reports': reports})

def equpload(request):
    if request.method == 'POST':
        equploadform = EqUploadForm(request.POST, request.FILES)
        meter_formset = MeterFormSet(request.POST, prefix='meters')
        if equploadform.is_valid() and meter_formset.is_valid():
            equipment_instance = equploadform.save(commit=False)
            asset_name = equploadform.cleaned_data.get('Asset_Type')
            eq_type_obj = equploadform.cleaned_data.get('Equipment_Type')
            try:
                if asset_name:
                    equipment_instance.Asset_Type = AssetType.objects.get(
                        name=asset_name,
                        is_active=True,
                    )
                if eq_type_obj:
                    equipment_instance.Equipment_Type = eq_type_obj
                equipment_instance.save()
                meter_formset.instance = equipment_instance
                meter_formset.save()
                messages.success(request, 'Equipment added successfully.')
                return redirect('equipment:equipment')
            except (AssetType.DoesNotExist, EQ_Type.DoesNotExist):
                messages.error(request, "There was an error saving the equipment. Please check the fields below.")
                equploadform.add_error(None, "Critical Error: Type selection lost.")
    else:
        equploadform = EqUploadForm()
        meter_formset = MeterFormSet(prefix='meters')
    return render(request, 'equipment/create_eq.html', {
        'equploadform': equploadform,
        'meter_formset': meter_formset
    })
    
def create_eq(request):
    equploadform = EqUploadForm()
    meter_formset = MeterFormSet(prefix='meters')
    return render(request, 'equipment/create_eq.html', {'equploadform': equploadform, 'meter_formset': meter_formset})

def get_next_equipment_number(eq_type):
    prefix = eq_type.Prefix
    last_eq = Equipment.objects.filter(
        Equipment_Number__startswith=prefix
    ).order_by('Equipment_Number').last()
    if last_eq:
        last_num = int(last_eq.Equipment_Number.split('-')[-1])
        next_num = str(last_num + 1).zfill(3)
    else:
        next_num = "001"
    return f"{prefix}-{next_num}"

def load_equipment_types(request):
    asset_name = request.GET.get('asset_id')
    types = EQ_Type.objects.filter(
        Asset_Type__name=asset_name,
        Asset_Type__is_active=True,
        is_active=True,
    ).values(
        'id',
        'Equipment_Type'
    ).order_by('Equipment_Type')
    return JsonResponse(list(types), safe=False)

def load_equipment_options(request):
    form = EqUploadForm(data=request.GET)
    asset_val = request.GET.get('Asset_Type')
    if asset_val:
        form.fields['Equipment_Type'].queryset = EQ_Type.objects.filter(
            Asset_Type__name=asset_val,
            Asset_Type__is_active=True,
            is_active=True,
        ).order_by('Equipment_Type')
    eq_type_id = request.GET.get('Equipment_Type')
    if eq_type_id and eq_type_id.isdigit():
        try:
            eq_type_obj = EQ_Type.objects.get(
                id=int(eq_type_id),
                is_active=True,
            )
            new_num = get_next_equipment_number(eq_type_obj)
            form.data = form.data.copy()
            form.data['Equipment_Number'] = new_num
            form.initial['Equipment_Number'] = new_num
        except EQ_Type.DoesNotExist:
            pass
    return render(request, 'equipment/equipment_fields.html', {'equploadform': form})
    
def generate_eq_number(request):
    type_id = request.GET.get('type_id')
    try:
        eq_type_obj = EQ_Type.objects.get(
            id=type_id,
            is_active=True,
        )
        prefix = eq_type_obj.Prefix
        last_eq = Equipment.objects.filter(
            Equipment_Number__startswith=prefix
        ).order_by('Equipment_Number').last()
        if last_eq:
            parts = last_eq.Equipment_Number.split('-')
            last_num = int(parts[-1]) if len(parts) > 1 else 0
            new_num = str(last_num + 1).zfill(3)
        else:
            new_num = "001"
        new_id = f"{prefix}-{new_num}"
        return JsonResponse({'new_number': new_id})
    except (EQ_Type.DoesNotExist, ValueError):
        return JsonResponse({'new_number': 'Error: Type Not Found'}, status=404)
    
def search_eq(request):
    equipment_list = Equipment.objects.all().select_related('Asset_Type', 'Equipment_Type')
    equipment_number = request.GET.get('Equipment_Number', '').strip()
    asset_type = request.GET.get('Asset_Type', '').strip()
    equipment_type = request.GET.get('Equipment_Type', '').strip()
    equipment_status = request.GET.get('Equipment_Status', '').strip()
    make = request.GET.get('Make', '').strip()
    model = request.GET.get('Model', '').strip()

    if equipment_number:
        equipment_list = equipment_list.filter(Equipment_Number__icontains=equipment_number)
    if asset_type:
        equipment_list = equipment_list.filter(Asset_Type__name__icontains=asset_type)
    if equipment_type:
        equipment_list = equipment_list.filter(Equipment_Type__Equipment_Type__icontains=equipment_type)
    if equipment_status:
        equipment_list = equipment_list.filter(Equipment_Status__icontains=equipment_status)
    if make:
        equipment_list = equipment_list.filter(Make__icontains=make)
    if model:
        equipment_list = equipment_list.filter(Model__icontains=model)

    sort_by = request.GET.get('sort', 'Equipment_Number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')

    sort_mapping = {
        'Equipment_Number': 'Equipment_Number',
        'Asset_Type': 'Asset_Type__name',
        'Equipment_Type': 'Equipment_Type__Equipment_Type',
        'Equipment_Status': 'Equipment_Status',
        'Equipment_Description': 'Equipment_Description',
        'Commissioning_Date': 'Commissioning_Date',
        'Decommissioning_Date': 'Decommissioning_Date',
        'Make': 'Make',
        'Model': 'Model',
    }
    
    if clean_sort_key in sort_mapping:
        db_field = sort_mapping[clean_sort_key]
        order_field = f"-{db_field}" if is_descending else db_field
        equipment_list = equipment_list.order_by(order_field)
    else:
        equipment_list = equipment_list.order_by('Equipment_Number')

    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')
    
    all_asset_types = Equipment.objects.exclude(Asset_Type__name__isnull=True).values_list('Asset_Type__name', flat=True).distinct().order_by('Asset_Type__name')
    all_eq_types = Equipment.objects.exclude(Equipment_Type__Equipment_Type__isnull=True).values_list('Equipment_Type__Equipment_Type', flat=True).distinct().order_by('Equipment_Type__Equipment_Type')
    all_statuses = Equipment.objects.exclude(Equipment_Status__isnull=True).values_list('Equipment_Status', flat=True).distinct().order_by('Equipment_Status')
    all_makes = Equipment.objects.exclude(Make__isnull=True).values_list('Make', flat=True).distinct().order_by('Make')
    all_models = Equipment.objects.exclude(Model__isnull=True).values_list('Model', flat=True).distinct().order_by('Model')


    return render(request, 'equipment/search_eq.html', {
        'equipment_list': equipment_list, 
        'sort': sort_by,
         'all_equipment_suggestions': all_equipment_suggestions,
        'all_asset_types': all_asset_types,
        'all_eq_types': all_eq_types,
        'all_statuses': all_statuses,
        'all_makes': all_makes,
        'all_models': all_models,
        'eq_num_val': equipment_number,
        'asset_type_val': asset_type,
        'eq_type_val': equipment_type,
        'status_val': equipment_status,
        'make_val': make,
        'model_val': model,
    })

def export_equipment(request):
    sort_by = request.GET.get('sort', 'Equipment_Number')
    equipment_list = Equipment.objects.all()
    equipment_number = request.GET.get('Equipment_Number')
    asset_type = request.GET.get('Asset_Type')
    equipment_type = request.GET.get('Equipment_Type')
    equipment_status = request.GET.get('Equipment_Status')
    make = request.GET.get('Make')
    model = request.GET.get('Model')
    if equipment_number:
        equipment_list = equipment_list.filter(Equipment_Number__icontains=equipment_number)
    if asset_type:
        equipment_list = equipment_list.filter(Asset_Type__name__icontains=asset_type)
    if equipment_type:
        equipment_list = equipment_list.filter(Equipment_Type__Equipment_Type__icontains=equipment_type)
    if equipment_status:
        equipment_list = equipment_list.filter(Equipment_Status__icontains=equipment_status)
    if make:
        equipment_list = equipment_list.filter(Make__icontains=make)
    if model:
        equipment_list = equipment_list.filter(Model__icontains=model)
    eqdata=list(equipment_list.values('Equipment_Number', 'Asset_Type__name', 'Equipment_Type__Equipment_Type', 'Equipment_Status', 'Make', 'Model'))
    df = pd.DataFrame(eqdata)
    df.rename(columns={
        'Equipment_Number': 'Equipment Number',
        'Asset_Type__name': 'Asset Type',
        'Equipment_Type__Equipment_Type': 'Equipment Type',
        'Equipment_Status': 'Status',
    }, inplace=True)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Equipment.xlsx"'
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Search Results')
        worksheet = writer.sheets['Search Results']
        for i, col in enumerate(df.columns):
            column_len = df[col].astype(str).str.len().max()
            column_len = max(column_len, len(col)) + 2
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column_len
        header_font = openpyxl.styles.Font(bold=True)
        for cell in worksheet[1]:
            cell.font = header_font
    return response

def edit_eq(request):
    all_equipment = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')
    eq_number = request.GET.get('eqedit')
    
    if request.method == 'POST':
        eq_number = request.POST.get('Equipment_Number')
        item = get_object_or_404(Equipment, Equipment_Number=eq_number)
        
        form = EqEditForm(request.POST, request.FILES, instance=item)
        meter_formset = MeterFormSet(request.POST, instance=item, prefix='meters')
                
        if form.is_valid() and meter_formset.is_valid():
            form.save()
            meter_formset.save()
            messages.success(request, 'Equipment updated successfully.')
            return redirect('equipment:equipment') 
    else:
        item = None
        form = None
        meter_formset = None
    
    if eq_number:
            item = Equipment.objects.filter(Equipment_Number=eq_number).first()
            if item:
                form = EqEditForm(instance=item)
                meter_formset = MeterFormSet(instance=item, prefix='meters')
            
    return render(request, 'equipment/edit_eq.html', {
        'all_equipment': all_equipment,
        'item': item,
        'equploadform': form,
        'meter_formset': meter_formset,
        'eq_number': eq_number,
    })

def add_component(request):
    eq_query = request.GET.get('eqedit')
    item = None
    initial_data = {}

    if eq_query:
        item = Equipment.objects.filter(Equipment_Number__iexact=eq_query).first()
        if item:
            initial_data = {
                'Equipment': item, 
                'Equipment_Number': item.Equipment_Number,
                'Equipment_Description': item.Equipment_Description
            }

    if request.method == 'POST':
        form = CompUploadForm(request.POST, request.FILES)
        if form.is_valid():
            component = form.save(commit=False)
            
            comp_type_name = component.Component_Type.name if component.Component_Type else ""
            
            seq_num = ""
            if component.Component_Number and '-' in component.Component_Number:
                seq_num = component.Component_Number.split('-')[-1]
            
            desc_parts = [
                comp_type_name,
                component.Make,
                component.Model,
                seq_num
            ]
            
            generated_desc = ", ".join([str(p).strip() for p in desc_parts if p])
            component.Component_Description = generated_desc

            try:
                component.save()
                messages.success(request, 'Component added successfully.')
                return redirect('equipment:equipment')
            except Exception as e:
                form.add_error('Component_Description', f"A component with this description already exists: {generated_desc}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    field_label = form.fields[field].label or field.replace('_', ' ').title()
                    messages.error(request, f"{field_label}: {error}")
                    
            for error in form.non_field_errors():
                messages.error(request, error)
    else:
        form = CompUploadForm(initial=initial_data)

    return render(request, 'equipment/add_comp.html', {
        'compuploadform': form,
        'item': item,
    })

def get_equipment_details(request):
    eq_number = request.GET.get('eq_num')
    try:
        equipment = Equipment.objects.get(Equipment_Number=eq_number)
        comp_types = ComponentType.objects.filter(
            asset_type=equipment.Asset_Type,
            is_active=True,
        ).order_by('name')
        comp_type_list = [
            {'id': ct.id, 'name': ct.name} for ct in comp_types
        ]
        return JsonResponse({
            'success': True,
            'id': equipment.pk,
            'number': equipment.Equipment_Number,
            'description': equipment.Equipment_Description,
            'asset_type_id': equipment.Asset_Type.id,
            'component_types': comp_type_list 
        })
    except Equipment.DoesNotExist:
        return JsonResponse({'success': False})

def get_next_component_id(request):
    eq_id = request.GET.get('eq_id')
    type_id = request.GET.get('type_id')
    
    if not eq_id or not type_id:
        return JsonResponse({'full_id': ''})

    equipment = get_object_or_404(Equipment, id=eq_id)
    comp_type = get_object_or_404(ComponentType, id=type_id)
    
    full_id = f"{equipment.Equipment_Number}-{comp_type.short_code}"
    
    return JsonResponse({'full_id': full_id})

def change_component(request):
    all_equipment = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')

    if request.method == 'POST':
        form = CompChangeForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    comp_id = form.cleaned_data['Component'].id
                    comp_slot = Component.objects.get(id=comp_id) 
                    
                    history = form.save(commit=False)
                    history.Old_Serial = comp_slot.Serial_Number
                    
                    history.New_PO = form.cleaned_data.get('New_PO')
                    
                    history.save()

                    comp_slot.Make = form.cleaned_data.get('New_Make')
                    comp_slot.Model = form.cleaned_data.get('New_Model')
                    comp_slot.Serial_Number = form.cleaned_data.get('New_Serial')
                    comp_slot.Installation_Date = form.cleaned_data.get('Change_Date')
                    comp_slot.PO_Number = form.cleaned_data.get('New_PO')
                    comp_slot.Expected_Lifespan = form.cleaned_data.get('New_Lifespan')
                    comp_slot.UoM = form.cleaned_data.get('New_UoM')
                    comp_slot.Warranty_Duration = form.cleaned_data.get('New_Wty_Dur')
                    comp_slot.Wty_UoM = form.cleaned_data.get('New_Wty_UoM')
                    comp_slot.Warranty_Start_Date = form.cleaned_data.get('New_Wty_Start')
                    comp_slot.Warranty_End_Date = form.cleaned_data.get('New_Wty_End')
                    
                    comp_slot.save()
                messages.success(request, 'Component updated successfully.')
                return redirect('equipment:equipment')
            
            except Exception as e:
                print(f"DATABASE ERROR: {e}")
                form.add_error(None, f"Internal Error: {e}")
    else:
        form = CompChangeForm()
    
    return render(request, 'equipment/change_comp.html', {
        'all_equipment': all_equipment,
        'compchangeform': form,
    })

def get_equipment_components(request):
    eq_num = request.GET.get('eq_num')
    try:
        equipment = Equipment.objects.get(Equipment_Number=eq_num)
        components = Component.objects.filter(Equipment=equipment).values('id', 'Component_Number', 'Component_Description')
        return JsonResponse({
            'success': True,
            'components': list(components)
        })
    except Equipment.DoesNotExist:
        return JsonResponse({'success': False})

def get_component_details_by_id(request):
    comp_id = request.GET.get('comp_id')
    try:
        c = Component.objects.get(id=comp_id)
        return JsonResponse({
            'id': c.id,
            'equipment_id': c.Equipment.id,
            'number': c.Component_Number,
            'description': c.Component_Description,
            'make': c.Make,
            'model': c.Model,
            'serial': c.Serial_Number,
            'po': c.PO_Number,
            'lifespan': c.Expected_Lifespan,
            'image_url': c.Component_Image.url if c.Component_Image else None,
        })
    except Component.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    
def search_component_history(request):
    history_query = ComponentHistory.objects.select_related('Equipment', 'Component').all()
    equipment_list = Equipment.objects.all().select_related('Asset_Type', 'Equipment_Type')

    eq_num = request.GET.get('Equipment_Number', '').strip()
    asset_type = request.GET.get('Asset_Type', '').strip()
    eq_type = request.GET.get('Equipment_Type', '').strip()
    status = request.GET.get('Equipment_Status', '').strip()
    make = request.GET.get('Make', '').strip()
    model = request.GET.get('Model', '').strip()

    if eq_num:
        history_query = history_query.filter(Equipment__Equipment_Number__icontains=eq_num)
    if asset_type:
        history_query = history_query.filter(Equipment__Asset_Type__name__icontains=asset_type)
    if eq_type:
        history_query = history_query.filter(Equipment__Equipment_Type__Equipment_Type__icontains=eq_type)
    if status:
        history_query = history_query.filter(Equipment__Equipment_Status__icontains=status)
    if make:
        history_query = history_query.filter(Equipment__Make__icontains=make)
    if model:
        history_query = history_query.filter(Equipment__Model__icontains=model)

    sort_by = request.GET.get('sort', '-Change_Date') 
    sort_mapping = {
        'Equipment_Number': 'Equipment__Equipment_Number',
        'Component_Number': 'Component__Component_Number',
        'Change_Date': 'Change_Date',
        'Change_Type': 'Change_Type',
        'Meter_Desc': 'Meter_Description',
        'Meter_Reading': 'Meter_Reading',
        'New_Make': 'New_Make',
        'New_Model': 'New_Model',
        'New_Serial': 'New_Serial',
        'Old_Serial': 'Old_Serial',
        'New_PO': 'New_PO',
        'New_Lifespan': 'New_Lifespan',
        'New_UoM': 'New_UoM',
        'New_Wty_Dur': 'New_Wty_Dur',
        'New_Wty_UoM': 'New_Wty_UoM',
        'New_Wty_Start': 'New_Wty_Start',
        'New_Wty_End': 'New_Wty_End',
    }
    
    is_descending = sort_by.startswith('-')
    clean_key = sort_by.lstrip('-')

    if clean_key in sort_mapping:
        order_field = f"-{sort_mapping[clean_key]}" if is_descending else sort_mapping[clean_key]
    else:
        order_field = '-Change_Date'

    history_query = history_query.order_by(order_field)
    
    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')
    all_asset_types = Equipment.objects.exclude(Asset_Type__name__isnull=True).values_list('Asset_Type__name', flat=True).distinct().order_by('Asset_Type__name')
    all_eq_types = Equipment.objects.exclude(Equipment_Type__Equipment_Type__isnull=True).values_list('Equipment_Type__Equipment_Type', flat=True).distinct().order_by('Equipment_Type__Equipment_Type')
    all_statuses = Equipment.objects.exclude(Equipment_Status__isnull=True).values_list('Equipment_Status', flat=True).distinct().order_by('Equipment_Status')
    all_makes = Equipment.objects.exclude(Make__isnull=True).values_list('Make', flat=True).distinct().order_by('Make')
    all_models = Equipment.objects.exclude(Model__isnull=True).values_list('Model', flat=True).distinct().order_by('Model')


    return render(request, 'equipment/search_comp.html', {
        'equipment_list': history_query,
        'sort': sort_by,
        'all_equipment_suggestions': all_equipment_suggestions,
        'all_asset_types': all_asset_types,
        'all_eq_types': all_eq_types,
        'all_statuses': all_statuses,
        'all_makes': all_makes,
        'all_models': all_models,
        'eq_num_val': eq_num,
        'asset_type_val': asset_type,
        'eq_type_val': eq_type,
        'status_val': status,
        'make_val': make,
        'model_val': model,
    })
    
def export_component_history(request):
    history_query = ComponentHistory.objects.select_related('Equipment', 'Component').all()

    eq_num = request.GET.get('Equipment_Number')
    if eq_num:
        history_query = history_query.filter(Equipment__Equipment_Number__icontains=eq_num)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Component History"

    headers = [
        'Equipment Number', 'Component Number', 'Change Date', 'Change Type', 
        'Meter Reading', 'New Serial', 'Old Serial', 'PO'
    ]
    ws.append(headers)

    for item in history_query:
        ws.append([
            item.Equipment.Equipment_Number,
            item.Component.Component_Number,
            item.Change_Date,
            item.Change_Type,
            item.Meter_Reading,
            item.New_Serial,
            item.Old_Serial,
            item.New_PO
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Component_History.xlsx"'
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    wb.save(response)
    return response

def _build_shift_report_context(form, report=None, request=None):
    reports = ShiftReport.objects.all().order_by('-date', '-id')
    asset_types = Equipment.objects.values_list('Asset_Type__name', flat=True).distinct().order_by('Asset_Type__name')
    garages = Equipment.objects.values_list('Garage__Facility_Name', flat=True).distinct().order_by('Garage__Facility_Name')
    equipment_list = Equipment.objects.all().select_related('Equipment_Type').order_by('Equipment_Type__Equipment_Type', 'Equipment_Number')
    hours_range = range(6, 18)
    range_48 = range(48)
    
    if request:
        asset_filter = request.GET.get('asset_type') or request.POST.get('asset_type')
        garage_filter = request.GET.get('garage') or request.POST.get('garage')
        
        if asset_filter:
            equipment_list = equipment_list.filter(Asset_Type__name=asset_filter)
        if garage_filter:
            equipment_list = equipment_list.filter(Garage__Facility_Name=garage_filter)

    hours_range = range(6, 18)
    range_49 = range(48)
    saved_data = {}
    
    if report:
        statuses = report.statuses.select_related('equipment').all()
        for status in statuses:
            saved_data[status.equipment.Equipment_Number] = status.grid_data.split(',') if status.grid_data else []
            
    context = {
        'form': form,
        'report': report,
        'is_edit': report is not None,
        'reports': reports,
        'equipment_list': equipment_list,
        'asset_types': asset_types,
        'garages': garages,
        'hours_range': hours_range,
        'range_48': range_48,
        'saved_data': saved_data,
    }
    return context

def _save_machine_statuses(report, machines_json):
    if not machines_json:
        return
    try:
        machines = json.loads(machines_json)
    except json.JSONDecodeError:
        machines = []
        
    report.statuses.all().delete()
    for machine in machines:
        equipment_number = machine.get('id')
        if not equipment_number:
            continue
        equipment = Equipment.objects.filter(Equipment_Number=equipment_number).first()
        if not equipment:
            continue
        MachineShiftStatus.objects.create(
            report=report,
            equipment=equipment,
            total_down=float(machine.get('td') or 0),
            total_worked=float(machine.get('tw') or 0),
            available=float(machine.get('av') or 0),
            final_status=machine.get('status') or 'Available',
            grid_data=machine.get('grid') or ''
        )
        
def shift_report(request):
    if request.method == 'POST':
        date = request.POST.get('date')
        shift = request.POST.get('shift')
        machines_json = request.POST.get('machines_json') or request.GET.get('machines_json')
        
        existing = ShiftReport.objects.filter(date=date, shift=shift).first()
        if existing:
            return redirect('equipment:shift_report_edit', pk=existing.pk)
        
        form = ShiftReportForm(request.POST)
        if form.is_valid():
            report = form.save()
            _save_machine_statuses(report, machines_json)
            messages.success(request,"Status Report created successfully!")
            return redirect('equipment:equipment')
        else:
            messages.error(request, "There was an error saving the status report!")
    else:
        form = ShiftReportForm()
        
    context = _build_shift_report_context(form=form, report=None, request=request)
    return render(request, 'equipment/shift_report.html', context)

def shift_report_edit(request, pk):
    report = get_object_or_404(ShiftReport, pk=pk)
    if request.method == 'POST':
        form = ShiftReportForm(request.POST, instance=report)
        if form.is_valid():
            report = form.save()
            _save_machine_statuses(report, request.POST.get('machines_json'))
            messages.success(request,"Status Report updated successfully!")
            return redirect('equipment:equipment')
        else:
            messages.error(request, "There was an error saving the status report!")
    else:
        form = ShiftReportForm(instance=report)
    context = _build_shift_report_context(form=form, report=report, request=request)
    return render(request, 'equipment/shift_report.html', context)

def export_shift_report_excel(request, report_id):
    report = get_object_or_404(ShiftReport, pk=report_id)
    asset_filter = request.GET.get('asset_type')
    garage_filter = request.GET.get('garage')
    all_equipment = Equipment.objects.all()
    
    if asset_filter:
        all_equipment = all_equipment.filter(Asset_Type__name=asset_filter)
    if garage_filter:
        all_equipment = all_equipment.filter(Garage__Facility_Name=garage_filter)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report_{report.date}".replace("/", "-")[:30]

    headers = [
        'Date', 'Shift', 'Equipment', 'Asset Type', 'Garage', 
        'Mining Supervisor', 'Maintenance Supervisor', 
        'Total Downtime', 'Available Time', 'Final Status', 'Grid Data'
    ]
    ws.append(headers)
    
    r_date = report.date.strftime('%Y-%m-%d') if report.date else ''
    for equip in all_equipment:
        status = report.statuses.filter(equipment=equip).first()
        
        # Access related names safely
        e_type = equip.Asset_Type.name if equip.Asset_Type else ''
        e_garage = equip.Garage.Facility_Name if equip.Garage else ''

        row = [
            r_date, report.shift, equip.Equipment_Number, e_type, e_garage,
            report.mining_supervisor, report.maint_supervisor,
        ]
        
        if status:
            row.extend([status.total_down, status.available, status.final_status, status.grid_data])
        else:
            row.extend([0.0, 12.0, "Available"])
            
        ws.append(row)
            
    for i, _ in enumerate(ws.columns, 1):
        column_letter = get_column_letter(i)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 2
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Shift_Report_{r_date}_{report.shift}.xlsx".replace(" ", "_")
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_shift_archive_excel(request):
    asset_filter = request.GET.get('asset_type')
    garage_filter = request.GET.get('garage')
    reports = ShiftReport.objects.all().order_by('-date', 'shift')
    all_equipment = Equipment.objects.all().select_related('Asset_Type', 'Garage')
    
    if asset_filter:
        all_equipment = all_equipment.filter(Asset_Type__name=asset_filter)
    if garage_filter:
        all_equipment = all_equipment.filter(Garage__Facility_Name=garage_filter)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipment_Archive"
    headers = [
        'Date', 'Shift', 'Mining Supervisor', 'Maintenance Supervisor', 
        'Equipment', 'Asset Type', 'Garage',
        'Total Downtime', 'Available Time', 'Final Status', 'Grid Data'
    ]
    ws.append(headers)
    
    for report in reports:
        r_date = report.date.strftime('%Y-%m-%d') if report.date else ''
        for equip in all_equipment:
            status = report.statuses.filter(equipment=equip).first()
            e_type = equip.Asset_Type.name if equip.Asset_Type else ''
            e_garage = equip.Garage.Facility_Name if equip.Garage else ''
            
            row = [
                r_date, report.shift, report.mining_supervisor, report.maint_supervisor,
                equip.Equipment_Number, e_type, e_garage,
            ]
            
            if status:
                row.extend([status.total_down, status.available, status.final_status, status.grid_data])
            else:
                row.extend([0.0, 12.0, "Available"])
            ws.append(row)
            
    for i, _ in enumerate(ws.columns, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = 20 # Start with a fixed width to test

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Equipment_Archive.xlsx"'
    return response

def search_comp_list(request):
    comp_list = Component.objects.select_related(
        'Equipment', 
        'Equipment__Asset_Type', 
        'Equipment__Equipment_Type',
    ).all()
    
    equip_val = request.GET.get('Equipment_Number', '').strip()
    asset_type_val = request.GET.get('Asset_Type', '').strip()
    equip_type_val = request.GET.get('Equipment_Type', '').strip()
    comp_num_val = request.GET.get('Component_Number', '').strip()
    comp_desc_val = request.GET.get('Component_Description', '').strip()
    make_val = request.GET.get('Make', '').strip()
    model_val = request.GET.get('Model', '').strip()
    
    if equip_val:
        comp_list = comp_list.filter(Equipment__Equipment_Number__icontains=equip_val)
    if asset_type_val:
        comp_list = comp_list.filter(Equipment__Asset_Type__name__icontains=asset_type_val)
    if equip_type_val:
        comp_list = comp_list.filter(Equipment__Equipment_Type__Equipment_Type__icontains=equip_type_val)
    if comp_num_val:
        comp_list = comp_list.filter(Component_Number__icontains=comp_num_val)
    if comp_desc_val:
        comp_list = comp_list.filter(Component_Description__icontains=comp_desc_val)
    if make_val:
        comp_list = comp_list.filter(Make__icontains=make_val)
    if model_val:
        comp_list = comp_list.filter(Model__icontains=model_val)
            
    sort_by = request.GET.get('sort', 'Component_Number')
    is_descending = sort_by.startswith('-')
    clean_sort_key = sort_by.lstrip('-')
    
    sort_mapping = {
        'Equipment_Number': 'Equipment__Equipment_Number',
        'Asset_Type': 'Equipment__Asset_Type__name',
        'Equipment_Type': 'Equipment__Equipment_Type__Equipment_Type',
        'Component_Number': 'Component_Number',
        'Component_Description': 'Component_Description',
        'Make': 'Make',
        'Model': 'Model',
        'Serial_Number': 'Serial_Number',
        'Warranty_Duration': 'Warranty_Duration',
        'Wty_UoM': 'Wty_UoM',
        'Warranty_Start_Date': 'Warranty_Start_Date',
        'Warranty_End_Date': 'Warranty_End_Date',
        'PO_Number': 'PO_Number',
    }

    if clean_sort_key in sort_mapping:
        db_field = sort_mapping[clean_sort_key]
        order_field = f"-{db_field}" if is_descending else db_field
        comp_list = comp_list.order_by(order_field)
    else:
        comp_list = comp_list.order_by('Component_Number')

    all_equipment_suggestions = Equipment.objects.all().only('Equipment_Number', 'Equipment_Description').order_by('Equipment_Number')
    
    all_component_suggestions = Component.objects.all().only('Component_Number', 'Component_Description').order_by('Component_Number')
    
    all_asset_types = Equipment.objects.exclude(Asset_Type__name__isnull=True).values_list('Asset_Type__name', flat=True).distinct().order_by('Asset_Type__name')
    all_eq_types = Equipment.objects.exclude(Equipment_Type__Equipment_Type__isnull=True).values_list('Equipment_Type__Equipment_Type', flat=True).distinct().order_by('Equipment_Type__Equipment_Type')
    all_makes = Component.objects.exclude(Make__isnull=True).values_list('Make', flat=True).distinct().order_by('Make')
    all_models = Component.objects.exclude(Model__isnull=True).values_list('Model', flat=True).distinct().order_by('Model')


    context = {
        'comp_list': comp_list,
        'sort': sort_by,
        'all_equipment_suggestions': all_equipment_suggestions,
        'all_component_suggestions': all_component_suggestions,
        'all_asset_types': all_asset_types,
        'all_eq_types': all_eq_types,
        'all_makes': all_makes,
        'all_models': all_models,
        'equip_val': equip_val,
        'asset_type_val': asset_type_val,
        'equip_type_val': equip_type_val,
        'comp_num_val': comp_num_val,
        'comp_desc_val': comp_desc_val,
        'make_val': make_val,
        'model_val': model_val,
    }
    
    return render(request, 'equipment/search_comp_list.html', context)

def export_list_excel(request):
    comp_list = Component.objects.select_related(
        'Equipment', 
        'Equipment__Asset_Type', 
        'Equipment__Equipment_Type'
    ).all()
    
    equip_val = request.GET.get('Equipment_Number')
    asset_type_val = request.GET.get('Asset_Type')
    equip_type_val = request.GET.get('Equipment_Type')
    comp_num_val = request.GET.get('Component_Number')
    comp_desc_val = request.GET.get('Component_Description')
    make_val = request.GET.get('Make')
    model_val = request.GET.get('Model')
    
    if equip_val:
        comp_list = comp_list.filter(Equipment__Equipment_Number__icontains=equip_val)
    if asset_type_val:
        comp_list = comp_list.filter(Equipment__Asset_Type__name__icontains=asset_type_val)
    if equip_type_val:
        comp_list = comp_list.filter(Equipment__Equipment_Type__Equipment_Type__icontains=equip_type_val)
    if comp_num_val:
        comp_list = comp_list.filter(Component_Number__icontains=comp_num_val)
    if comp_desc_val:
        comp_list = comp_list.filter(Component_Description__icontains=comp_desc_val)
    if make_val:
        comp_list = comp_list.filter(Make__icontains=make_val)
    if model_val:
        comp_list = comp_list.filter(Model__icontains=model_val)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Component_List"
    headers = [
        'Equipment Number', 'Asset Type', 'Component Number', 'Component Description',  
        'Make', 'Model', 'Serial Number', 'Warranty Duration', 'Warranty UoM', 'Warranty Start Date', 'Warranty End Date', 
        'PO Number', 'Expected Lifespan', 'UoM',
    ]
    ws.append(headers)
    
    for comp in comp_list:
        ws.append([
            comp.Equipment.Equipment_Number if comp.Equipment else '',
            comp.Equipment.Asset_Type.name if comp.Equipment and comp.Equipment.Asset_Type else '',
            comp.Component_Number,
            comp.Component_Description,
            comp.Make,
            comp.Model,
            comp.Serial_Number,
            comp.Warranty_Duration,
            comp.Wty_UoM,
            comp.Warranty_Start_Date.strftime('%Y-%m-%d') if comp.Warranty_Start_Date else '',
            comp.Warranty_End_Date.strftime('%Y-%m-%d') if comp.Warranty_End_Date else '',
            comp.PO_Number,
            comp.Expected_Lifespan,
            comp.UoM,
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Component_List.xlsx"'
    return response

def asset_type_table(request):
    asset_types = AssetType.objects.all().order_by("id")
    form = AssetTypeForm()
    edit_id = None
    editing_asset_type = None
    if request.method == "POST":
        action = request.POST.get("action")
        asset_type_id = request.POST.get("asset_type_id")
        name = request.POST.get("name", "").strip()
        if action == "save":
            if not name:
                messages.error(request, "Asset type name is required.")
                return redirect("equipment:asset_type_table")
            if asset_type_id:
                asset_type = get_object_or_404(AssetType, id=asset_type_id)
                duplicate_asset_type = AssetType.objects.filter(
                    name__iexact=name
                ).exclude(
                    id=asset_type.id
                ).first()
                if duplicate_asset_type:
                    messages.error(request, "That asset type already exists.")
                    return redirect("equipment:asset_type_table")
                asset_type.name = name
                asset_type.is_active = request.POST.get("is_active") == "on"
                asset_type.save()
                messages.success(request, "Successfully updated entry.")
                return redirect("equipment:asset_type_table")
            existing_asset_type = AssetType.objects.filter(name__iexact=name).first()
            if existing_asset_type:
                if not existing_asset_type.is_active:
                    existing_asset_type.is_active = True
                    existing_asset_type.save()
                    messages.success(request, "Successfully reactivated existing entry.")
                    return redirect("equipment:asset_type_table")
                messages.error(request, "That asset type already exists.")
                return redirect("equipment:asset_type_table")
            AssetType.objects.create(
                name=name,
                is_active=True,
            )
            messages.success(request, "Successfully added entry.")
            return redirect("equipment:asset_type_table")
        if action == "deactivate":
            inactive_ids = request.POST.getlist("inactive_ids")
            if inactive_ids:
                updated_count = AssetType.objects.filter(id__in=inactive_ids).update(
                    is_active=False
                )
                if updated_count == 1:
                    messages.success(request, "Successfully set entry inactive.")
                else:
                    messages.success(request, f"Successfully set {updated_count} entries inactive.")
            else:
                messages.warning(request, "No entries were selected.")
            return redirect("equipment:asset_type_table")
    if request.GET.get("edit"):
        edit_id = request.GET.get("edit")
        editing_asset_type = get_object_or_404(AssetType, id=edit_id)
        form = AssetTypeForm(instance=editing_asset_type)
    context = {
        "form": form,
        "asset_types": asset_types,
        "edit_id": edit_id,
        "editing_asset_type": editing_asset_type,
    }
    return render(request, "equipment/asset_type_table.html", context)

def eq_type_table(request):
    eq_types = EQ_Type.objects.select_related("Asset_Type").all().order_by(
        "Asset_Type__name",
        "Equipment_Type",
    )
    form = EQTypeForm()
    edit_id = None
    editing_eq_type = None
    if request.method == "POST":
        action = request.POST.get("action")
        eq_type_id = request.POST.get("eq_type_id")
        if action == "save":
            if eq_type_id:
                editing_eq_type = get_object_or_404(EQ_Type, id=eq_type_id)
                form = EQTypeForm(request.POST, instance=editing_eq_type)
            else:
                form = EQTypeForm(request.POST)
            if form.is_valid():
                eq_type = form.save(commit=False)
                eq_type.is_active = request.POST.get("is_active") == "on" if eq_type_id else True
                duplicate_prefix = EQ_Type.objects.filter(
                    Prefix__iexact=eq_type.Prefix
                )
                if eq_type_id:
                    duplicate_prefix = duplicate_prefix.exclude(id=eq_type.id)
                if duplicate_prefix.exists():
                    messages.error(request, "That prefix already exists.")
                    return redirect("equipment:eq_type_table")
                duplicate_equipment_type = EQ_Type.objects.filter(
                    Asset_Type=eq_type.Asset_Type,
                    Equipment_Type__iexact=eq_type.Equipment_Type,
                )
                if eq_type_id:
                    duplicate_equipment_type = duplicate_equipment_type.exclude(id=eq_type.id)
                if duplicate_equipment_type.exists():
                    messages.error(request, "That equipment type already exists for this asset type.")
                    return redirect("equipment:eq_type_table")
                try:
                    eq_type.save()
                except IntegrityError:
                    messages.error(request, "Unable to save. Please check for duplicate values.")
                    return redirect("equipment:eq_type_table")
                if eq_type_id:
                    messages.success(request, "Successfully updated entry.")
                else:
                    messages.success(request, "Successfully added entry.")
                return redirect("equipment:eq_type_table")
            messages.error(request, "Please correct the errors below.")
        if action == "deactivate":
            inactive_ids = request.POST.getlist("inactive_ids")
            if inactive_ids:
                updated_count = EQ_Type.objects.filter(id__in=inactive_ids).update(
                    is_active=False
                )
                if updated_count == 1:
                    messages.success(request, "Successfully set entry inactive.")
                else:
                    messages.success(request, f"Successfully set {updated_count} entries inactive.")
            else:
                messages.warning(request, "No entries were selected.")
            return redirect("equipment:eq_type_table")
    if request.GET.get("edit"):
        edit_id = request.GET.get("edit")
        editing_eq_type = get_object_or_404(EQ_Type, id=edit_id)
        form = EQTypeForm(instance=editing_eq_type)
    context = {
        "form": form,
        "eq_types": eq_types,
        "edit_id": edit_id,
        "editing_eq_type": editing_eq_type,
    }
    return render(request, "equipment/eq_type_table.html", context)

def component_type_table(request):
    component_types = ComponentType.objects.select_related("asset_type").all().order_by(
        "asset_type__name",
        "name",
    )
    form = ComponentTypeForm()
    edit_id = None
    editing_component_type = None
    if request.method == "POST":
        action = request.POST.get("action")
        component_type_id = request.POST.get("component_type_id")
        if action == "save":
            if component_type_id:
                editing_component_type = get_object_or_404(
                    ComponentType,
                    id=component_type_id,
                )
                form = ComponentTypeForm(request.POST, instance=editing_component_type)
            else:
                form = ComponentTypeForm(request.POST)
            if form.is_valid():
                component_type = form.save(commit=False)
                if component_type_id:
                    component_type.is_active = request.POST.get("is_active") == "on"
                else:
                    component_type.is_active = True
                duplicate_code = ComponentType.objects.filter(
                    short_code__iexact=component_type.short_code
                )
                if component_type_id:
                    duplicate_code = duplicate_code.exclude(id=component_type.id)
                if duplicate_code.exists():
                    messages.error(request, "That code already exists.")
                    return redirect("equipment:component_type_table")
                duplicate_component_type = ComponentType.objects.filter(
                    asset_type=component_type.asset_type,
                    name__iexact=component_type.name,
                )
                if component_type_id:
                    duplicate_component_type = duplicate_component_type.exclude(
                        id=component_type.id
                    )
                if duplicate_component_type.exists():
                    messages.error(
                        request,
                        "That component type already exists for this asset type.",
                    )
                    return redirect("equipment:component_type_table")
                try:
                    component_type.save()
                except IntegrityError:
                    messages.error(
                        request,
                        "Unable to save. Please check for duplicate values.",
                    )
                    return redirect("equipment:component_type_table")
                if component_type_id:
                    messages.success(request, "Successfully updated entry.")
                else:
                    messages.success(request, "Successfully added entry.")
                return redirect("equipment:component_type_table")
            messages.error(request, "Please correct the errors below.")
        if action == "deactivate":
            inactive_ids = request.POST.getlist("inactive_ids")
            if inactive_ids:
                updated_count = ComponentType.objects.filter(
                    id__in=inactive_ids
                ).update(
                    is_active=False
                )
                if updated_count == 1:
                    messages.success(request, "Successfully set entry inactive.")
                else:
                    messages.success(
                        request,
                        f"Successfully set {updated_count} entries inactive.",
                    )
            else:
                messages.warning(request, "No entries were selected.")
            return redirect("equipment:component_type_table")
    if request.GET.get("edit"):
        edit_id = request.GET.get("edit")
        editing_component_type = get_object_or_404(ComponentType, id=edit_id)
        form = ComponentTypeForm(instance=editing_component_type)
    context = {
        "form": form,
        "component_types": component_types,
        "edit_id": edit_id,
        "editing_component_type": editing_component_type,
    }
    return render(request, "equipment/component_type_table.html", context)