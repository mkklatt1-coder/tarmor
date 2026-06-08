from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime, time
from django.db.models import Sum, Q, Value
from django.db.models.functions import Concat
from .models import WorkWeek, Schedule, ScheduleSnapshot, DailyCrewCapacity, WeekSetup, TimeOffLog
from .forms import WeekSetupForm
from work_orders.models import WorkOrder
from facilities.models import Facility
from personnel.models import CrewShiftRotation, Employee, Crew
from collections import defaultdict
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.utils.timezone import make_aware
from django.db import transaction
import io
import json

def scheduling(request):
    return render(request, 'scheduling/scheduling.html')

def scheduling_view(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            date_str = data.get('date')
            garage_id = data.get('garage_id')
            shift_type = data.get('shift_type')
            hours = int(data.get('hours', 0) or 0)

            target_facility = Facility.objects.get(id=garage_id)
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()

            log, created = TimeOffLog.objects.get_or_create(
                facility=target_facility,
                date=target_date
            )
            if shift_type == 'DS':
                log.ds_off = hours
            else:
                log.ns_off = hours
            log.save()
            return JsonResponse({'success': True, 'msg': 'Time off updated.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    week_num = request.GET.get("week")
    garage_id = request.GET.get("garage")
    weeks = WorkWeek.objects.all().order_by("week_number")
    garages = Facility.objects.all().order_by("Facility_Name")
    selected_week = WorkWeek.objects.filter(week_number=week_num).first() or weeks.first()
    selected_garage = Facility.objects.filter(id=garage_id).first() or garages.first()
    planned_ds = [0.0 for _ in range(7)]
    planned_ns = [0.0 for _ in range(7)]
    available_ds = [0.0 for _ in range(7)]
    available_ns = [0.0 for _ in range(7)]
    week_days = []
    week_orders = WorkOrder.objects.none()
    day_rows = []

    if selected_week:
        start = selected_week.start_date.date() if hasattr(selected_week.start_date, "date") else selected_week.start_date
        end = selected_week.end_date.date() if hasattr(selected_week.end_date, "date") else selected_week.end_date
        week_days = [start + timedelta(days=i) for i in range(7)]
        if selected_garage:
            week_orders = WorkOrder.objects.filter(
                equipment__Garage=selected_garage,
                plan_start_date__range=(selected_week.start_date, selected_week.end_date),
                date_closed__isnull=True,
                job_status__status_choice__in=["Waiting to Schedule", "Reschedule", "Execution"],
            ).select_related("equipment", "job_status").order_by("plan_start_date")

            # Planned hours
            for wo in week_orders:
                if wo.plan_start_date:
                    wo_date = wo.plan_start_date.date() if hasattr(wo.plan_start_date, "date") else wo.plan_start_date
                    offset = (wo_date - start).days
                    if 0 <= offset < 7:
                        shift = (wo.plan_shift or "DS").upper()
                        if shift == "NS":
                            planned_ns[offset] += float(wo.est_work_hours or 0)
                        else:
                            planned_ds[offset] += float(wo.est_work_hours or 0)

            # Available hours
            facility_code = str(selected_garage.Facility_Code).strip()
            crews = Crew.objects.filter(location_code=facility_code).select_related('rotation', 'pattern')
            crew_data = [{
                'instance': crew,
                'emp_count': Employee.objects.filter(crew=crew).count(),
                'hrs_per_shift': float(crew.rotation.hrs_per_shift or 0) if crew.rotation else 0.0
            } for crew in crews]

            time_off_map = {
                log.date: log 
                for log in TimeOffLog.objects.filter(facility=selected_garage, date__range=(week_days[0], week_days[-1]))
            }

            for i in range(7):
                date_i = week_days[i]
                for data in crew_data:
                    crew = data['instance']
                    shift_status = crew.get_status_for_date(date_i)
                    daily_available_hrs = data['emp_count'] * data['hrs_per_shift']

                    if shift_status == "DAY":
                        available_ds[i] += daily_available_hrs
                    elif shift_status == "NIGHT":
                        available_ns[i] += daily_available_hrs

                day_log = time_off_map.get(date_i)
                ds_off_hrs = day_log.ds_off if day_log else 0
                ns_off_hrs = day_log.ns_off if day_log else 0

                final_avail_ds = max(available_ds[i] - float(ds_off_hrs), 0.0)
                final_avail_ns = max(available_ns[i] - float(ns_off_hrs), 0.0)

                day_rows.append({
                    "day": date_i,
                    "date_string": date_i.strftime("%Y-%m-%d"),
                    "ds_off": ds_off_hrs,
                    "ns_off": ns_off_hrs,
                    "ds_planned": planned_ds[i],
                    "ds_available": final_avail_ds,
                    "ns_planned": planned_ns[i],
                    "ns_available": final_avail_ns,
                    "ds_over": planned_ds[i] > final_avail_ds,
                    "ns_over": planned_ns[i] > final_avail_ns,
                    "total_planned": planned_ds[i] + planned_ns[i],
                    "total_available": final_avail_ds + final_avail_ns,
                })
    context = {
        "weeks": weeks,
        "garages": garages,
        "selected_week": selected_week,
        "selected_garage": selected_garage,
        "week_days": week_days,
        "work_orders": week_orders,
        "day_rows": day_rows,
    }
    return render(request, "scheduling/schedule.html", context)

def gantt_view(request, week, garage_id):
    week_obj = get_object_or_404(WorkWeek, week_number=week)
    garage = get_object_or_404(Facility, id=garage_id)

    work_orders = WorkOrder.objects.filter(
        plan_start_date__range=(week_obj.start_date, week_obj.end_date),
        equipment__Garage=garage,
        date_closed__isnull=True,
    ).select_related('equipment')
    
    data = [
        {
            "task": str(wo.work_order),
            "start": wo.plan_start_date.strftime("%Y-%m-%d") if wo.plan_start_date else "",
            "duration": float(wo.est_work_hours or 0) / 8,
            "description": wo.repair_description or "",
        }
        for wo in work_orders
    ]

    context = {
        "week": week_obj,
        "garage": garage,
        "data": data
    }
    
    return render(request, "scheduling/gantt.html", context)

def forecast_view(request):
    today = timezone.now().date()
    garages = Facility.objects.all().order_by("Facility_Name")
    garage_id = request.GET.get("garage")
    current_week = WorkWeek.objects.filter(start_date__lte=today, end_date__gte=today).first()
    start_week_num = (current_week.week_number + 1) if current_week else 1

    upcoming_weeks = WorkWeek.objects.filter(
        week_number__range=(start_week_num, start_week_num + 2)
    ).order_by('week_number')
    
    forecast_data = []

    if garage_id:
        selected_garage = Facility.objects.filter(pk=garage_id).first()
    else:
        selected_garage = garages.first()
        
        current_week = WorkWeek.objects.filter(start_date__lte=today, end_date__gte=today).first()
        start_week_num = (current_week.week_number + 1) if current_week else 1
        
        upcoming_weeks = WorkWeek.objects.filter(
            week_number__range=(start_week_num, start_week_num + 2)
        ).order_by('week_number')
        
        forecast_data = []

    if selected_garage:
        facility_code = str(selected_garage.Facility_Code)
        crews = Crew.objects.filter(location_code=facility_code)
        employee_counts = {c.pk: Employee.objects.filter(crew=c).count() for c in crews}
        
        rotations = CrewShiftRotation.objects.filter(Location=selected_garage)
        rotations_by_letter = {r.Shift_ID.split("-")[-1]: r for r in rotations}
        
        for wk in upcoming_weeks:
            week_orders = WorkOrder.objects.filter(
            equipment__Garage=selected_garage,
            plan_start_date__range=(wk.start_date, wk.end_date),
            date_closed__isnull=True,
            job_status__status_choice__in=["Waiting to Schedule", "Reschedule", "Execution"],
            ).select_related("equipment", "job_status").order_by("plan_start_date")

            total_est = week_orders.aggregate(Sum("est_work_hours"))["est_work_hours__sum"] or 0

            total_available = 0.0
            for i in range(7):
                date_i = wk.start_date + timedelta(days=i)
                for crew in crews:
                    # Logic check: ensuring crew status matches DAY/NIGHT
                    if crew.get_status_for_date(date_i) in ["DAY", "NIGHT"]:
                        rotation = rotations_by_letter.get(crew.shift_letter)
                        if rotation:
                            hrs = float(rotation.hrs_per_shift or 0)
                            total_available += (employee_counts.get(crew.pk, 0) * hrs)


            forecast_data.append({
                "week": wk,
                "estimated_hours": float(total_est),
                "available_hours": total_available,
                "sched_load": round((float(total_est) / total_available * 100), 1) if total_available > 0 else 0,
                "work_orders": week_orders,
            })

    context = {
        "forecast_data": forecast_data,
        "garages": garages,
        "selected_garage": selected_garage,
    }
    return render(request, "scheduling/forecast.html", context)

@require_POST
def update_workorder_date(request, pk):
    wo = get_object_or_404(WorkOrder, pk=pk)
    updated_fields = []
    new_date = request.POST.get("plan_start_date")
    if new_date:
        wo.plan_start_date = new_date
        updated_fields.append("plan_start_date")
    new_shift = request.POST.get("plan_shift")
    if new_shift:
        wo.plan_shift = new_shift
        updated_fields.append("plan_shift")
    if updated_fields:
        wo.save(update_fields=updated_fields)
    return JsonResponse({
        "ok": True,
        "plan_start_date": wo.plan_start_date.strftime("%Y-%m-%d") if wo.plan_start_date else "",
        "plan_shift": getattr(wo, "plan_shift", ""),
    })

def weeksetup_view(request):
    """Single page: create/update WeekSetup + display weeks."""
    active = WeekSetup.objects.filter(active=True).first()
    setup = active or WeekSetup()
    if request.method == "POST":
        form = WeekSetupForm(request.POST, instance=setup)
        if form.is_valid():
            form.save()                       # rebuilds WorkWeek table
            return redirect("scheduling:scheduling")   # ✅ after save
    else:
        form = WeekSetupForm(instance=setup)
    today = timezone.now().date()
    weeks, current_week = [], None
    if active:
        all_weeks = WorkWeek.objects.filter(setup=active).order_by("week_number")
        current_week = all_weeks.filter(start_date__lte=today, end_date__gte=today).first()
        if current_week:
            start_range = max(1, current_week.week_number - 4)
            end_range = min(52, current_week.week_number + 26)
            weeks = all_weeks.filter(week_number__range=(start_range, end_range))
    context = {"form": form, "active": active, "weeks": weeks, "current_week": current_week}
    return render(request, "scheduling/weeksetup.html", context)

def export_weeks_excel(request):
    import pandas as pd
    active = WeekSetup.objects.filter(active=True).first()
    if not active:
        return HttpResponse("No active week setup to export.", content_type="text/plain")
    weeks = WorkWeek.objects.filter(setup=active).order_by("week_number")
    df = pd.DataFrame(
        [{"Week": w.week_number, "Start Date": w.start_date, "End Date": w.end_date} for w in weeks]
    )
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="WorkWeeks")
    buffer.seek(0)
    filename = f"WorkWeeks_{active.week1_start_date}.xlsx"
    response = HttpResponse(
        buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

@require_POST
def save_working_copy(request):
    week = request.POST.get('week')
    garage_id = request.POST.get('garage')
    
    for key, value in request.POST.items():
        if key.startswith('date_'):
            wo_id = key.replace('date_', '')
            if value:
                WorkOrder.objects.filter(id=wo_id).update(plan_start_date=value)
    messages.success(request, "Schedule working copy saved successfully!")

    return redirect('scheduling:scheduling')

def export_forecast_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    today = timezone.now().date()
    garage_id = request.GET.get("garage")
    garages = Facility.objects.all().order_by("Facility_Name")
    selected_garage = Facility.objects.filter(pk=garage_id).first()

    if not selected_garage:
        selected_garage = garages.first()
    
    if not selected_garage:
        return HttpResponse("No garages found.", status=404)

    current_week = WorkWeek.objects.filter(start_date__lte=today, end_date__gte=today).first()
    start_week_num = (current_week.week_number + 1) if current_week else 1
    upcoming_weeks = WorkWeek.objects.filter(week_number__range=(start_week_num, start_week_num + 2)).order_by('week_number')

    wb = Workbook()
    ws = wb.active
    if ws is None:
        return HttpResponse("Could not create spreadsheet.", status=500)
    ws.title = "3-Week Forecast"

    header_fill = PatternFill(start_color="555555", end_color="555555", fill_type="solid")
    subheader_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
    white_text = Font(color="FFFFFF", bold=True)

    row_num = 1

    facility_code = str(selected_garage.Facility_Code)
    crews = Crew.objects.filter(location_code=facility_code)
    employee_counts = {c.pk: Employee.objects.filter(crew=c).count() for c in crews}
    rotations = CrewShiftRotation.objects.filter(Location=selected_garage)
    rotations_by_letter = {r.Shift_ID.split("-")[-1]: r for r in rotations}

    for wk in upcoming_weeks:
        total_available = 0.0
        for i in range(7):
            date_i = wk.start_date + timedelta(days=i)
            for crew in crews:
                if crew.get_status_for_date(date_i) in ["DAY", "NIGHT"]:
                    rotation = rotations_by_letter.get(crew.shift_letter)
                    if rotation:
                        hrs = float(rotation.hrs_per_shift or 0)
                        total_available += (employee_counts.get(crew.pk, 0) * hrs)

        week_orders = WorkOrder.objects.filter(
            equipment__Garage=selected_garage,
            plan_start_date__range=(wk.start_date, wk.end_date),
            date_closed__isnull=True,
            job_status__status_choice__in=["Waiting to Schedule", "Reschedule", "Execution"],
        ).select_related("equipment", "job_status")
        total_est = week_orders.aggregate(Sum("est_work_hours"))["est_work_hours__sum"] or 0

        summary_data = [
            f"Week {wk.week_number} ({wk.start_date} to {wk.end_date})",
            "Available:", total_available,
            "Planned:", float(total_est),
            "Load:", f"{(float(total_est)/total_available*100 if total_available > 0 else 0):.1f}%"
        ]
        
        for col_num, value in enumerate(summary_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = header_fill
            cell.font = white_text
        row_num += 1

        cols = ["Work Order", "Eq Number", "Eq Description", "Type", "Description", "Status", "Start Date", "Hours"]
        for col_num, column_title in enumerate(cols, 1):
            cell = ws.cell(row=row_num, column=col_num, value=column_title)
            cell.fill = subheader_fill
            cell.font = white_text
        row_num += 1

        for wo in week_orders:
            ws.cell(row=row_num, column=1, value=str(wo.work_order))
            ws.cell(row=row_num, column=2, value=str(wo.equipment.Equipment_Number ))
            ws.cell(row=row_num, column=3, value=str(wo.equipment.Equipment_Description))
            ws.cell(row=row_num, column=4, value=str(wo.work_type)) 
            ws.cell(row=row_num, column=5, value=str(wo.repair_description or ""))
            ws.cell(row=row_num, column=6, value=str(wo.job_status.status_choice))
            ws.cell(row=row_num, column=7, value=wo.plan_start_date.strftime('%Y-%m-%d') if wo.plan_start_date else "")
            ws.cell(row=row_num, column=8, value=float(wo.est_work_hours or 0))
            row_num += 1
        
        row_num += 1 

    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        
        for cell in col:
            if cell.value:
                val_len = len(str(cell.value))
                if val_len > max_length:
                    max_length = val_len
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Forecast_{selected_garage.Facility_Name}.xlsx'
    wb.save(response)
    return response

def shop_plan_view(request):
    week_num = request.GET.get("week")
    garage_id = request.GET.get("garage")
    
    # Standard selection logic
    selected_week = get_object_or_404(WorkWeek, week_number=week_num)
    selected_garage = get_object_or_404(Facility, id=garage_id)
    
    start = selected_week.start_date
    week_days = [start + timedelta(days=i) for i in range(7)]
    day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

    # 1. Get all Work Orders for this week/garage
    work_orders = WorkOrder.objects.filter(
        equipment__Garage=selected_garage,
        plan_start_date__range=(selected_week.start_date, selected_week.end_date),
        date_closed__isnull=True
    ).select_related('equipment')

    # 2. Identify unique equipment involved this week
    unique_equipment = sorted(
        list(set(wo.equipment for wo in work_orders)), 
        key=lambda e: e.Equipment_Number
    )

    # 3. Build the Grid Data
    # Structure: { equipment_obj: [ {day_index: { 'DS': hrs, 'NS': hrs } } ] }
    shop_data = []
    for equip in unique_equipment:
        equip_row = {'equipment': equip, 'days': []}
        
        for i in range(7):
            current_date = week_days[i]
            # Filter orders for this specific equipment and day
            day_orders = [wo for wo in work_orders if wo.equipment == equip and wo.plan_start_date.date() == current_date]
            
            # Logic for DS (Day Shift) vs NS (Night Shift)
            # Adjust 'hour' threshold based on your actual shift change time (e.g., 6 PM)
            ds_hours = sum(float(wo.est_work_hours or 0) for wo in day_orders if wo.plan_start_date.hour < 18)
            ns_hours = sum(float(wo.est_work_hours or 0) for wo in day_orders if wo.plan_start_date.hour >= 18)
            
            equip_row['days'].append({
                'ds': ds_hours if ds_hours > 0 else "",
                'ns': ns_hours if ns_hours > 0 else ""
            })
        
        shop_data.append(equip_row)

    context = {
        "selected_week": selected_week,
        "selected_garage": selected_garage,
        "day_names": day_names,
        "week_days": week_days,
        "shop_data": shop_data,
    }
    return render(request, "scheduling/shop_plan.html", context)

@require_POST
def commit_schedule(request):
    week_num_val = request.POST.get('week')
    garage_id = request.POST.get('garage')
    
    week_instance = get_object_or_404(
        WorkWeek, 
        week_number=week_num_val, 
        setup__active=True
    )

    garage_instance = get_object_or_404(Facility, pk=garage_id)

    schedule, created = Schedule.objects.get_or_create(
        week=week_instance,
        responsible_garage=garage_instance
    )

    with transaction.atomic():
        for key, value in request.POST.items():
            if key.startswith('date_') and value:
                wo_id = key.replace('date_', '')
                work_order = get_object_or_404(WorkOrder, id=wo_id)
                
                ScheduleSnapshot.objects.update_or_create(
                    schedule=schedule,
                    work_order=work_order,
                    defaults={
                        'plan_start_snapshot': value,
                        'estimated_hours_snapshot': work_order.est_work_hours or 0,
                        'job_status_snapshot': work_order.job_status.status_choice,
                        'date_closed_snapshot': work_order.date_closed
                    }
                )
        
    messages.success(request, f"Schedule for Week {week_num_val} successfully COMMITTED.")
    return redirect(f"/scheduling/?week={week_num_val}&garage={garage_id}")